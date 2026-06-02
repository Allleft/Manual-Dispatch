from io import BytesIO
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


FINAL_SUMMARY_HEADERS = [
    "No.",
    "Customer Name",
    "Suburb",
    "Estimated Distance From Warehouse (km)",
    "Invoice #",
    "Product Details",
    "Load",
]
OPSHOP_PICKUP_HEADERS = [
    "No.",
    "Category",
    "Route Group",
    "OP SHOP Name",
    "Suburb",
    "Address",
    "Pickup Date",
    "Run Type",
    "Frequency",
    "Time Window",
    "Contact",
    "Phone",
    "Access",
    "Key Required",
    "Trailer Restriction",
    "Notes",
]
INVALID_SHEET_CHARACTERS = re.compile(r"[\[\]\*:/\\?]")
EMPTY_ORDER_SUMMARY_MESSAGE = "No Delivery Orders included."


def build_final_summary_excel(final_summaries, dispatch_date, delivery_date=None):
    """Build an Excel workbook from saved Final Trip Summary snapshots."""
    workbook = Workbook()

    if not final_summaries:
        worksheet = workbook.active
        worksheet.title = "Final Summaries"
        empty_scope = (
            f"{dispatch_date} / {delivery_date}"
            if delivery_date
            else dispatch_date
        )
        worksheet["A1"] = f"No saved Final Trip Summaries for {empty_scope}"
        worksheet["A1"].font = Font(bold=True)
        worksheet.column_dimensions["A"].width = 48
        return _save_workbook(workbook)

    default_sheet = workbook.active
    workbook.remove(default_sheet)
    used_titles = set()

    for summary in final_summaries:
        worksheet = workbook.create_sheet(_safe_sheet_title(summary, used_titles))
        _write_summary_sheet(worksheet, summary)

    return _save_workbook(workbook)


def _write_summary_sheet(worksheet, summary):
    meta_rows = [
        ("Dispatch Date", summary.dispatch_date),
        ("Delivery Date", summary.delivery_date),
        ("Driver", summary.driver_name_snapshot),
        ("Rego #", summary.vehicle_rego_snapshot or "No vehicle selected"),
        ("Saved By", summary.saved_by_account_name or "Unknown"),
        ("Total Pallets", summary.total_pallets),
        ("Total Loose Bags", summary.total_loose_bags),
    ]

    for row_index, (label, value) in enumerate(meta_rows, start=1):
        worksheet.cell(row=row_index, column=1, value=label).font = Font(bold=True)
        worksheet.cell(row=row_index, column=2, value=value)

    row_index = len(meta_rows) + 2
    row_number = 1

    wrote_order_rows = False
    for trip in summary.trips:
        if not trip.orders:
            continue
        wrote_order_rows = True

        worksheet.cell(row=row_index, column=1, value=_trip_label(trip.trip_no)).font = Font(bold=True)
        row_index += 1

        for column_index, header in enumerate(FINAL_SUMMARY_HEADERS, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=header)
            cell.font = Font(bold=True)
        row_index += 1

        for order in trip.orders:
            values = [
                row_number,
                order.company_name_snapshot or "",
                order.suburb_snapshot or "",
                _format_estimated_distance(order),
                order.invoice_number_snapshot or "",
                _format_product_details(order),
                _format_load_quantity(order),
            ]
            for column_index, value in enumerate(values, start=1):
                cell = worksheet.cell(row=row_index, column=column_index, value=value)
                if column_index == 6:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            row_index += 1
            row_number += 1

        row_index += 1

    if not wrote_order_rows:
        worksheet.cell(row=row_index, column=1, value=EMPTY_ORDER_SUMMARY_MESSAGE).font = Font(bold=True)
        row_index += 2

    opshop_pickups = getattr(summary, "opshop_pickups", None) or []
    if opshop_pickups:
        worksheet.cell(row=row_index, column=1, value="OP SHOP PICKUPS").font = Font(bold=True)
        row_index += 1
        for column_index, header in enumerate(OPSHOP_PICKUP_HEADERS, start=1):
            worksheet.cell(row=row_index, column=column_index, value=header).font = Font(bold=True)
        row_index += 1

        for row_number, pickup in enumerate(opshop_pickups, start=1):
            values = [
                row_number,
                _format_opshop_pickup_category(pickup),
                pickup.route_group_name_snapshot or "",
                pickup.opshop_name_snapshot or "",
                pickup.suburb_snapshot or "",
                pickup.street_address_snapshot or "",
                pickup.pickup_date_snapshot or "",
                pickup.run_type_snapshot or "",
                pickup.pickup_frequency_snapshot or "",
                pickup.time_window_snapshot or "",
                pickup.primary_contact_snapshot or "",
                pickup.primary_phone_snapshot or "",
                pickup.access_type_snapshot or "",
                "Yes" if pickup.key_required_snapshot else "No",
                pickup.trailer_restriction_snapshot or "",
                pickup.notes_snapshot or "",
            ]
            for column_index, value in enumerate(values, start=1):
                cell = worksheet.cell(row=row_index, column=column_index, value=value)
                if column_index == 16:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            row_index += 1

    worksheet.freeze_panes = "A8"
    _apply_column_widths(worksheet)


def _safe_sheet_title(summary, used_titles):
    base = summary.driver_name_snapshot or summary.driver_id or "Summary"
    title = INVALID_SHEET_CHARACTERS.sub("-", base).strip() or "Summary"
    title = title[:31]
    candidate = title
    suffix = 2

    while candidate in used_titles:
        suffix_text = f" {suffix}"
        candidate = f"{title[: 31 - len(suffix_text)]}{suffix_text}"
        suffix += 1

    used_titles.add(candidate)
    return candidate


def _trip_label(trip_no):
    return "Trip 1" if trip_no == "trip1" else "Trip 2"


def _format_product_details(order):
    product_lines = getattr(order, "product_lines_snapshot", None) or []
    if not product_lines:
        return "No product details recorded."
    return "\n".join(
        f"{index}. {line.product_name} - {line.quantity} {_pluralized_unit(line.unit, line.quantity)}"
        for index, line in enumerate(product_lines, start=1)
    )


def _format_load_quantity(order):
    pallets = int(getattr(order, "pallet_quantity_snapshot", 0) or 0)
    loose_bags = int(getattr(order, "loose_bags_quantity_snapshot", 0) or 0)
    if pallets > 0:
        return f"{pallets} {_pluralized_unit('PALLETS', pallets)}"
    if loose_bags > 0:
        return f"{loose_bags} {_pluralized_unit('BAGS', loose_bags)}"
    return "-"


def _format_estimated_distance(order):
    distance = getattr(
        order,
        "estimated_distance_km_from_warehouse_snapshot",
        None,
    )
    if distance in ("", None):
        return "Unknown"
    return float(distance)


def _format_opshop_pickup_category(pickup):
    category = (getattr(pickup, "pickup_category_snapshot", None) or "").strip().upper()
    if category == "COUNTRYSIDE":
        return "Countryside"
    if category in {"ON_CALL", "ONCALL"}:
        return "Oncall"
    if category in {"REGULAR", "STANDARD"}:
        return "Regular"

    run_type = (getattr(pickup, "run_type_snapshot", None) or "").strip().upper()
    if run_type in {"REGULAR", "STANDARD"}:
        return "Regular"
    if run_type == "ON_CALL":
        return "Oncall"
    return category


def _pluralized_unit(unit, quantity):
    normalized = str(unit or "").upper()
    singular = "Pallet" if normalized == "PALLETS" else "Bag"
    return singular if quantity == 1 else f"{singular}s"


def _apply_column_widths(worksheet):
    for column_cells in worksheet.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        column_letter = column_cells[0].column_letter
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 34)


def _save_workbook(workbook):
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
