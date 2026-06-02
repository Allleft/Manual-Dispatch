from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


OPSHOP_RUN_SHEET_HEADERS = [
    "Driver / Assigned To",
    "Pickup Date",
    "Run Type",
    "Route Group",
    "OP SHOP Name",
    "Suburb",
    "Address",
    "Area / Region",
    "Primary Contact",
    "Primary Phone",
    "Secondary Contact",
    "Secondary Phone",
    "Time Window",
    "Call Before Arrival",
    "Access Type",
    "Key Required",
    "Trailer Restriction",
    "Notes",
    "Status",
]

REGULAR_SECTION_TITLE = "REGULAR OP SHOP PICKUPS"
ONCALL_SECTION_TITLE = "ONCALL OP SHOP PICKUPS"
COUNTRYSIDE_SECTION_TITLE = "COUNTRYSIDE OP SHOP PICKUPS"
UNKNOWN_ROUTE_GROUP_TITLE = "Unknown Route Group"
UNASSIGNED_DRIVER_LABEL = "Unassigned"
UNASSIGNED_SECTION_TITLE = "Unassigned OP SHOP Pickups"


def build_opshop_pickup_run_sheet_excel(board_data, dispatch_date):
    """Build an independent OP SHOP pickup run sheet workbook."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "OP SHOP Run Sheet"

    worksheet["A1"] = "OP SHOP Pickup Run Sheet"
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet["A2"] = "Dispatch Date"
    worksheet["A2"].font = Font(bold=True)
    worksheet["B2"] = dispatch_date

    row_index = 4
    regular_pickups, oncall_pickups, countryside_groups = _build_sections(board_data)
    row_index = _write_pickup_section(
        worksheet,
        row_index,
        REGULAR_SECTION_TITLE,
        regular_pickups,
        "No Regular OP SHOP pickups.",
        board_data,
    )
    row_index = _write_pickup_section(
        worksheet,
        row_index,
        ONCALL_SECTION_TITLE,
        oncall_pickups,
        "No Oncall OP SHOP pickups.",
        board_data,
    )
    row_index = _write_countryside_section(
        worksheet,
        row_index,
        countryside_groups,
        board_data,
    )

    worksheet.freeze_panes = "A5"
    _apply_column_widths(worksheet)
    return _save_workbook(workbook)


def _build_sections(board_data):
    pickups_by_id = {}

    for pickup in list(getattr(board_data, "scheduled_opshop_pickups", []) or []):
        pickups_by_id[pickup.pickup_task_id] = pickup
    for pickup in list(getattr(board_data, "oncall_opshop_pickups", []) or []):
        pickups_by_id[pickup.pickup_task_id] = pickup
    for pickup in list(getattr(board_data, "countryside_opshop_pickups", []) or []):
        pickups_by_id[pickup.pickup_task_id] = pickup
    for pickup in list(getattr(board_data, "assigned_opshop_pickups", []) or []):
        pickups_by_id[pickup.pickup_task_id] = pickup

    regular_pickups = []
    oncall_pickups = []
    countryside_pickups = []
    for pickup in pickups_by_id.values():
        pickup_category = (pickup.pickup_category or "NORMAL").upper()
        run_type = (pickup.run_type or "").upper()
        if pickup_category == "COUNTRYSIDE":
            countryside_pickups.append(pickup)
        elif run_type in {"REGULAR", "STANDARD"}:
            regular_pickups.append(pickup)
        elif run_type == "ON_CALL":
            oncall_pickups.append(pickup)

    countryside_groups = {}
    for pickup in countryside_pickups:
        route_group_name = _route_group_name(pickup)
        countryside_groups.setdefault(route_group_name, []).append(pickup)

    return (
        _sort_pickups(regular_pickups, board_data),
        _sort_pickups(oncall_pickups, board_data),
        {
            route_group_name: _sort_pickups(pickups, board_data)
            for route_group_name, pickups in sorted(
                countryside_groups.items(),
                key=lambda item: item[0].lower(),
            )
        },
    )


def _write_pickup_section(worksheet, row_index, section_title, pickups, empty_message, board_data):
    worksheet.cell(row=row_index, column=1, value=section_title).font = Font(bold=True)
    row_index += 1
    if not pickups:
        worksheet.cell(row=row_index, column=1, value=empty_message)
        row_index += 2
        return row_index

    row_index = _write_header_row(worksheet, row_index)
    for pickup in pickups:
        row_index = _write_pickup_row(worksheet, row_index, pickup, board_data)
    return row_index + 1


def _write_countryside_section(worksheet, row_index, countryside_groups, board_data):
    worksheet.cell(row=row_index, column=1, value=COUNTRYSIDE_SECTION_TITLE).font = Font(bold=True)
    row_index += 1
    if not countryside_groups:
        worksheet.cell(row=row_index, column=1, value="No Countryside OP SHOP pickups.")
        row_index += 2
        return row_index

    for route_group_name, pickups in countryside_groups.items():
        worksheet.cell(row=row_index, column=1, value=f"Route Group: {route_group_name}").font = Font(bold=True)
        row_index += 1
        row_index = _write_header_row(worksheet, row_index)
        for pickup in pickups:
            row_index = _write_pickup_row(worksheet, row_index, pickup, board_data)
        row_index += 1
    return row_index


def _write_header_row(worksheet, row_index):
    for column_index, header in enumerate(OPSHOP_RUN_SHEET_HEADERS, start=1):
        cell = worksheet.cell(row=row_index, column=column_index, value=header)
        cell.font = Font(bold=True)
    return row_index + 1


def _write_pickup_row(worksheet, row_index, pickup, board_data):
    for column_index, value in enumerate(_pickup_row(pickup, board_data), start=1):
        cell = worksheet.cell(row=row_index, column=column_index, value=value)
        if column_index == 18:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    return row_index + 1


def _sort_pickups(pickups, board_data):
    return sorted(
        pickups,
        key=lambda pickup: (
            pickup.pickup_date or "",
            _driver_sort_key(pickup, board_data),
            pickup.suburb or "",
            pickup.opshop_name or "",
            pickup.pickup_task_id or "",
        ),
    )


def _pickup_row(pickup, board_data):
    return [
        _assigned_driver_name(pickup, board_data),
        _clean_cell(pickup.pickup_date),
        _clean_cell(pickup.run_type),
        _clean_cell(_route_group_name(pickup, blank_for_normal=True)),
        _clean_cell(pickup.opshop_name),
        _clean_cell(pickup.suburb),
        _clean_cell(pickup.street_address),
        _clean_cell(pickup.area_region),
        _clean_cell(pickup.primary_contact),
        _clean_cell(pickup.primary_phone),
        _clean_cell(pickup.secondary_contact),
        _clean_cell(pickup.secondary_phone),
        _clean_cell(pickup.time_window),
        _format_bool(pickup.call_before_arrival),
        _clean_cell(pickup.access_type),
        _format_bool(pickup.key_required),
        _clean_cell(pickup.trailer_restriction),
        _format_notes(pickup),
        _clean_cell(pickup.status),
    ]


def _assigned_driver_name(pickup, board_data):
    drivers_by_id = {driver.driver_id: driver for driver in getattr(board_data, "drivers", [])}
    driver_id = pickup.driver_id or pickup.assigned_driver_id
    if not driver_id:
        return UNASSIGNED_DRIVER_LABEL
    return _clean_cell(
        pickup.assigned_driver_name
        or getattr(drivers_by_id.get(driver_id), "name", None)
        or driver_id
    )


def _driver_sort_key(pickup, board_data):
    driver_name = _assigned_driver_name(pickup, board_data)
    if driver_name == UNASSIGNED_DRIVER_LABEL:
        return (1, "")
    return (0, driver_name.lower())


def _route_group_name(pickup, blank_for_normal=False):
    if (pickup.pickup_category or "NORMAL").upper() != "COUNTRYSIDE":
        return "" if blank_for_normal else UNKNOWN_ROUTE_GROUP_TITLE
    return _clean_cell(pickup.route_group_name) or UNKNOWN_ROUTE_GROUP_TITLE


def _clean_cell(value):
    if value is None:
        return ""
    return str(value)


def _format_bool(value):
    return "Yes" if bool(value) else "No"


def _format_notes(pickup):
    notes = []
    if pickup.status_notes:
        notes.append(f"Status: {_clean_cell(pickup.status_notes)}")
    if pickup.task_notes:
        notes.append(f"Task: {_clean_cell(pickup.task_notes)}")
    return "\n".join(notes)


def _apply_column_widths(worksheet):
    for column_cells in worksheet.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        column_letter = column_cells[0].column_letter
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 34)


def _save_workbook(workbook):
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
