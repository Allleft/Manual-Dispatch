from io import BytesIO
from datetime import datetime
import math
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins


COLLECTION_HEADERS = [
    "OPSHOP NAME",
    "Suburb",
    "CLOTHING KG",
    "SHOES KG",
    "TIME IN",
    "TIME OUT",
    "TROLLEYS OUT TO OPSHOPS",
    "TROLLEYS IN TO MCC ",
    "HARD TOYS",
    "SOFT TOYS",
    "BLACK BAGS",
    "SHOE BAGS",
]
MIN_PICKUP_ROWS = 11
MIN_PICKUP_ROW_HEIGHT = 25.5
PICKUP_ROW_LINE_HEIGHT = 15
MAX_PICKUP_ROW_HEIGHT = 100.5
COLLECTION_COLUMN_WIDTHS = {
    "A": 30.71,
    "B": 14.14,
    "C": 12.86,
    "D": 12,
    "E": 7.71,
    "F": 7.57,
    "G": 8.43,
    "H": 8.29,
    "I": 9.57,
    "J": 13,
    "K": 10.71,
    "L": 9,
}
_INVALID_SHEET_NAME_CHARACTERS = re.compile(r"[\\/*?:\[\]]")


def build_opshop_pickup_collection_excel(collection):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "WEIGHT SHEET"
    _configure_worksheet(worksheet)
    _write_form_header(worksheet, collection)
    _write_pickup_table(worksheet, collection)
    return _save(workbook)


def build_opshop_pickup_collections_excel(collections, pickup_date):
    if not collections:
        raise ValueError(
            "No OP SHOP Pickup Collections available for this pickup date."
        )

    workbook = Workbook()
    workbook.remove(workbook.active)
    used_sheet_names = set()
    for collection in sorted(
        collections,
        key=lambda item: (
            str(item.driver_name_snapshot or item.driver_id or ""),
            str(item.collection_id or ""),
        ),
    ):
        worksheet = workbook.create_sheet(
            _unique_sheet_name(collection.driver_name_snapshot, used_sheet_names)
        )
        _configure_worksheet(worksheet)
        _write_form_header(worksheet, collection, pickup_date=pickup_date)
        _write_pickup_table(worksheet, collection)
    return _save(workbook)


def _configure_worksheet(worksheet):
    worksheet.sheet_view.showGridLines = False
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.scale = None
    worksheet.page_margins = PageMargins(
        left=0.35,
        right=0.35,
        top=0.4,
        bottom=0.4,
        header=0.2,
        footer=0.2,
    )
    for column_letter, width in COLLECTION_COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column_letter].width = width


def _write_form_header(worksheet, collection, pickup_date=None):
    for range_ref in (
        "A1:L1",
        "A2:L2",
        "A3:L3",
        "A4:L4",
        "A5:L5",
        "A8:L8",
    ):
        worksheet.merge_cells(range_ref)

    worksheet["A1"] = "DAILY OP SHOP COLLECTIONS - WEIGHT SHEET"
    worksheet["A1"].font = Font(bold=True, size=16)
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["A3"] = (
        "REMINDER : NO BOARD GAMES/ PUZZLES  - can not send overseas, "
        "don't understand english"
    )
    worksheet["A4"] = "**Please ensure  HARD & SOFT TOYS are in separate bags**"
    worksheet["A5"] = (
        f"DRIVER NAME: {collection.driver_name_snapshot or ''}    "
        f"PICK UP DATE: {_display_date(pickup_date or collection.pickup_date)}    "
        f"DAY: {_display_day(pickup_date or collection.pickup_date)}"
    )
    worksheet["A6"] = "REGO # ________________________"
    worksheet["A8"] = "PLEASE RECORD WEIGHT OF BAGS FOR EACH OP SHOP "
    for coordinate in ("A3", "A4", "A6", "A8"):
        worksheet[coordinate].font = Font(bold=True, size=11)
        worksheet[coordinate].alignment = Alignment(vertical="center", wrap_text=True)
    worksheet["A5"].font = Font(bold=True, size=11)
    worksheet["A5"].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=False,
        shrink_to_fit=True,
    )
    worksheet["A8"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["A3"].fill = PatternFill("solid", fgColor="1F1F1F")
    worksheet["A3"].font = Font(bold=True, size=11, color="FFFFFF")
    worksheet["A4"].fill = PatternFill("solid", fgColor="FCE4D6")

    worksheet.row_dimensions[1].height = 26.25
    worksheet.row_dimensions[2].height = 26.25
    worksheet.row_dimensions[3].height = 27
    worksheet.row_dimensions[4].height = 23.25
    worksheet.row_dimensions[5].height = 21
    worksheet.row_dimensions[6].height = 21
    worksheet.row_dimensions[7].height = 18.75
    worksheet.row_dimensions[8].height = 27.75


def _write_pickup_table(worksheet, collection):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="E2F0D9")
    header_row = 10
    for column_index, header in enumerate(COLLECTION_HEADERS, start=1):
        cell = worksheet.cell(row=header_row, column=column_index, value=header)
        cell.font = Font(bold=True, size=9)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[header_row].height = 36.75

    row_index = 12
    for pickup in collection.pickups:
        _write_pickup_row(worksheet, row_index, pickup, border)
        row_index += 1
    for _ in range(max(MIN_PICKUP_ROWS - len(collection.pickups), 0)):
        _write_empty_pickup_row(worksheet, row_index, border)
        row_index += 1

    worksheet.freeze_panes = "A12"
    worksheet.print_area = f"A1:L{row_index - 1}"


def _write_pickup_row(worksheet, row_index, pickup, border):
    values = [
        pickup.opshop_name_snapshot or "",
        pickup.suburb_snapshot or "",
        pickup.clothing_kg_snapshot,
        pickup.shoes_kg_snapshot,
        pickup.time_in_snapshot,
        pickup.time_out_snapshot,
        pickup.trolleys_out_to_opshops_snapshot,
        pickup.trolleys_in_to_mcc_snapshot,
        pickup.hard_toys_snapshot,
        pickup.soft_toys_snapshot,
        pickup.black_bags_snapshot,
        pickup.shoe_bags_snapshot,
    ]
    for column_index, value in enumerate(values, start=1):
        cell = worksheet.cell(row=row_index, column=column_index, value=value)
        cell.border = border
        if column_index in {3, 4}:
            cell.number_format = '0.## "KG"'
        cell.alignment = Alignment(
            horizontal="center" if column_index >= 3 else "left",
            vertical="center",
            wrap_text=column_index in {1, 2},
        )
    wrapped_lines = max(
        _wrapped_line_count(values[0], COLLECTION_COLUMN_WIDTHS["A"]),
        _wrapped_line_count(values[1], COLLECTION_COLUMN_WIDTHS["B"]),
    )
    worksheet.row_dimensions[row_index].height = min(
        MAX_PICKUP_ROW_HEIGHT,
        MIN_PICKUP_ROW_HEIGHT
        + PICKUP_ROW_LINE_HEIGHT * max(wrapped_lines - 1, 0),
    )


def _write_empty_pickup_row(worksheet, row_index, border):
    for column_index in range(1, len(COLLECTION_HEADERS) + 1):
        cell = worksheet.cell(row=row_index, column=column_index, value=None)
        cell.border = border
        if column_index in {3, 4}:
            cell.number_format = '0.## "KG"'
        cell.alignment = Alignment(horizontal="center" if column_index >= 3 else "left", vertical="center")
    worksheet.row_dimensions[row_index].height = MIN_PICKUP_ROW_HEIGHT


def _wrapped_line_count(value, column_width):
    usable_width = max(1, int(float(column_width)))
    return sum(
        max(1, math.ceil(len(line) / usable_width))
        for line in str(value or "").split("\n")
    )


def _display_date(value):
    try:
        return datetime.strptime(value, "%Y-%m-%d").strftime("%d/%m/%Y")
    except (TypeError, ValueError):
        return str(value or "")


def _display_day(value):
    try:
        return datetime.strptime(value, "%Y-%m-%d").strftime("%A").upper()
    except (TypeError, ValueError):
        return ""


def _unique_sheet_name(driver_name, used_sheet_names):
    base_name = _INVALID_SHEET_NAME_CHARACTERS.sub(" ", str(driver_name or "Driver"))
    base_name = " ".join(base_name.split()).strip("'") or "Driver"
    base_name = base_name[:31]
    candidate = base_name
    sequence = 2
    while candidate.casefold() in used_sheet_names:
        suffix = f" {sequence}"
        candidate = f"{base_name[:31 - len(suffix)]}{suffix}"
        sequence += 1
    used_sheet_names.add(candidate.casefold())
    return candidate


def _save(workbook):
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
