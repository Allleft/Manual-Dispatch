from io import BytesIO

from datetime import datetime
import math
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins


DAILY_RUN_SHEET_HEADERS = [
    "Sequence",
    "Customer Name",
    "Suburb",
    "Invoice #",
    "PRODUCT",
    "KG'S",
    "Pallets",
    "COD",
    "CQ",
    "Time In",
    "Time Out",
    "PRINT NAME",
    "SIGNATURE",
    "RETND",
]
MIN_TABLE_ROWS = 18
BODY_FONT_SIZE = 8.5
PRODUCT_FONT_SIZE = 7.5
MIN_BODY_ROW_HEIGHT = 21.95
BODY_LINE_HEIGHT = 10.5
PRODUCT_LINE_HEIGHT = 10.5
BODY_ROW_PADDING = 2
MAX_BODY_ROW_HEIGHT = 60
COLUMN_WIDTHS = {
    "A": 3.855,
    "B": 27.57,
    "C": 15,
    "D": 12,
    "E": 16,
    "F": 5.57,
    "G": 6.855,
    "H": 5.855,
    "I": 3,
    "J": 8,
    "K": 8,
    "L": 12.57,
    "M": 17.285,
    "N": 7.71,
}
_INVALID_SHEET_NAME_CHARACTERS = re.compile(r"[\\/*?:\[\]]")
_PRODUCT_CODE_KG_SUFFIX = re.compile(r"([0-9]+(?:\.[0-9]+)?)KG$", re.IGNORECASE)
_PACKAGE_UNIT_BAG_WEIGHT = re.compile(
    r"BAG\s*([0-9]+(?:\.[0-9]+)?)",
    re.IGNORECASE,
)


def build_delivery_run_sheet_excel(run_sheet):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Daily Run Sheet"
    _configure_print_layout(worksheet)
    _write_daily_run_sheet_form(worksheet, run_sheet, run_sheet.delivery_date)
    return _save(workbook)


def build_delivery_run_sheets_excel(run_sheets, delivery_date):
    if not run_sheets:
        raise ValueError(
            "No Generated or Saved Delivery Run Sheets are available for this Delivery Date."
        )

    workbook = Workbook()
    workbook.remove(workbook.active)
    used_sheet_names = set()
    for run_sheet in run_sheets:
        worksheet = workbook.create_sheet(
            _unique_sheet_name(run_sheet.driver_name_snapshot, used_sheet_names)
        )
        _configure_print_layout(worksheet)
        _write_daily_run_sheet_form(worksheet, run_sheet, delivery_date)
    return _save(workbook)


def _configure_print_layout(worksheet):
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins = PageMargins(
        left=0.118,
        right=0.197,
        top=0.394,
        bottom=0.157,
        header=0.315,
        footer=0.315,
    )


def _write_daily_run_sheet_form(worksheet, run_sheet, delivery_date):
    worksheet.sheet_view.showGridLines = False
    worksheet.print_options.horizontalCentered = False
    worksheet.print_title_rows = "7:8"

    worksheet["A1"] = "DAILY RUN SHEET"
    worksheet["C1"] = f"DATE  {_display_date(delivery_date)}"
    worksheet["F1"] = f"DRIVER: {run_sheet.driver_name_snapshot or ''}"
    worksheet["L1"] = f"REGO #: {_rego_snapshot_display(run_sheet)}"
    for coordinate, font_size in (("A1", 16), ("C1", 11), ("F1", 11), ("L1", 14)):
        worksheet[coordinate].font = Font(bold=True, size=font_size)
        worksheet[coordinate].alignment = Alignment(
            vertical="center",
            horizontal="center" if coordinate == "L1" else "left",
            wrap_text=False,
            shrink_to_fit=False,
        )
    highlight_fill = PatternFill(fill_type="solid", fgColor="C6E0B4")
    worksheet["F1"].fill = highlight_fill
    worksheet["L1"].fill = highlight_fill

    worksheet["B3"] = "START TIME: ____________________________________"
    worksheet.merge_cells("B5:E5")
    worksheet["B5"] = "TIME LOADING STARTED(TO BE FILLED IN BY STOREMAN)___________"
    worksheet.merge_cells("F5:N5")
    worksheet["F5"] = "TIME LOADING COMPLETED(TO BE FILLED IN BY STOREMAN)_____________"
    for coordinate, font_size in (("B3", 16), ("B5", 12), ("F5", 12)):
        worksheet[coordinate].font = Font(bold=True, size=font_size)
        worksheet[coordinate].alignment = Alignment(vertical="center", wrap_text=False)

    worksheet["N6"] = "NO. #"
    worksheet["J7"] = "Time"
    worksheet["K7"] = "Time"
    worksheet["M7"] = "Comments"
    worksheet["N7"] = "PALLETS"
    for coordinate in ("J7", "K7", "N6", "N7"):
        worksheet[coordinate].font = Font(bold=True, size=11)
        worksheet[coordinate].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["M7"].font = Font(size=11)
    worksheet["M7"].alignment = Alignment(horizontal="center", vertical="center")
    returned_pallets_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    for coordinate in ("N6", "N7"):
        worksheet[coordinate].fill = returned_pallets_fill

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_row = 8
    for column_index, header in enumerate(DAILY_RUN_SHEET_HEADERS, start=1):
        cell = worksheet.cell(row=header_row, column=column_index, value=header)
        cell.font = Font(bold=column_index != 5, size=9)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet["N8"].fill = returned_pallets_fill
    worksheet.row_dimensions[header_row].height = 21

    snapshot_rows = _ordered_snapshot_rows(run_sheet)
    row_index = header_row + 1
    for row_no, order in enumerate(snapshot_rows, start=1):
        _write_order_row(worksheet, row_index, row_no, order, border)
        row_index += 1
    for row_no in range(len(snapshot_rows) + 1, MIN_TABLE_ROWS + 1):
        _write_empty_row(worksheet, row_index, row_no, border)
        row_index += 1

    finish = worksheet.cell(row=row_index + 1, column=2, value="FINISH TIME:____________________________________")
    finish.font = Font(bold=True, size=16)
    finish.alignment = Alignment(vertical="center")
    worksheet.row_dimensions[row_index + 1].height = 21
    worksheet.print_area = f"A1:N{row_index + 1}"
    _apply_column_widths(worksheet)
    _apply_row_heights(worksheet)
    worksheet.page_setup.fitToHeight = 1 if len(snapshot_rows) <= MIN_TABLE_ROWS else 0


def _ordered_snapshot_rows(run_sheet):
    return [
        order
        for trip in (run_sheet.trips or [])
        for order in (trip.orders or [])
    ]


def delivery_run_sheet_product_display(order):
    displays = []
    for line in getattr(order, "product_lines_snapshot", None) or []:
        product_name = str(getattr(line, "product_name", "") or "").strip()
        product_code = str(getattr(line, "product_code", "") or "").strip()
        quantity = getattr(line, "quantity", None)
        unit = str(getattr(line, "unit", "") or "").strip()
        if not product_name and not product_code:
            continue
        identity = product_code or product_name
        quantity_text = " ".join(
            part for part in (_display_number(quantity), unit) if part
        )
        package_quantity = getattr(line, "package_quantity", None)
        package_unit = str(getattr(line, "package_unit", "") or "").strip()
        package_text = (
            f"{_display_number(package_quantity)} {package_unit}"
            if package_quantity is not None and package_unit
            else ""
        )
        detail = package_text or quantity_text
        display = f"{identity} - {detail}" if detail else identity
        displays.append(display)
    if displays:
        return "\n".join(displays)
    return str(getattr(order, "product_snapshot", "") or "").strip()


def _delivery_run_sheet_line_weight_kg(line):
    unit = str(getattr(line, "unit", "") or "").strip().upper()
    if unit in {"KG", "KGS"}:
        return _numeric_value(getattr(line, "quantity", None))

    package_quantity = _numeric_value(getattr(line, "package_quantity", None))
    if package_quantity is None or package_quantity <= 0:
        return None

    product_code = str(getattr(line, "product_code", "") or "").strip()
    code_match = _PRODUCT_CODE_KG_SUFFIX.search(product_code)
    if code_match:
        return float(code_match.group(1)) * package_quantity

    package_unit = str(getattr(line, "package_unit", "") or "").strip()
    package_match = _PACKAGE_UNIT_BAG_WEIGHT.fullmatch(package_unit)
    if package_match:
        weight_per_bag = float(package_match.group(1))
        return weight_per_bag * package_quantity if weight_per_bag > 0 else None
    return None


def _delivery_run_sheet_weight_total(order):
    total = sum(
        weight
        for line in getattr(order, "product_lines_snapshot", None) or []
        if (weight := _delivery_run_sheet_line_weight_kg(line)) is not None
    )
    return "" if total == 0 else int(total) if total.is_integer() else total


def _numeric_value(value):
    if value is None or value == "":
        return None
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    return numeric if math.isfinite(numeric) else None


def _display_number(value):
    if value is None or value == "":
        return ""
    numeric = float(value)
    return str(int(numeric)) if numeric.is_integer() else str(numeric)


def _wrapped_line_count(value, column_width, font_size=BODY_FONT_SIZE):
    usable_width = max(1, int(float(column_width) * 11 / font_size))
    return sum(
        max(1, math.ceil(len(line) / usable_width))
        for line in str(value or "").split("\n")
    )


def _unique_sheet_name(driver_name, used_sheet_names):
    base_name = _INVALID_SHEET_NAME_CHARACTERS.sub(" ", str(driver_name or "Driver"))
    base_name = " ".join(base_name.split()).strip("'") or "Driver"
    base_name = base_name[:31]
    candidate = base_name
    sequence = 2
    while candidate.casefold() in used_sheet_names:
        suffix = f" ({sequence})"
        candidate = f"{base_name[:31 - len(suffix)]}{suffix}"
        sequence += 1
    used_sheet_names.add(candidate.casefold())
    return candidate


def _display_date(value):
    return datetime.strptime(value, "%Y-%m-%d").strftime("%d/%m/%Y")


def _write_order_row(worksheet, row_index, row_no, order, border):
    product_display = delivery_run_sheet_product_display(order)
    values = [
        row_no,
        order.company_name_snapshot or "",
        order.suburb_snapshot or "",
        order.invoice_number_snapshot or "",
        product_display,
        _delivery_run_sheet_weight_total(order),
        _number_or_blank(order.pallet_quantity_snapshot),
        "",
        "",
        "",
        "",
        "",
        "",
        "",
    ]
    for column_index, value in enumerate(values, start=1):
        cell = worksheet.cell(row=row_index, column=column_index, value=value)
        cell.font = Font(
            bold=column_index == 5,
            size=PRODUCT_FONT_SIZE if column_index == 5 else BODY_FONT_SIZE,
        )
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center"
            if column_index in {1, 4, 6, 7, 8, 9, 10, 11, 14}
            else "left",
            vertical="top",
            wrap_text=column_index in {2, 3, 4, 5, 13},
            shrink_to_fit=False,
        )
        if column_index in {6, 7, 14} and value != "":
            cell.number_format = "General"
    body_lines = max(
        _wrapped_line_count(order.company_name_snapshot, COLUMN_WIDTHS["B"]),
        _wrapped_line_count(order.suburb_snapshot, COLUMN_WIDTHS["C"]),
        _wrapped_line_count(order.invoice_number_snapshot, COLUMN_WIDTHS["D"]),
    )
    product_lines = _wrapped_line_count(
        product_display,
        COLUMN_WIDTHS["E"],
        PRODUCT_FONT_SIZE,
    )
    worksheet.row_dimensions[row_index].height = min(
        MAX_BODY_ROW_HEIGHT,
        max(
            MIN_BODY_ROW_HEIGHT,
            BODY_LINE_HEIGHT * body_lines + BODY_ROW_PADDING,
            PRODUCT_LINE_HEIGHT * product_lines + BODY_ROW_PADDING,
        ),
    )


def _write_empty_row(worksheet, row_index, row_no, border):
    for column_index in range(1, 15):
        cell = worksheet.cell(row=row_index, column=column_index, value=row_no if column_index == 1 else "")
        cell.font = Font(size=BODY_FONT_SIZE)
        cell.border = border
        cell.alignment = Alignment(horizontal="center" if column_index == 1 else "left", vertical="top")
    worksheet.row_dimensions[row_index].height = MIN_BODY_ROW_HEIGHT


def _number_or_blank(value):
    return "" if value is None else value


def _rego_snapshot_display(run_sheet):
    rego = str(getattr(run_sheet, "vehicle_rego_snapshot", "") or "").strip()
    return rego or "Not selected"


def _apply_column_widths(worksheet):
    for column_letter, width in COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column_letter].width = width
    worksheet.freeze_panes = "B9"


def _apply_row_heights(worksheet):
    row_heights = {
        1: 21,
        2: 21,
        3: 21,
        4: 15,
        5: 15.75,
        6: 15.75,
        7: 15,
        8: 21,
    }
    for row_index, height in row_heights.items():
        worksheet.row_dimensions[row_index].height = height


def _save(workbook):
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
