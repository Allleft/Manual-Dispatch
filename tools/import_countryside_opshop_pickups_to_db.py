"""Import Countryside OP SHOP route groups into SQLite source tables.

The Countryside workbook is a template source only. Each sheet name becomes a
route group, and each active row creates or updates an ON_CALL + COUNTRYSIDE
schedule membership for that route group. The importer does not create actual
pickup tasks.
"""

from __future__ import annotations

import argparse
import hashlib
import os
import sqlite3
import sys
from collections import Counter
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from backend.repositories.sqlite_manual_dispatch_repository import (  # noqa: E402
    SQLiteManualDispatchRepository,
)
from backend.schemas import (  # noqa: E402
    OpShopCountrysideRouteGroup,
    OpShopLocation,
    OpShopPickupSchedule,
)
from tools.maintenance_logbook import (  # noqa: E402
    add_maintenance_logbook_arguments,
    record_maintenance_event,
    resolve_maintenance_actor,
    safe_basename,
    sanitized_failure_metadata,
    workbook_import_failure_phase,
    workbook_import_metadata,
)


DEFAULT_DB_PATH = PROJECT_ROOT / "data" / "manual_dispatch.sqlite3"
REQUIRED_COLUMNS = [
    "Op_Shop_Name",
    "Run_Day",
    "Run_Type",
    "Active_Flag",
    "Suburb",
    "Street_Address",
    "Area_Region",
    "Primary_Contact",
    "Primary_Phone",
    "Secondary_Contact",
    "Secondary_Phone",
    "Pickup_Frequency",
    "Time_Window",
    "Call_Before_Arrival",
    "Call_Timing",
    "Access_Type",
    "Key_Required",
    "Trailer_Restriction",
    "Status",
    "Status_Start_Date",
    "Status_Notes",
    "Assigned to",
]
DRIVER_ALIAS_TO_NAME = {
    "john g": "John Georgiadis",
    "gavin": "Gavin Fynn",
    "nonda": "Epaminondas Tsatsoulis",
    "lee": "Guanlin Li",
}
WORKBOOK_IMPORT_COUNTRYSIDE = "WORKBOOK_IMPORT_COUNTRYSIDE"
VALID_RUN_DAYS = {"MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY"}


@dataclass
class CountrysideOpShopImportSummary:
    sheets_read: int = 0
    rows_read: int = 0
    rows_imported: int = 0
    route_groups_inserted: int = 0
    route_groups_updated: int = 0
    route_groups_deactivated: int = 0
    locations_inserted: int = 0
    locations_updated: int = 0
    schedules_inserted: int = 0
    schedules_updated: int = 0
    schedules_deactivated: int = 0
    duplicate_locations_reused: int = 0
    unresolved_assigned_to: dict[str, int] = field(default_factory=dict)
    backup_path: str | None = None


@dataclass
class PreparedCountrysideRow:
    location_key: str
    route_group: OpShopCountrysideRouteGroup
    location: OpShopLocation
    schedule: OpShopPickupSchedule


def import_countryside_opshop_pickups_to_db(file_path, db_path=None):
    workbook_path = Path(file_path)
    target_db_path = resolve_db_path(db_path)
    rows = read_countryside_workbook_rows(workbook_path)
    backup_path = backup_database_if_exists(target_db_path)
    repository = SQLiteManualDispatchRepository(target_db_path)
    driver_lookup = build_driver_lookup(repository)
    prepared_rows, unresolved = prepare_countryside_rows(
        rows,
        driver_lookup,
    )

    summary = CountrysideOpShopImportSummary(
        sheets_read=len({row["__sheet_name"] for row in rows}),
        rows_read=len(rows),
        rows_imported=len(prepared_rows),
        unresolved_assigned_to=dict(unresolved),
        backup_path=str(backup_path) if backup_path else None,
    )
    imported_route_group_ids = set()
    imported_locations = set()
    imported_schedule_ids = set()
    processed_route_groups = set()
    route_group_ids_by_key = {}

    with sqlite3.connect(target_db_path) as connection:
        connection.row_factory = sqlite3.Row
        for prepared in prepared_rows:
            group_key = normalize_key(prepared.route_group.route_group_name)
            if group_key not in processed_route_groups:
                existing_group = find_route_group_by_name(
                    connection,
                    prepared.route_group.route_group_name,
                )
                if existing_group:
                    prepared.route_group.route_group_id = existing_group["route_group_id"]
                    prepared.route_group.created_at = existing_group["created_at"]
                    summary.route_groups_updated += 1
                else:
                    summary.route_groups_inserted += 1
                repository.upsert_countryside_route_group(prepared.route_group)
                processed_route_groups.add(group_key)
                route_group_ids_by_key[group_key] = prepared.route_group.route_group_id
            else:
                prepared.route_group.route_group_id = route_group_ids_by_key[group_key]
            imported_route_group_ids.add(prepared.route_group.route_group_id)

            if prepared.location_key in imported_locations:
                summary.duplicate_locations_reused += 1
            else:
                existing_location_id = find_location_id_by_key(
                    connection,
                    prepared.location_key,
                )
                if existing_location_id:
                    prepared.location.opshop_id = existing_location_id
                    summary.locations_updated += 1
                else:
                    summary.locations_inserted += 1
                repository.upsert_opshop_location(prepared.location)
                imported_locations.add(prepared.location_key)

            prepared.schedule.opshop_id = prepared.location.opshop_id
            prepared.schedule.route_group_id = prepared.route_group.route_group_id
            prepared.schedule.schedule_id = deterministic_id(
                "OPSHOP-SCHEDULE",
                schedule_key(prepared.schedule),
            )
            existing_schedule_id = find_schedule_id_by_key(
                connection,
                schedule_key(prepared.schedule),
            )
            if existing_schedule_id:
                prepared.schedule.schedule_id = existing_schedule_id
                summary.schedules_updated += 1
            else:
                summary.schedules_inserted += 1
            repository.upsert_opshop_pickup_schedule(prepared.schedule)
            imported_schedule_ids.add(prepared.schedule.schedule_id)

        summary.schedules_deactivated = deactivate_missing_countryside_schedules(
            connection,
            imported_schedule_ids,
        )
        summary.route_groups_deactivated = deactivate_missing_countryside_route_groups(
            connection,
            imported_route_group_ids,
        )

    return summary


def resolve_db_path(db_path=None):
    if db_path:
        return Path(db_path)
    configured = os.environ.get("MANUAL_DISPATCH_DB_PATH")
    if configured:
        return Path(configured)
    return DEFAULT_DB_PATH


def backup_database_if_exists(db_path):
    db_path = Path(db_path)
    if not db_path.exists():
        return None

    backup_dir = db_path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp_value = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = (
        backup_dir
        / f"manual_dispatch_before_countryside_opshop_import_{timestamp_value}.sqlite3"
    )

    source = sqlite3.connect(db_path)
    try:
        destination = sqlite3.connect(backup_path)
        try:
            source.backup(destination)
        finally:
            destination.close()
    finally:
        source.close()
    return backup_path


def read_countryside_workbook_rows(workbook_path):
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    rows = []
    for worksheet in workbook.worksheets:
        header_map = _find_header_map(worksheet)
        for row in worksheet.iter_rows(
            min_row=header_map["__header_row"] + 1,
            values_only=False,
        ):
            record = {
                column: cell_text(row[header_map[column]])
                for column in REQUIRED_COLUMNS
            }
            if any(record.values()):
                record["__sheet_name"] = worksheet.title
                rows.append(record)
    return rows


def _find_header_map(worksheet):
    for row_number, row in enumerate(worksheet.iter_rows(values_only=False), start=1):
        values = [cell_text(cell) for cell in row]
        possible_map = {value: index for index, value in enumerate(values) if value}
        if all(column in possible_map for column in REQUIRED_COLUMNS):
            possible_map["__header_row"] = row_number
            return possible_map
    missing = ", ".join(REQUIRED_COLUMNS)
    raise ValueError(f"{worksheet.title} is missing required columns: {missing}")


def build_driver_lookup(repository):
    lookup = {}
    for driver in repository.list_specification_drivers():
        lookup[normalize_key(driver.name)] = driver
    return lookup


def prepare_countryside_rows(rows, driver_lookup):
    prepared_rows = []
    unresolved = Counter()
    now = timestamp()

    for index, row in enumerate(rows):
        status = clean_text(row.get("Status"))
        active_flag = normalize_bool(
            row.get("Active_Flag"),
            default=normalize_key(status) == "active",
        )
        if normalize_key(status) == "on_hold" or active_flag is False:
            continue
        if normalize_key(status) != "active":
            continue

        assigned_to_alias = clean_text(row.get("Assigned to"))
        resolved_name = resolve_driver_name(assigned_to_alias)
        driver = driver_lookup.get(normalize_key(resolved_name)) if resolved_name else None
        if assigned_to_alias and not driver:
            unresolved[assigned_to_alias] += 1

        route_group_name = clean_text(row.get("__sheet_name")) or "Countryside"
        route_group = OpShopCountrysideRouteGroup(
            route_group_id=deterministic_id(
                "OPSHOP-COUNTRYSIDE-GROUP",
                normalize_key(route_group_name),
            ),
            route_group_name=route_group_name,
            status="Active",
            active_flag=True,
            display_order=index,
            source_marker=WORKBOOK_IMPORT_COUNTRYSIDE,
            created_at=now,
            updated_at=now,
        )

        location_key = location_dedupe_key(row)
        opshop_id = deterministic_id("OPSHOP", location_key)
        pickup_frequency = clean_text(row.get("Pickup_Frequency")) or "On Call"
        run_day = normalize_run_day(row.get("Run_Day"))
        status_notes = merge_status_notes(
            row.get("Status_Start_Date"),
            row.get("Status_Notes"),
        )
        schedule = OpShopPickupSchedule(
            schedule_id=deterministic_id(
                "OPSHOP-SCHEDULE",
                "|".join(
                    [
                        opshop_id,
                        route_group.route_group_id,
                        run_day or "",
                        "ON_CALL",
                        "COUNTRYSIDE",
                        normalize_key(pickup_frequency),
                        normalize_key(row.get("Time_Window")),
                    ]
                ),
            ),
            opshop_id=opshop_id,
            run_day=run_day,
            run_type="ON_CALL",
            pickup_frequency=pickup_frequency,
            time_window=clean_text(row.get("Time_Window")),
            call_before_arrival=normalize_bool(
                row.get("Call_Before_Arrival"),
                default=False,
            ),
            call_timing=clean_text(row.get("Call_Timing")),
            status="Active",
            active_flag=True,
            fortnight_group=None,
            review_required=False,
            review_reason=WORKBOOK_IMPORT_COUNTRYSIDE,
            created_at=now,
            updated_at=now,
            default_driver_id=driver.driver_id if driver else None,
            default_driver_alias=assigned_to_alias,
            default_driver_name_snapshot=resolved_name,
            pickup_category="COUNTRYSIDE",
            route_group_id=route_group.route_group_id,
        )
        location = OpShopLocation(
            opshop_id=opshop_id,
            name=clean_text(row.get("Op_Shop_Name")) or "Unknown OP SHOP",
            suburb=clean_text(row.get("Suburb")),
            street_address=clean_text(row.get("Street_Address")),
            area_region=clean_text(row.get("Area_Region")),
            primary_contact=clean_text(row.get("Primary_Contact")),
            primary_phone=clean_text(row.get("Primary_Phone")),
            secondary_contact=clean_text(row.get("Secondary_Contact")),
            secondary_phone=clean_text(row.get("Secondary_Phone")),
            access_type=clean_text(row.get("Access_Type")),
            key_required=normalize_bool(row.get("Key_Required"), default=False),
            trailer_restriction=clean_text(row.get("Trailer_Restriction")),
            status_notes=status_notes,
            is_active=True,
            created_at=now,
            updated_at=now,
        )
        prepared_rows.append(
            PreparedCountrysideRow(
                location_key=location_key,
                route_group=route_group,
                location=location,
                schedule=schedule,
            )
        )

    return prepared_rows, unresolved


def merge_status_notes(status_start_date, status_notes):
    notes = clean_text(status_notes)
    start_date = normalize_excel_date(status_start_date)
    if start_date and notes:
        return f"Status start date: {start_date}\n{notes}"
    if start_date:
        return f"Status start date: {start_date}"
    return notes


def normalize_excel_date(value):
    text = clean_text(value)
    if not text:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date().isoformat()
        except (TypeError, ValueError):
            return None
    if text.replace(".", "", 1).isdigit():
        try:
            return from_excel(float(text)).date().isoformat()
        except (TypeError, ValueError):
            return None
    try:
        return datetime.fromisoformat(text).date().isoformat()
    except ValueError:
        return text


def resolve_driver_name(alias):
    normalized = normalize_key(alias)
    if not normalized:
        return None
    return DRIVER_ALIAS_TO_NAME.get(normalized, alias)


def find_route_group_by_name(connection, route_group_name):
    row = connection.execute(
        """
        SELECT *
        FROM opshop_countryside_route_groups
        WHERE lower(trim(route_group_name)) = ?
        """,
        (normalize_key(route_group_name),),
    ).fetchone()
    return row


def find_location_id_by_key(connection, dedupe_key):
    stable_opshop_id = deterministic_id("OPSHOP", dedupe_key)
    row = connection.execute(
        "SELECT opshop_id FROM opshop_locations WHERE opshop_id = ?",
        (stable_opshop_id,),
    ).fetchone()
    if row:
        return row["opshop_id"]

    normalized_name, normalized_suburb, normalized_address = dedupe_key.split("|", 2)
    row = connection.execute(
        """
        SELECT opshop_id
        FROM opshop_locations
        WHERE lower(trim(name)) = ?
            AND lower(trim(COALESCE(suburb, ''))) = ?
            AND lower(trim(COALESCE(street_address, ''))) = ?
        """,
        (normalized_name, normalized_suburb, normalized_address),
    ).fetchone()
    return row["opshop_id"] if row else None


def find_schedule_id_by_key(connection, key):
    stable_schedule_id = deterministic_id("OPSHOP-SCHEDULE", key)
    row = connection.execute(
        "SELECT schedule_id FROM opshop_pickup_schedules WHERE schedule_id = ?",
        (stable_schedule_id,),
    ).fetchone()
    if row:
        return row["schedule_id"]

    (
        opshop_id,
        route_group_id,
        run_day,
        run_type,
        pickup_category,
        pickup_frequency,
        time_window,
    ) = key.split("|", 6)
    row = connection.execute(
        """
        SELECT schedule_id
        FROM opshop_pickup_schedules
        WHERE opshop_id = ?
            AND COALESCE(route_group_id, '') = ?
            AND COALESCE(run_day, '') = ?
            AND run_type = ?
            AND COALESCE(pickup_category, 'NORMAL') = ?
            AND lower(trim(COALESCE(pickup_frequency, ''))) = ?
            AND lower(trim(COALESCE(time_window, ''))) = ?
        """,
        (
            opshop_id,
            route_group_id,
            run_day,
            run_type,
            pickup_category,
            pickup_frequency,
            time_window,
        ),
    ).fetchone()
    return row["schedule_id"] if row else None


def deactivate_missing_countryside_schedules(connection, imported_schedule_ids):
    timestamp_value = timestamp()
    if imported_schedule_ids:
        placeholders = ", ".join("?" for _ in imported_schedule_ids)
        parameters = [
            "On_Hold",
            0,
            timestamp_value,
            WORKBOOK_IMPORT_COUNTRYSIDE,
            *sorted(imported_schedule_ids),
        ]
        cursor = connection.execute(
            f"""
            UPDATE opshop_pickup_schedules
            SET status = ?,
                active_flag = ?,
                updated_at = ?
            WHERE run_type = 'ON_CALL'
                AND COALESCE(pickup_category, 'NORMAL') = 'COUNTRYSIDE'
                AND active_flag = 1
                AND review_reason = ?
                AND schedule_id NOT IN ({placeholders})
            """,
            parameters,
        )
    else:
        cursor = connection.execute(
            """
            UPDATE opshop_pickup_schedules
            SET status = ?,
                active_flag = ?,
                updated_at = ?
            WHERE run_type = 'ON_CALL'
                AND COALESCE(pickup_category, 'NORMAL') = 'COUNTRYSIDE'
                AND active_flag = 1
                AND review_reason = ?
            """,
            ("On_Hold", 0, timestamp_value, WORKBOOK_IMPORT_COUNTRYSIDE),
        )
    return cursor.rowcount


def deactivate_missing_countryside_route_groups(connection, imported_route_group_ids):
    timestamp_value = timestamp()
    if imported_route_group_ids:
        placeholders = ", ".join("?" for _ in imported_route_group_ids)
        parameters = [
            "On_Hold",
            0,
            timestamp_value,
            WORKBOOK_IMPORT_COUNTRYSIDE,
            *sorted(imported_route_group_ids),
        ]
        cursor = connection.execute(
            f"""
            UPDATE opshop_countryside_route_groups
            SET status = ?,
                active_flag = ?,
                updated_at = ?
            WHERE active_flag = 1
                AND source_marker = ?
                AND route_group_id NOT IN ({placeholders})
            """,
            parameters,
        )
    else:
        cursor = connection.execute(
            """
            UPDATE opshop_countryside_route_groups
            SET status = ?,
                active_flag = ?,
                updated_at = ?
            WHERE active_flag = 1
                AND source_marker = ?
            """,
            ("On_Hold", 0, timestamp_value, WORKBOOK_IMPORT_COUNTRYSIDE),
        )
    return cursor.rowcount


def schedule_key(schedule):
    return "|".join(
        [
            schedule.opshop_id,
            schedule.route_group_id or "",
            schedule.run_day or "",
            schedule.run_type,
            schedule.pickup_category,
            normalize_key(schedule.pickup_frequency),
            normalize_key(schedule.time_window),
        ]
    )


def location_dedupe_key(row):
    return "|".join(
        [
            normalize_key(row.get("Op_Shop_Name")),
            normalize_key(row.get("Suburb")),
            normalize_key(row.get("Street_Address")),
        ]
    )


def deterministic_id(prefix, key):
    digest = hashlib.sha1(key.encode("utf-8")).hexdigest()[:16].upper()
    return f"{prefix}-{digest}"


def normalize_run_day(value):
    normalized = normalize_key(value).upper()
    if not normalized:
        return None
    normalized = normalized.replace("-", "_").replace(" ", "_")
    return normalized if normalized in VALID_RUN_DAYS else None


def normalize_bool(value, default=False):
    normalized = normalize_key(value)
    if not normalized:
        return default
    if normalized in {"1", "yes", "y", "true"}:
        return True
    if normalized in {"0", "no", "n", "false"}:
        return False
    return default


def cell_text(cell):
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.replace(tzinfo=None).isoformat(sep=" ")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_text(value):
    if value is None:
        return None
    cleaned = "\n".join(
        " ".join(part.strip().split())
        for part in str(value).strip().splitlines()
        if part.strip()
    )
    return cleaned or None


def normalize_key(value):
    return " ".join(str(value or "").strip().lower().replace("-", "_").split())


def timestamp():
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def print_summary(summary):
    data = asdict(summary)
    for key in [
        "sheets_read",
        "rows_read",
        "rows_imported",
        "route_groups_inserted",
        "route_groups_updated",
        "route_groups_deactivated",
        "locations_inserted",
        "locations_updated",
        "schedules_inserted",
        "schedules_updated",
        "schedules_deactivated",
        "duplicate_locations_reused",
    ]:
        print(f"{key}: {data[key]}")
    print("unresolved_assigned_to:")
    if data["unresolved_assigned_to"]:
        for alias, count in sorted(data["unresolved_assigned_to"].items()):
            print(f"  - {alias}: {count}")
    else:
        print("  - none")
    print(f"backup_path: {data['backup_path'] or 'no existing database'}")


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Import Countryside OP SHOP workbook route groups and templates.",
    )
    parser.add_argument("--file", required=True, help="Path to Opshop countryside pickup.xlsx")
    parser.add_argument("--db-path", help="Target SQLite database path")
    add_maintenance_logbook_arguments(parser)
    args = parser.parse_args(argv)

    actor = resolve_maintenance_actor(args.actor)
    database_path = resolve_db_path(args.db_path)
    workbook_filename = safe_basename(args.file)
    try:
        summary = import_countryside_opshop_pickups_to_db(args.file, args.db_path)
        print_summary(summary)
    except Exception as error:
        record_maintenance_event(
            action="COUNTRYSIDE_WORKBOOK_IMPORT_COMPLETED",
            result="FAILED",
            workspace="OPSHOP",
            actor=actor,
            entity_type="OPSHOP_WORKBOOK_IMPORT",
            entity_id=f"countryside:{workbook_filename}",
            summary="Countryside OP SHOP workbook import failed.",
            metadata={
                "mode": "apply",
                "workbook_filename": workbook_filename,
                "database_filename": safe_basename(database_path),
                **sanitized_failure_metadata(
                    error,
                    workbook_import_failure_phase(error, args.file),
                ),
            },
            logbook_dir=args.logbook_dir,
        )
        raise

    metadata = workbook_import_metadata(
        summary,
        workbook_path=args.file,
        database_path=database_path,
        count_fields=(
            "sheets_read",
            "rows_read",
            "rows_imported",
            "route_groups_inserted",
            "route_groups_updated",
            "route_groups_deactivated",
            "locations_inserted",
            "locations_updated",
            "schedules_inserted",
            "schedules_updated",
            "schedules_deactivated",
            "duplicate_locations_reused",
        ),
    )
    has_unresolved_aliases = metadata["unresolved_alias_occurrence_count"] > 0
    result = "PARTIAL" if has_unresolved_aliases else "SUCCESS"
    event_summary = (
        "Countryside OP SHOP workbook import completed with unresolved driver aliases: "
        f"{summary.rows_imported} rows imported from {summary.sheets_read} sheets."
        if has_unresolved_aliases
        else (
            "Countryside OP SHOP workbook import completed: "
            f"{summary.rows_imported} rows imported from {summary.sheets_read} sheets."
        )
    )
    record_maintenance_event(
        action="COUNTRYSIDE_WORKBOOK_IMPORT_COMPLETED",
        result=result,
        workspace="OPSHOP",
        actor=actor,
        entity_type="OPSHOP_WORKBOOK_IMPORT",
        entity_id=f"countryside:{workbook_filename}",
        summary=event_summary,
        metadata=metadata,
        logbook_dir=args.logbook_dir,
    )


if __name__ == "__main__":
    main()
