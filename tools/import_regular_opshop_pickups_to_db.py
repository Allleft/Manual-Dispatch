"""Import the Regular OP SHOP Pickup workbook into SQLite source tables.

This importer reads the MON/TUE/WED/THU/FRI sheets from the current Regular
OP SHOP workbook and updates locations plus pickup schedules only. It does not
create pickup tasks; board loading is responsible for ensuring visible tasks.
"""

from __future__ import annotations

import argparse
import hashlib
import os
import re
import sqlite3
import sys
from collections import Counter
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from backend.repositories.sqlite_manual_dispatch_repository import (  # noqa: E402
    SQLiteManualDispatchRepository,
)
from backend.db.connection import connect  # noqa: E402
from backend.schemas import OpShopLocation, OpShopPickupSchedule  # noqa: E402
from tools.maintenance_logbook import (  # noqa: E402
    add_maintenance_logbook_arguments,
    record_maintenance_event,
    resolve_maintenance_actor,
    safe_basename,
    sanitized_failure_metadata,
    workbook_import_failure_phase,
    workbook_import_metadata,
)


DEFAULT_DB_PATH = PROJECT_ROOT / "data" / "manual_dispatch.sqlite3"
SHEET_RUN_DAYS = {
    "MON": "MONDAY",
    "TUE": "TUESDAY",
    "WED": "WEDNESDAY",
    "THU": "THURSDAY",
    "FRI": "FRIDAY",
}
REQUIRED_COLUMNS = [
    "Op_Shop_Name",
    "Run_Day",
    "Run_Type",
    "Active_Flag",
    "Suburb",
    "Street_Address",
    "Area_Region",
    "Primary_Contact",
    "Primary_Phone",
    "Secondary_Contact",
    "Secondary_Phone",
    "Pickup_Frequency",
    "Time_Window",
    "Call_Before_Arrival",
    "Call_Timing",
    "Access_Type",
    "Key_Required",
    "Trailer_Restriction",
    "Status",
    "Status_Start_Date",
    "Status_Notes",
    "Assigned to",
]
DRIVER_ALIAS_TO_NAME = {
    "john g": "John Georgiadis",
    "gavin": "Gavin Fynn",
    "nonda": "Epaminondas Tsatsoulis",
    "lee": "Guanlin Li",
}
WORKBOOK_IMPORT_REVIEW_REASON = "WORKBOOK_IMPORT"


@dataclass
class RegularOpShopImportSummary:
    rows_read: int = 0
    rows_imported: int = 0
    rows_skipped_inactive: int = 0
    locations_inserted: int = 0
    locations_updated: int = 0
    schedules_inserted: int = 0
    schedules_updated: int = 0
    schedules_deactivated: int = 0
    unresolved_assigned_to: dict[str, int] = field(default_factory=dict)
    default_driver_mapping_counts: dict[str, int] = field(default_factory=dict)
    backup_path: str | None = None


@dataclass
class PreparedRegularRow:
    location_key: str
    location: OpShopLocation
    schedule: OpShopPickupSchedule
    assigned_to_alias: str | None
    resolved_driver_name: str | None


@dataclass
class RegularImportPlan:
    locations: list[OpShopLocation]
    schedules: list[OpShopPickupSchedule]
    imported_schedule_ids: set[str]
    deactivation_schedule_ids: set[str]
    locations_inserted: int
    locations_updated: int
    schedules_inserted: int
    schedules_updated: int


def import_regular_opshop_pickups_to_db(file_path, db_path=None):
    workbook_path = Path(file_path)
    target_db_path = resolve_db_path(db_path)
    rows = read_regular_workbook_rows(workbook_path)
    backup_path = backup_database_if_exists(target_db_path)
    repository = SQLiteManualDispatchRepository(target_db_path)
    driver_lookup = build_driver_lookup(repository)
    prepared_rows, skipped_count, unresolved, mapping_counts = prepare_regular_rows(
        rows,
        driver_lookup,
    )

    summary = RegularOpShopImportSummary(
        rows_read=len(rows),
        rows_imported=len(prepared_rows),
        rows_skipped_inactive=skipped_count,
        unresolved_assigned_to=dict(unresolved),
        default_driver_mapping_counts=dict(mapping_counts),
        backup_path=str(backup_path) if backup_path else None,
    )
    with connect(target_db_path) as connection:
        connection.execute("BEGIN IMMEDIATE")
        try:
            plan = preflight_regular_import(connection, prepared_rows)
            for location in plan.locations:
                upsert_opshop_location_in_transaction(connection, location)
            for schedule in plan.schedules:
                upsert_opshop_schedule_in_transaction(connection, schedule)
            summary.schedules_deactivated = deactivate_regular_schedules(
                connection,
                plan.deactivation_schedule_ids,
            )
            connection.commit()
        except Exception:
            connection.rollback()
            raise

    summary.locations_inserted = plan.locations_inserted
    summary.locations_updated = plan.locations_updated
    summary.schedules_inserted = plan.schedules_inserted
    summary.schedules_updated = plan.schedules_updated

    return summary


def resolve_db_path(db_path=None):
    if db_path:
        return Path(db_path)
    configured = os.environ.get("MANUAL_DISPATCH_DB_PATH")
    if configured:
        return Path(configured)
    return DEFAULT_DB_PATH


def backup_database_if_exists(db_path):
    db_path = Path(db_path)
    if not db_path.exists():
        return None

    backup_dir = db_path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"manual_dispatch_before_regular_opshop_import_{timestamp}.sqlite3"

    source = sqlite3.connect(db_path)
    try:
        destination = sqlite3.connect(backup_path)
        try:
            source.backup(destination)
        finally:
            destination.close()
    finally:
        source.close()
    return backup_path


def read_regular_workbook_rows(workbook_path):
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    missing_sheets = [sheet for sheet in SHEET_RUN_DAYS if sheet not in workbook.sheetnames]
    if missing_sheets:
        raise ValueError(f"Workbook is missing required sheets: {', '.join(missing_sheets)}")

    rows = []
    for sheet_name, run_day in SHEET_RUN_DAYS.items():
        worksheet = workbook[sheet_name]
        header_map = _find_header_map(worksheet)
        for row in worksheet.iter_rows(min_row=header_map["__header_row"] + 1, values_only=False):
            record = {
                column: cell_text(row[header_map[column]])
                for column in REQUIRED_COLUMNS
            }
            if any(record.values()):
                record["__sheet_name"] = sheet_name
                record["__run_day"] = run_day
                record["__row_number"] = row[0].row
                rows.append(record)
    return rows


def _find_header_map(worksheet):
    for row_number, row in enumerate(worksheet.iter_rows(values_only=False), start=1):
        values = [cell_text(cell) for cell in row]
        possible_map = {value: index for index, value in enumerate(values) if value}
        if all(column in possible_map for column in REQUIRED_COLUMNS):
            possible_map["__header_row"] = row_number
            return possible_map
    missing = ", ".join(REQUIRED_COLUMNS)
    raise ValueError(f"{worksheet.title} is missing required columns: {missing}")


def build_driver_lookup(repository):
    lookup = {}
    for driver in repository.list_specification_drivers():
        lookup[normalize_key(driver.name)] = driver
    return lookup


def prepare_regular_rows(rows, driver_lookup):
    prepared_rows = []
    skipped_count = 0
    unresolved = Counter()
    mapping_counts = Counter()
    now = timestamp()

    for row in rows:
        status = clean_text(row.get("Status"))
        active_flag = normalize_bool(
            row.get("Active_Flag"),
            default=normalize_key(status) == "active",
        )
        if normalize_key(status) == "on_hold" or active_flag is False:
            skipped_count += 1
            continue
        if normalize_key(status) != "active":
            skipped_count += 1
            continue

        assigned_to_alias = clean_text(row.get("Assigned to"))
        resolved_name = resolve_driver_name(assigned_to_alias)
        driver = driver_lookup.get(normalize_key(resolved_name)) if resolved_name else None
        if assigned_to_alias and not driver:
            unresolved[assigned_to_alias] += 1
        if driver:
            mapping_counts[driver.name] += 1

        location_key = location_dedupe_key(row)
        opshop_id = deterministic_id("OPSHOP", location_key)
        pickup_frequency = clean_text(row.get("Pickup_Frequency"))
        schedule = OpShopPickupSchedule(
            schedule_id=deterministic_id(
                "OPSHOP-SCHEDULE",
                "|".join(
                    [
                        opshop_id,
                        row["__run_day"],
                        "REGULAR",
                        "NORMAL",
                    ]
                ),
            ),
            opshop_id=opshop_id,
            run_day=row["__run_day"],
            run_type="REGULAR",
            pickup_frequency=pickup_frequency,
            time_window=clean_text(row.get("Time_Window")),
            call_before_arrival=normalize_bool(
                row.get("Call_Before_Arrival"),
                default=False,
            ),
            call_timing=clean_text(row.get("Call_Timing")),
            status="Active",
            active_flag=True,
            fortnight_group=None,
            review_required=False,
            review_reason=WORKBOOK_IMPORT_REVIEW_REASON,
            created_at=now,
            updated_at=now,
            default_driver_id=driver.driver_id if driver else None,
            default_driver_alias=assigned_to_alias,
            default_driver_name_snapshot=resolved_name,
        )
        location = OpShopLocation(
            opshop_id=opshop_id,
            name=clean_text(row.get("Op_Shop_Name")) or "Unknown OP SHOP",
            suburb=clean_text(row.get("Suburb")),
            street_address=clean_text(row.get("Street_Address")),
            area_region=clean_text(row.get("Area_Region")),
            primary_contact=clean_text(row.get("Primary_Contact")),
            primary_phone=clean_text(row.get("Primary_Phone")),
            secondary_contact=clean_text(row.get("Secondary_Contact")),
            secondary_phone=clean_text(row.get("Secondary_Phone")),
            access_type=clean_text(row.get("Access_Type")),
            key_required=normalize_bool(row.get("Key_Required"), default=False),
            trailer_restriction=clean_text(row.get("Trailer_Restriction")),
            status_notes=clean_text(row.get("Status_Notes")),
            is_active=True,
            created_at=now,
            updated_at=now,
        )
        prepared_rows.append(
            PreparedRegularRow(
                location_key=location_key,
                location=location,
                schedule=schedule,
                assigned_to_alias=assigned_to_alias,
                resolved_driver_name=resolved_name,
            )
        )

    return prepared_rows, skipped_count, unresolved, mapping_counts


def preflight_regular_import(connection, prepared_rows):
    imported_location_ids = {}
    imported_schedule_ids = set()
    source_schedule_keys = set()
    locations = []
    schedules = []
    locations_inserted = 0
    locations_updated = 0
    schedules_inserted = 0
    schedules_updated = 0

    for prepared in prepared_rows:
        if prepared.location_key not in imported_location_ids:
            existing_location_id = find_location_id_by_key(
                connection,
                prepared.location_key,
            )
            if existing_location_id:
                prepared.location.opshop_id = existing_location_id
                locations_updated += 1
            else:
                locations_inserted += 1
            imported_location_ids[prepared.location_key] = prepared.location.opshop_id
            locations.append(prepared.location)

        prepared.location.opshop_id = imported_location_ids[prepared.location_key]
        prepared.schedule.opshop_id = prepared.location.opshop_id
        source_schedule_key = schedule_key(prepared.schedule)
        if source_schedule_key in source_schedule_keys:
            raise ValueError(
                "Duplicate Regular workbook schedule slot for "
                f"{prepared.location.name} {prepared.schedule.run_day}"
            )
        source_schedule_keys.add(source_schedule_key)

        existing_schedule_id = find_schedule_id_by_key(
            connection,
            source_schedule_key,
            prepared.location_key,
        )
        if existing_schedule_id:
            prepared.schedule.schedule_id = existing_schedule_id
            schedules_updated += 1
        else:
            schedules_inserted += 1
        imported_schedule_ids.add(prepared.schedule.schedule_id)
        schedules.append(prepared.schedule)

    deactivation_schedule_ids = find_missing_regular_schedule_ids(
        connection,
        imported_schedule_ids,
    )
    return RegularImportPlan(
        locations=locations,
        schedules=schedules,
        imported_schedule_ids=imported_schedule_ids,
        deactivation_schedule_ids=deactivation_schedule_ids,
        locations_inserted=locations_inserted,
        locations_updated=locations_updated,
        schedules_inserted=schedules_inserted,
        schedules_updated=schedules_updated,
    )


def resolve_driver_name(alias):
    normalized = normalize_key(alias)
    if not normalized:
        return None
    return DRIVER_ALIAS_TO_NAME.get(normalized, alias)


def find_location_id_by_key(connection, dedupe_key):
    matches = []
    for candidate in connection.execute(
        """
        SELECT opshop_id, name, suburb, street_address
        FROM opshop_locations
        ORDER BY opshop_id
        """
    ).fetchall():
        candidate_key = location_key_from_values(
            candidate["name"],
            candidate["suburb"],
            candidate["street_address"],
        )
        if candidate_key == dedupe_key:
            matches.append(candidate["opshop_id"])
    if len(matches) > 1:
        raise ValueError(
            "Duplicate OP SHOP physical location identity for imported Regular row"
        )
    return matches[0] if matches else None


def find_schedule_id_by_key(connection, key, location_key=None):
    stable_schedule_id = deterministic_id("OPSHOP-SCHEDULE", key)
    stable_row = connection.execute(
        """
        SELECT schedule_id
        FROM opshop_pickup_schedules
        WHERE schedule_id = ?
        """,
        (stable_schedule_id,),
    ).fetchone()

    opshop_id, run_day, run_type, pickup_category = key.split("|", 3)
    rows = connection.execute(
        """
        SELECT schedule_id
        FROM opshop_pickup_schedules
        WHERE opshop_id = ?
            AND COALESCE(run_day, '') = ?
            AND run_type = ?
            AND COALESCE(pickup_category, 'NORMAL') = ?
        ORDER BY schedule_id
        """,
        (opshop_id, run_day, run_type, pickup_category),
    ).fetchall()
    if len(rows) > 1:
        raise ValueError(
            f"Duplicate OP SHOP schedule slot for {opshop_id} {run_day}"
        )
    if rows:
        return rows[0]["schedule_id"]
    if stable_row:
        raise ValueError(
            f"OP SHOP schedule identity conflict for {opshop_id} {run_day}"
        )

    if not location_key:
        return None
    physical_matches = []
    candidates = connection.execute(
        """
        SELECT
            schedule.schedule_id,
            location.name,
            location.suburb,
            location.street_address
        FROM opshop_pickup_schedules AS schedule
        JOIN opshop_locations AS location
            ON location.opshop_id = schedule.opshop_id
        WHERE COALESCE(schedule.run_day, '') = ?
            AND schedule.run_type = ?
            AND COALESCE(schedule.pickup_category, 'NORMAL') = ?
        ORDER BY schedule.schedule_id
        """,
        (run_day, run_type, pickup_category),
    ).fetchall()
    for candidate in candidates:
        candidate_key = location_key_from_values(
            candidate["name"],
            candidate["suburb"],
            candidate["street_address"],
        )
        if candidate_key == location_key:
            physical_matches.append(candidate["schedule_id"])
    if len(physical_matches) > 1:
        raise ValueError(
            f"Duplicate OP SHOP schedule slot for physical location {run_day}"
        )
    return physical_matches[0] if physical_matches else None


def upsert_opshop_location_in_transaction(connection, location):
    connection.execute(
        """
        INSERT INTO opshop_locations (
            opshop_id,
            name,
            suburb,
            street_address,
            area_region,
            primary_contact,
            primary_phone,
            secondary_contact,
            secondary_phone,
            access_type,
            key_required,
            trailer_restriction,
            status_notes,
            is_active,
            created_at,
            updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(opshop_id)
        DO UPDATE SET
            name = excluded.name,
            suburb = excluded.suburb,
            street_address = excluded.street_address,
            area_region = excluded.area_region,
            primary_contact = excluded.primary_contact,
            primary_phone = excluded.primary_phone,
            secondary_contact = excluded.secondary_contact,
            secondary_phone = excluded.secondary_phone,
            access_type = excluded.access_type,
            key_required = excluded.key_required,
            trailer_restriction = excluded.trailer_restriction,
            status_notes = excluded.status_notes,
            is_active = excluded.is_active,
            updated_at = excluded.updated_at
        """,
        (
            location.opshop_id,
            location.name,
            location.suburb,
            location.street_address,
            location.area_region,
            location.primary_contact,
            location.primary_phone,
            location.secondary_contact,
            location.secondary_phone,
            location.access_type,
            int(location.key_required),
            location.trailer_restriction,
            location.status_notes,
            int(location.is_active),
            location.created_at,
            location.updated_at,
        ),
    )


def upsert_opshop_schedule_in_transaction(connection, schedule):
    connection.execute(
        """
        INSERT INTO opshop_pickup_schedules (
            schedule_id,
            opshop_id,
            run_day,
            run_type,
            pickup_category,
            route_group_id,
            pickup_frequency,
            time_window,
            call_before_arrival,
            call_timing,
            status,
            active_flag,
            fortnight_group,
            review_required,
            review_reason,
            default_driver_id,
            default_driver_alias,
            default_driver_name_snapshot,
            created_at,
            updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(schedule_id)
        DO UPDATE SET
            opshop_id = excluded.opshop_id,
            run_day = excluded.run_day,
            run_type = excluded.run_type,
            pickup_category = excluded.pickup_category,
            route_group_id = excluded.route_group_id,
            pickup_frequency = excluded.pickup_frequency,
            time_window = excluded.time_window,
            call_before_arrival = excluded.call_before_arrival,
            call_timing = excluded.call_timing,
            status = excluded.status,
            active_flag = excluded.active_flag,
            fortnight_group = excluded.fortnight_group,
            review_required = excluded.review_required,
            review_reason = excluded.review_reason,
            default_driver_id = excluded.default_driver_id,
            default_driver_alias = excluded.default_driver_alias,
            default_driver_name_snapshot = excluded.default_driver_name_snapshot,
            updated_at = excluded.updated_at
        """,
        (
            schedule.schedule_id,
            schedule.opshop_id,
            schedule.run_day,
            schedule.run_type,
            schedule.pickup_category,
            schedule.route_group_id,
            schedule.pickup_frequency,
            schedule.time_window,
            int(schedule.call_before_arrival),
            schedule.call_timing,
            schedule.status,
            int(schedule.active_flag),
            schedule.fortnight_group,
            int(schedule.review_required),
            schedule.review_reason,
            schedule.default_driver_id,
            schedule.default_driver_alias,
            schedule.default_driver_name_snapshot,
            schedule.created_at,
            schedule.updated_at,
        ),
    )


def find_missing_regular_schedule_ids(connection, imported_schedule_ids):
    """Treat the workbook as the complete active Regular source list.

    Only workbook-backed schedules are deactivated. UI-created templates have
    no workbook source marker and must remain under office control.
    """
    parameters = [WORKBOOK_IMPORT_REVIEW_REASON]
    imported_filter = ""
    if imported_schedule_ids:
        placeholders = ", ".join("?" for _ in imported_schedule_ids)
        imported_filter = f"AND schedule_id NOT IN ({placeholders})"
        parameters.extend(sorted(imported_schedule_ids))
    rows = connection.execute(
        f"""
        SELECT schedule_id
        FROM opshop_pickup_schedules
        WHERE run_type = 'REGULAR'
            AND active_flag = 1
            AND (
                review_reason = ?
                OR default_driver_alias IS NOT NULL
            )
            {imported_filter}
        ORDER BY schedule_id
        """,
        parameters,
    ).fetchall()
    return {row["schedule_id"] for row in rows}


def deactivate_regular_schedules(connection, schedule_ids):
    if not schedule_ids:
        return 0
    placeholders = ", ".join("?" for _ in schedule_ids)
    cursor = connection.execute(
        f"""
        UPDATE opshop_pickup_schedules
        SET status = ?,
            active_flag = ?,
            updated_at = ?
        WHERE schedule_id IN ({placeholders})
            AND run_type = 'REGULAR'
            AND active_flag = 1
        """,
        ("On_Hold", 0, timestamp(), *sorted(schedule_ids)),
    )
    return cursor.rowcount


def schedule_key(schedule):
    return "|".join(
        [
            schedule.opshop_id,
            schedule.run_day or "",
            schedule.run_type,
            getattr(schedule, "pickup_category", "NORMAL") or "NORMAL",
        ]
    )


def location_dedupe_key(row):
    return location_key_from_values(
        row.get("Op_Shop_Name"),
        row.get("Suburb"),
        row.get("Street_Address"),
    )


def location_key_from_values(name, suburb, street_address):
    return "|".join(
        [
            normalize_source_component(name),
            normalize_source_component(suburb),
            normalize_source_address(street_address),
        ]
    )


def normalize_source_component(value):
    return " ".join(
        re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).split()
    )


def normalize_source_address(value):
    replacements = {
        "road": "rd",
        "street": "st",
        "avenue": "ave",
        "highway": "hwy",
        "parade": "pde",
        "drive": "dr",
        "court": "ct",
        "place": "pl",
        "lane": "ln",
    }
    return " ".join(
        replacements.get(token, token)
        for token in normalize_source_component(value).split()
    )


def deterministic_id(prefix, key):
    digest = hashlib.sha1(key.encode("utf-8")).hexdigest()[:16].upper()
    return f"{prefix}-{digest}"


def normalize_bool(value, default=False):
    normalized = normalize_key(value)
    if not normalized:
        return default
    if normalized in {"1", "yes", "y", "true"}:
        return True
    if normalized in {"0", "no", "n", "false"}:
        return False
    return default


def cell_text(cell):
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).strip().split())


def clean_text(value):
    cleaned = " ".join(str(value or "").strip().split())
    return cleaned or None


def normalize_key(value):
    return " ".join(str(value or "").strip().lower().replace("-", "_").split())


def timestamp():
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def print_summary(summary):
    data = asdict(summary)
    print(f"Rows read: {data['rows_read']}")
    print(f"Rows imported: {data['rows_imported']}")
    print(f"Rows skipped inactive/on-hold: {data['rows_skipped_inactive']}")
    print(f"Locations inserted: {data['locations_inserted']}")
    print(f"Locations updated: {data['locations_updated']}")
    print(f"Schedules inserted: {data['schedules_inserted']}")
    print(f"Schedules updated: {data['schedules_updated']}")
    print(f"Schedules deactivated: {data['schedules_deactivated']}")
    print("Default driver mapping counts:")
    if data["default_driver_mapping_counts"]:
        for driver_name, count in sorted(data["default_driver_mapping_counts"].items()):
            print(f"  - {driver_name}: {count}")
    else:
        print("  - none")
    print("Unresolved Assigned to aliases:")
    if data["unresolved_assigned_to"]:
        for alias, count in sorted(data["unresolved_assigned_to"].items()):
            print(f"  - {alias}: {count}")
    else:
        print("  - none")
    if data["backup_path"]:
        print(f"Backup created: {data['backup_path']}")
    else:
        print("Backup created: no existing database")


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Import Regular OP SHOP workbook into SQLite locations and schedules.",
    )
    parser.add_argument("--file", required=True, help="Path to Opshop reuglar pickup.xlsx")
    parser.add_argument("--db-path", help="Target SQLite database path")
    add_maintenance_logbook_arguments(parser)
    args = parser.parse_args(argv)

    actor = resolve_maintenance_actor(args.actor)
    database_path = resolve_db_path(args.db_path)
    workbook_filename = safe_basename(args.file)
    try:
        summary = import_regular_opshop_pickups_to_db(args.file, args.db_path)
        print_summary(summary)
    except Exception as error:
        record_maintenance_event(
            action="REGULAR_WORKBOOK_IMPORT_COMPLETED",
            result="FAILED",
            workspace="OPSHOP",
            actor=actor,
            entity_type="OPSHOP_WORKBOOK_IMPORT",
            entity_id=f"regular:{workbook_filename}",
            summary="Regular OP SHOP workbook import failed.",
            metadata={
                "mode": "apply",
                "workbook_filename": workbook_filename,
                "database_filename": safe_basename(database_path),
                **sanitized_failure_metadata(
                    error,
                    workbook_import_failure_phase(error, args.file),
                ),
            },
            logbook_dir=args.logbook_dir,
        )
        raise

    metadata = workbook_import_metadata(
        summary,
        workbook_path=args.file,
        database_path=database_path,
        count_fields=(
            "rows_read",
            "rows_imported",
            "rows_skipped_inactive",
            "locations_inserted",
            "locations_updated",
            "schedules_inserted",
            "schedules_updated",
            "schedules_deactivated",
        ),
    )
    has_unresolved_aliases = metadata["unresolved_alias_occurrence_count"] > 0
    result = "PARTIAL" if has_unresolved_aliases else "SUCCESS"
    event_summary = (
        "Regular OP SHOP workbook import completed with unresolved driver aliases: "
        f"{summary.rows_imported} rows imported."
        if has_unresolved_aliases
        else (
            "Regular OP SHOP workbook import completed: "
            f"{summary.rows_imported} rows imported."
        )
    )
    record_maintenance_event(
        action="REGULAR_WORKBOOK_IMPORT_COMPLETED",
        result=result,
        workspace="OPSHOP",
        actor=actor,
        entity_type="OPSHOP_WORKBOOK_IMPORT",
        entity_id=f"regular:{workbook_filename}",
        summary=event_summary,
        metadata=metadata,
        logbook_dir=args.logbook_dir,
    )


if __name__ == "__main__":
    main()
