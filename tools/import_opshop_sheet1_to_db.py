"""Import OP SHOP Sheet1 source data into SQLite master tables.

This imports OP SHOP locations and pickup schedules only. It does not create
opshop_pickup_tasks and does not alter the manual dispatch workflow.
"""

from __future__ import annotations

import argparse
import hashlib
import os
import sqlite3
import sys
from collections import Counter
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from backend.repositories.sqlite_manual_dispatch_repository import (  # noqa: E402
    SQLiteManualDispatchRepository,
)
from backend.schemas import OpShopLocation, OpShopPickupSchedule  # noqa: E402
from backend.services.manual_dispatch.opshop_pickup_service import (  # noqa: E402
    classify_pickup_frequency,
)


DEFAULT_DB_PATH = PROJECT_ROOT / "data" / "manual_dispatch.sqlite3"
REQUIRED_COLUMNS = [
    "Op_Shop_Name",
    "Run_Day",
    "Run_Type",
    "Active_Flag",
    "Suburb",
    "Street_Address",
    "Area_Region",
    "Primary_Contact",
    "Primary_Phone",
    "Secondary_Contact",
    "Secondary_Phone",
    "Pickup_Frequency",
    "Time_Window",
    "Call_Before_Arrival",
    "Call_Timing",
    "Access_Type",
    "Key_Required",
    "Trailer_Restriction",
    "Status",
    "Status_Start_Date",
    "Status_Notes",
]
RUN_DAY_MAP = {
    "monday": "MONDAY",
    "tuesday": "TUESDAY",
    "wednesday": "WEDNESDAY",
    "thursday": "THURSDAY",
    "friday": "FRIDAY",
}
RUN_TYPE_MAP = {
    "standard": "STANDARD",
    "regular": "REGULAR",
    "on call": "ON_CALL",
    "on_call": "ON_CALL",
}
@dataclass
class ImportSummary:
    rows_read: int = 0
    rows_imported: int = 0
    rows_skipped_inactive: int = 0
    locations_inserted: int = 0
    locations_updated: int = 0
    schedules_inserted: int = 0
    schedules_updated: int = 0
    review_required_count: int = 0
    review_required_by_reason: dict[str, int] = field(default_factory=dict)
    backup_path: str | None = None


@dataclass
class PreparedImportRow:
    source: dict
    location_key: str
    location: OpShopLocation
    schedule: OpShopPickupSchedule
    review_reasons: list[str]


def import_sheet1_to_db(file_path, db_path=None):
    workbook_path = Path(file_path)
    target_db_path = resolve_db_path(db_path)
    rows = read_sheet1_rows(workbook_path)
    prepared_rows, skipped_count = prepare_import_rows(rows)
    backup_path = backup_database_if_exists(target_db_path)

    repository = SQLiteManualDispatchRepository(target_db_path)
    summary = ImportSummary(
        rows_read=len(rows),
        rows_imported=len(prepared_rows),
        rows_skipped_inactive=skipped_count,
        backup_path=str(backup_path) if backup_path else None,
    )
    imported_locations = set()

    with sqlite3.connect(target_db_path) as connection:
        connection.row_factory = sqlite3.Row
        for prepared in prepared_rows:
            if prepared.location_key not in imported_locations:
                existing_location_id = find_location_id_by_key(connection, prepared.location_key)
                if existing_location_id:
                    prepared.location.opshop_id = existing_location_id
                    summary.locations_updated += 1
                else:
                    summary.locations_inserted += 1
                repository.upsert_opshop_location(prepared.location)
                imported_locations.add(prepared.location_key)

            prepared.schedule.opshop_id = prepared.location.opshop_id
            existing_schedule_id = find_schedule_id_by_key(
                connection,
                schedule_key(prepared.schedule),
            )
            if existing_schedule_id:
                prepared.schedule.schedule_id = existing_schedule_id
                summary.schedules_updated += 1
            else:
                summary.schedules_inserted += 1
            repository.upsert_opshop_pickup_schedule(prepared.schedule)

            if prepared.schedule.review_required:
                summary.review_required_count += 1
                for reason in prepared.review_reasons:
                    summary.review_required_by_reason[reason] = (
                        summary.review_required_by_reason.get(reason, 0) + 1
                    )

    return summary


def resolve_db_path(db_path=None):
    if db_path:
        return Path(db_path)
    configured = os.environ.get("MANUAL_DISPATCH_DB_PATH")
    if configured:
        return Path(configured)
    return DEFAULT_DB_PATH


def backup_database_if_exists(db_path):
    db_path = Path(db_path)
    if not db_path.exists():
        return None

    backup_dir = db_path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"manual_dispatch_before_opshop_sheet1_import_{timestamp}.sqlite3"

    source = sqlite3.connect(db_path)
    try:
        destination = sqlite3.connect(backup_path)
        try:
            source.backup(destination)
        finally:
            destination.close()
    finally:
        source.close()
    return backup_path


def read_sheet1_rows(workbook_path):
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    if "Sheet1" not in workbook.sheetnames:
        raise ValueError("Workbook must contain Sheet1")

    worksheet = workbook["Sheet1"]
    header_row = None
    header_map = None
    for row_number, row in enumerate(worksheet.iter_rows(values_only=False), start=1):
        values = [cell_text(cell) for cell in row]
        possible_map = {value: index for index, value in enumerate(values) if value}
        if all(column in possible_map for column in REQUIRED_COLUMNS):
            header_row = row_number
            header_map = possible_map
            break

    if header_map is None:
        missing = ", ".join(REQUIRED_COLUMNS)
        raise ValueError(f"Sheet1 is missing required columns: {missing}")

    rows = []
    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=False):
        record = {
            column: cell_text(row[header_map[column]])
            for column in REQUIRED_COLUMNS
        }
        if any(record.values()):
            rows.append(record)
    return rows


def prepare_import_rows(rows):
    prepared_rows = []
    skipped_count = 0
    now = timestamp()

    for row in rows:
        status = clean_text(row.get("Status"))
        active_flag = normalize_bool(row.get("Active_Flag"), default=True)
        if normalize_key(status) == "on_hold" or active_flag is False:
            skipped_count += 1
            continue
        if normalize_key(status) != "active":
            skipped_count += 1
            continue

        location_key = location_dedupe_key(row)
        opshop_id = deterministic_id("OPSHOP", location_key)
        run_type, run_type_unknown = normalize_run_type(row.get("Run_Type"))
        run_day = normalize_run_day(row.get("Run_Day"))
        pickup_frequency = clean_text(row.get("Pickup_Frequency"))
        review_reasons = review_reasons_for(
            run_type=run_type,
            run_type_unknown=run_type_unknown,
            run_day=run_day,
            pickup_frequency=pickup_frequency,
        )
        schedule = OpShopPickupSchedule(
            schedule_id=deterministic_id(
                "OPSHOP-SCHEDULE",
                "|".join(
                    [
                        opshop_id,
                        run_day or "",
                        run_type,
                        normalize_key(pickup_frequency),
                        normalize_key(row.get("Time_Window")),
                    ]
                ),
            ),
            opshop_id=opshop_id,
            run_day=run_day,
            run_type=run_type,
            pickup_frequency=pickup_frequency,
            time_window=clean_text(row.get("Time_Window")),
            call_before_arrival=normalize_bool(
                row.get("Call_Before_Arrival"),
                default=False,
            ),
            call_timing=clean_text(row.get("Call_Timing")),
            status="Active",
            active_flag=True,
            fortnight_group=None,
            review_required=bool(review_reasons),
            review_reason="; ".join(review_reasons) or None,
            created_at=now,
            updated_at=now,
        )
        location = OpShopLocation(
            opshop_id=opshop_id,
            name=clean_text(row.get("Op_Shop_Name")) or "Unknown OP SHOP",
            suburb=clean_text(row.get("Suburb")),
            street_address=clean_text(row.get("Street_Address")),
            area_region=clean_text(row.get("Area_Region")),
            primary_contact=clean_text(row.get("Primary_Contact")),
            primary_phone=clean_text(row.get("Primary_Phone")),
            secondary_contact=clean_text(row.get("Secondary_Contact")),
            secondary_phone=clean_text(row.get("Secondary_Phone")),
            access_type=clean_text(row.get("Access_Type")),
            key_required=normalize_bool(row.get("Key_Required"), default=False),
            trailer_restriction=clean_text(row.get("Trailer_Restriction")),
            status_notes=clean_text(row.get("Status_Notes")),
            is_active=True,
            created_at=now,
            updated_at=now,
        )
        prepared_rows.append(
            PreparedImportRow(
                source=row,
                location_key=location_key,
                location=location,
                schedule=schedule,
                review_reasons=review_reasons,
            )
        )

    return prepared_rows, skipped_count


def review_reasons_for(run_type, run_type_unknown, run_day, pickup_frequency):
    reasons = []
    normalized_frequency = normalize_key(pickup_frequency)
    frequency = classify_pickup_frequency(pickup_frequency)
    if run_type_unknown:
        reasons.append("Unknown run_type")
    if run_type in {"STANDARD", "REGULAR"} and not run_day:
        reasons.append("Missing run_day for STANDARD/REGULAR schedule")
    if run_type == "ON_CALL" and normalized_frequency in {"on call", "on_call"}:
        return reasons
    if frequency.frequency_type == "UNKNOWN":
        reasons.append("Blank or unknown pickup_frequency")
    if run_type in {"STANDARD", "REGULAR"} and frequency.frequency_type == "FORTNIGHTLY":
        reasons.append("Fortnightly schedule missing fortnight_group")
    if frequency.frequency_type == "MONTHLY":
        reasons.append("Monthly schedule requires review")
    return reasons


def find_location_id_by_key(connection, dedupe_key):
    stable_opshop_id = deterministic_id("OPSHOP", dedupe_key)
    row = connection.execute(
        """
        SELECT opshop_id
        FROM opshop_locations
        WHERE opshop_id = ?
        """,
        (stable_opshop_id,),
    ).fetchone()
    if row:
        return row["opshop_id"]

    normalized_name, normalized_suburb, normalized_address = dedupe_key.split("|", 2)
    row = connection.execute(
        """
        SELECT opshop_id
        FROM opshop_locations
        WHERE lower(trim(name)) = ?
            AND lower(trim(COALESCE(suburb, ''))) = ?
            AND lower(trim(COALESCE(street_address, ''))) = ?
        """,
        (normalized_name, normalized_suburb, normalized_address),
    ).fetchone()
    return row["opshop_id"] if row else None


def find_schedule_id_by_key(connection, key):
    stable_schedule_id = deterministic_id("OPSHOP-SCHEDULE", key)
    row = connection.execute(
        """
        SELECT schedule_id
        FROM opshop_pickup_schedules
        WHERE schedule_id = ?
        """,
        (stable_schedule_id,),
    ).fetchone()
    if row:
        return row["schedule_id"]

    opshop_id, run_day, run_type, pickup_frequency, time_window = key.split("|", 4)
    row = connection.execute(
        """
        SELECT schedule_id
        FROM opshop_pickup_schedules
        WHERE opshop_id = ?
            AND COALESCE(run_day, '') = ?
            AND run_type = ?
            AND lower(trim(COALESCE(pickup_frequency, ''))) = ?
            AND lower(trim(COALESCE(time_window, ''))) = ?
        """,
        (opshop_id, run_day, run_type, pickup_frequency, time_window),
    ).fetchone()
    return row["schedule_id"] if row else None


def schedule_key(schedule):
    return "|".join(
        [
            schedule.opshop_id,
            schedule.run_day or "",
            schedule.run_type,
            normalize_key(schedule.pickup_frequency),
            normalize_key(schedule.time_window),
        ]
    )


def location_dedupe_key(row):
    return "|".join(
        [
            normalize_key(row.get("Op_Shop_Name")),
            normalize_key(row.get("Suburb")),
            normalize_key(row.get("Street_Address")),
        ]
    )


def deterministic_id(prefix, key):
    digest = hashlib.sha1(key.encode("utf-8")).hexdigest()[:16].upper()
    return f"{prefix}-{digest}"


def normalize_run_type(value):
    normalized = normalize_key(value)
    if normalized in RUN_TYPE_MAP:
        return RUN_TYPE_MAP[normalized], False
    return "ON_CALL", True


def normalize_run_day(value):
    normalized = normalize_key(value)
    return RUN_DAY_MAP.get(normalized)


def normalize_bool(value, default=False):
    normalized = normalize_key(value)
    if not normalized:
        return default
    if normalized in {"1", "yes", "y", "true"}:
        return True
    if normalized in {"0", "no", "n", "false"}:
        return False
    return default


def cell_text(cell):
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).strip().split())


def clean_text(value):
    cleaned = " ".join(str(value or "").strip().split())
    return cleaned or None


def normalize_key(value):
    return " ".join(str(value or "").strip().lower().replace("-", "_").split())


def timestamp():
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def print_summary(summary):
    data = asdict(summary)
    print(f"Rows read: {data['rows_read']}")
    print(f"Rows imported: {data['rows_imported']}")
    print(f"Rows skipped inactive/on-hold: {data['rows_skipped_inactive']}")
    print(f"Locations inserted: {data['locations_inserted']}")
    print(f"Locations updated: {data['locations_updated']}")
    print(f"Schedules inserted: {data['schedules_inserted']}")
    print(f"Schedules updated: {data['schedules_updated']}")
    print(f"Review required count: {data['review_required_count']}")
    if data["review_required_by_reason"]:
        print("Review required by reason:")
        for reason, count in sorted(data["review_required_by_reason"].items()):
            print(f"  - {reason}: {count}")
    if data["backup_path"]:
        print(f"Backup created: {data['backup_path']}")
    else:
        print("Backup created: no existing database")


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Import OP SHOP Sheet1 data into SQLite locations and schedules.",
    )
    parser.add_argument("--file", required=True, help="Path to opshop_final_rechecked_v2.xlsx")
    parser.add_argument("--db-path", help="Target SQLite database path")
    args = parser.parse_args(argv)

    summary = import_sheet1_to_db(args.file, args.db_path)
    print_summary(summary)


if __name__ == "__main__":
    main()
