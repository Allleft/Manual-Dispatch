from io import BytesIO
import shutil
import sqlite3
import unittest
import uuid
from unittest.mock import patch
from pathlib import Path

from openpyxl import load_workbook

from backend.repositories.sqlite_manual_dispatch_repository import (
    SQLiteManualDispatchRepository,
)
from backend.schemas import (
    AssignTaskRequest,
    CreateOrderRequest,
    RegisterOperatorAccountRequest,
    SaveFinalTripSummaryRequest,
    UpdateOrderRequest,
)
from backend.services.final_summary_excel_export_service import (
    build_final_summary_excel,
)
from backend.services.manual_dispatch_service import ManualDispatchService


class ManualDispatchProductDetailsTest(unittest.TestCase):
    def setUp(self):
        temp_parent = Path.cwd() / "tmp"
        temp_parent.mkdir(exist_ok=True)
        self.temp_dir = temp_parent / f"product-details-test-{uuid.uuid4().hex}"
        self.temp_dir.mkdir()
        self.db_path = self.temp_dir / "manual_dispatch.sqlite3"
        with patch.dict("os.environ", {"MANUAL_DISPATCH_SEED_DEMO_DATA": "true"}):
            self.repository = SQLiteManualDispatchRepository(self.db_path)
        self.service = ManualDispatchService(self.repository)
        self.dispatch_date = "2026-05-05"
        self.account = self.service.register_operator_account(
            RegisterOperatorAccountRequest(
                account_name="Mandy",
                password="secret123",
                confirm_password="secret123",
            )
        )

    def tearDown(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_create_order_with_pallet_product_lines_persists_to_board(self):
        created = self.service.create_order(
            self._create_request(
                pallet_quantity=5,
                product_lines=self._pallet_lines(),
            )
        )

        board_order = next(
            order
            for order in self.service.get_board(self.dispatch_date).orders
            if order.order_id == created.order_id
        )

        self.assertEqual(
            ["colour singlet 10kg", "pure white singlet 10kg"],
            [line.product_name for line in board_order.product_lines],
        )
        self.assertEqual([3, 2], [line.quantity for line in board_order.product_lines])
        self.assertEqual(["PALLETS", "PALLETS"], [line.unit for line in board_order.product_lines])

    def test_create_order_with_bag_product_lines_persists_to_board(self):
        created = self.service.create_order(
            self._create_request(
                pallet_quantity=0,
                loose_bags_quantity=5,
                product_lines=[
                    {
                        "product_name": "colour singlet 10kg",
                        "quantity": 5,
                        "unit": "BAGS",
                    }
                ],
            )
        )

        board_order = next(
            order
            for order in self.service.get_board(self.dispatch_date).orders
            if order.order_id == created.order_id
        )

        self.assertEqual("BAGS", board_order.product_lines[0].unit)
        self.assertEqual(5, board_order.product_lines[0].quantity)

    def test_allows_mixed_product_detail_units_when_load_is_mixed(self):
        created = self.service.create_order(
            self._create_request(
                pallet_quantity=5,
                loose_bags_quantity=2,
                product_lines=[
                    {
                        "product_name": "colour singlet 10kg",
                        "quantity": 3,
                        "unit": "PALLETS",
                    },
                    {
                        "product_name": "pure white singlet 10kg",
                        "quantity": 2,
                        "unit": "BAGS",
                    },
                    {
                        "product_name": "COLOR RAGS 1.5KG BAG",
                        "quantity": 2,
                        "unit": "CARTONS",
                    },
                ],
            )
        )

        self.assertEqual(
            ["PALLETS", "BAGS", "CARTONS"],
            [line.unit for line in created.product_lines],
        )

    def test_pallet_order_allows_carton_product_details_without_changing_totals(self):
        created = self.service.create_order(
            self._create_request(
                invoice_number="184068",
                pallet_quantity=1,
                loose_bags_quantity=0,
                product_lines=self._pallet_and_carton_lines(),
            )
        )
        reloaded = self.repository.get_order(created.order_id)

        self.assertEqual(1, reloaded.pallet_quantity)
        self.assertEqual(0, reloaded.loose_bags_quantity)
        self.assertEqual(
            ["PALLETS", "CARTONS"],
            [line.unit for line in reloaded.product_lines],
        )

    def test_carton_product_detail_requires_a_pallet_quantity(self):
        for loose_bags_quantity in (0, 2):
            with self.subTest(loose_bags_quantity=loose_bags_quantity):
                with self.assertRaisesRegex(
                    ValueError,
                    "Product detail CARTONS requires a pallet quantity",
                ):
                    self.service.create_order(
                        self._create_request(
                            pallet_quantity=0,
                            loose_bags_quantity=loose_bags_quantity,
                            product_lines=[
                                {
                                    "product_name": "COLOR RAGS 1.5KG BAG",
                                    "quantity": 2,
                                    "unit": "CARTONS",
                                }
                            ],
                        )
                    )

    def test_rejects_product_detail_unit_that_conflicts_with_order_load_unit(self):
        with self.assertRaisesRegex(
            ValueError,
            "Product detail unit must align with the Order pallet or bag quantity",
        ):
            self.service.create_order(
                self._create_request(
                    pallet_quantity=5,
                    product_lines=[
                        {
                            "product_name": "colour singlet 10kg",
                            "quantity": 5,
                            "unit": "BAGS",
                        }
                    ],
                )
            )

    def test_rejects_invalid_product_detail_quantity(self):
        with self.assertRaisesRegex(
            ValueError,
            "quantity must be greater than 0",
        ):
            self.service.create_order(
                self._create_request(
                    pallet_quantity=1,
                    product_lines=[
                        {
                            "product_name": "colour singlet 10kg",
                            "quantity": 0,
                            "unit": "PALLETS",
                        }
                    ],
                )
            )

    def test_update_order_replaces_product_details(self):
        created = self.service.create_order(
            self._create_request(pallet_quantity=5, product_lines=self._pallet_lines())
        )

        updated = self.service.update_order(
            created.order_id,
            UpdateOrderRequest(
                invoice_number=created.invoice_number,
                company_name=created.company_name,
                phone=created.phone,
                delivery_address=created.delivery_address,
                suburb=created.suburb,
                postcode=created.postcode,
                delivery_date=created.delivery_date,
                zone=created.zone,
                urgency=created.urgency,
                preferred_driver_id=created.preferred_driver_id,
                pallet_quantity=0,
                loose_bags_quantity=5,
                start_time=created.start_time,
                end_time=created.end_time,
                note=created.note,
                product_lines=[
                    {
                        "product_name": "colour singlet 10kg",
                        "quantity": 5,
                        "unit": "BAGS",
                    }
                ],
            ),
        )

        self.assertEqual(0, updated.pallet_quantity)
        self.assertEqual(5, updated.loose_bags_quantity)
        self.assertEqual(["BAGS"], [line.unit for line in updated.product_lines])
        self.assertEqual(["colour singlet 10kg"], [line.product_name for line in updated.product_lines])

    def test_final_summary_snapshot_preserves_product_details(self):
        created = self.service.create_order(
            self._create_request(pallet_quantity=5, product_lines=self._pallet_lines())
        )
        self._assign_order(created.order_id)
        saved = self.service.save_final_trip_summary(
            self._summary_request(created.order_id)
        )

        with sqlite3.connect(self.db_path) as connection:
            connection.execute(
                """
                UPDATE order_product_lines
                SET product_name = ?
                WHERE order_id = ? AND line_no = 1
                """,
                ("edited after save", created.order_id),
            )
            connection.commit()

        detail = self.service.get_final_trip_summary(saved.summary_id)
        product_lines = detail.trips[0].orders[0].product_lines_snapshot

        self.assertEqual(
            ["colour singlet 10kg", "pure white singlet 10kg"],
            [line.product_name for line in product_lines],
        )

    def test_excel_export_includes_product_details_and_load_text(self):
        created = self.service.create_order(
            self._create_request(pallet_quantity=5, product_lines=self._pallet_lines())
        )
        self._assign_order(created.order_id)
        self.service.save_final_trip_summary(self._summary_request(created.order_id))

        workbook = load_workbook(
            BytesIO(
                build_final_summary_excel(
                    self.service.list_final_trip_summaries(self.dispatch_date),
                    self.dispatch_date,
                )
            ),
            data_only=True,
        )
        values = [
            cell
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows(values_only=True)
            for cell in row
            if cell is not None
        ]

        self.assertIn("Product Details", values)
        self.assertIn("Load", values)
        self.assertIn(
            "1. colour singlet 10kg - 3 Pallets\n2. pure white singlet 10kg - 2 Pallets",
            values,
        )
        self.assertIn("5 Pallets", values)

    def test_final_summary_excel_formats_cartons_without_changing_load(self):
        created = self.service.create_order(
            self._create_request(
                invoice_number="184068",
                pallet_quantity=1,
                loose_bags_quantity=0,
                product_lines=self._pallet_and_carton_lines(),
            )
        )
        self._assign_order(created.order_id)
        self.service.save_final_trip_summary(self._summary_request(created.order_id))

        workbook = load_workbook(
            BytesIO(
                build_final_summary_excel(
                    self.service.list_final_trip_summaries(self.dispatch_date),
                    self.dispatch_date,
                )
            ),
            data_only=True,
        )
        values = [
            cell
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows(values_only=True)
            for cell in row
            if cell is not None
        ]

        self.assertIn(
            "1. COLOUR RAGS 10KG NET - 1 Pallet\n"
            "2. COLOR RAGS 1.5KG BAG - 2 Cartons",
            values,
        )
        self.assertIn("1 Pallet", values)
        self.assertNotIn("2 Bags", values)

    def _create_request(self, **overrides):
        values = {
            "invoice_number": "VIC-PROD-001",
            "company_name": "Product Demo Customer",
            "phone": "0400 555 100",
            "delivery_address": "10 Product Lane",
            "suburb": "Richmond",
            "postcode": "3121",
            "delivery_date": self.dispatch_date,
            "zone": "Inner East",
            "urgency": "Normal",
            "preferred_driver_id": "",
            "pallet_quantity": 0,
            "loose_bags_quantity": 0,
            "start_time": "08:00",
            "end_time": "12:00",
            "note": "Phase 16 product detail test",
            "product_lines": [],
        }
        values.update(overrides)
        return CreateOrderRequest(**values)

    def _pallet_lines(self):
        return [
            {
                "product_name": "colour singlet 10kg",
                "quantity": 3,
                "unit": "PALLETS",
            },
            {
                "product_name": "pure white singlet 10kg",
                "quantity": 2,
                "unit": "PALLETS",
            },
        ]

    def _pallet_and_carton_lines(self):
        return [
            {
                "product_name": "COLOUR RAGS 10KG NET",
                "quantity": 1,
                "unit": "PALLETS",
            },
            {
                "product_name": "COLOR RAGS 1.5KG BAG",
                "quantity": 2,
                "unit": "CARTONS",
            },
        ]

    def _assign_order(self, order_id):
        return self.service.assign_task(
            AssignTaskRequest(
                dispatch_date=self.dispatch_date,
                task_type="ORDER",
                task_id=order_id,
                driver_id="D001",
                trip_no="trip1",
            )
        )

    def _summary_request(self, order_id):
        order = self.repository.get_order(order_id)
        return SaveFinalTripSummaryRequest(
            dispatch_date=self.dispatch_date,
            delivery_date=self.dispatch_date,
            driver_id="D001",
            driver_name_snapshot="John",
            vehicle_id=None,
            vehicle_rego_snapshot="No vehicle selected",
            total_pallets=order.pallet_quantity,
            total_loose_bags=order.loose_bags_quantity,
            generated_at="2026-05-05T00:00:00Z",
            saved_by_account_name=self.account.account_name,
            saved_by_account_id=self.account.account_id,
            trips=[
                {
                    "trip_no": "trip1",
                    "orders": [
                        {
                            "task_type": "ORDER",
                            "task_id": order.order_id,
                            "order_id_snapshot": order.order_id,
                            "invoice_number_snapshot": order.invoice_number,
                            "company_name_snapshot": order.company_name,
                            "suburb_snapshot": order.suburb,
                            "delivery_address_snapshot": order.delivery_address,
                            "pallet_quantity_snapshot": order.pallet_quantity,
                            "loose_bags_quantity_snapshot": order.loose_bags_quantity,
                            "note_snapshot": order.note,
                        }
                    ],
                }
            ],
        )


if __name__ == "__main__":
    unittest.main()
