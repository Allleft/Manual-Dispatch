import shutil
import unittest
import uuid
from pathlib import Path

from openpyxl import Workbook

from backend.repositories.sqlite_manual_dispatch_repository import (
    SQLiteManualDispatchRepository,
)
from tools.import_opshop_sheet1_to_db import REQUIRED_COLUMNS, import_sheet1_to_db


class ImportOpShopSheet1ToDbTest(unittest.TestCase):
    def setUp(self):
        self.temp_dir = Path.cwd() / "tmp" / f"opshop-sheet1-import-test-{uuid.uuid4().hex}"
        self.temp_dir.mkdir(parents=True)
        self.workbook_path = self.temp_dir / "opshop_final_rechecked_v2.xlsx"
        self.db_path = self.temp_dir / "manual_dispatch.sqlite3"

    def tearDown(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_import_creates_locations_and_schedules_without_pickup_tasks(self):
        self._save_workbook(
            [
                self._row(
                    Op_Shop_Name="Northside Op Shop",
                    Run_Day="Monday",
                    Run_Type="Standard",
                    Suburb="Coburg",
                    Street_Address="1 Sydney Road",
                    Primary_Phone="0400 111 222",
                    Pickup_Frequency="Weekly",
                )
            ]
        )

        summary = import_sheet1_to_db(self.workbook_path, self.db_path)
        repository = SQLiteManualDispatchRepository(self.db_path)

        self.assertEqual(1, summary.rows_read)
        self.assertEqual(1, summary.rows_imported)
        self.assertEqual(1, summary.locations_inserted)
        self.assertEqual(1, summary.schedules_inserted)
        self.assertEqual([], repository.list_opshop_pickup_tasks())

        location = repository.list_opshop_locations()[0]
        schedule = repository.list_opshop_pickup_schedules()[0]
        self.assertEqual("Northside Op Shop", location.name)
        self.assertEqual("Coburg", location.suburb)
        self.assertEqual("1 Sydney Road", location.street_address)
        self.assertEqual("0400 111 222", location.primary_phone)
        self.assertEqual(location.opshop_id, schedule.opshop_id)
        self.assertEqual("MONDAY", schedule.run_day)
        self.assertEqual("STANDARD", schedule.run_type)
        self.assertEqual("Weekly", schedule.pickup_frequency)

    def test_same_location_multiple_schedule_rows_create_one_location(self):
        self._save_workbook(
            [
                self._row(Run_Day="Monday", Run_Type="Standard", Pickup_Frequency="Weekly"),
                self._row(Run_Day="Wednesday", Run_Type="Regular", Pickup_Frequency="Fortnightly"),
            ]
        )

        summary = import_sheet1_to_db(self.workbook_path, self.db_path)
        repository = SQLiteManualDispatchRepository(self.db_path)

        self.assertEqual(2, summary.rows_imported)
        self.assertEqual(1, summary.locations_inserted)
        self.assertEqual(2, summary.schedules_inserted)
        self.assertEqual(1, len(repository.list_opshop_locations()))
        self.assertEqual(2, len(repository.list_opshop_pickup_schedules()))

    def test_rerunning_import_does_not_create_duplicate_locations_or_schedules(self):
        self._save_workbook(
            [
                self._row(Run_Day="Monday", Run_Type="Standard", Pickup_Frequency="Weekly"),
                self._row(Run_Day="Wednesday", Run_Type="Regular", Pickup_Frequency="Fortnightly"),
            ]
        )

        first_summary = import_sheet1_to_db(self.workbook_path, self.db_path)
        second_summary = import_sheet1_to_db(self.workbook_path, self.db_path)
        repository = SQLiteManualDispatchRepository(self.db_path)

        self.assertEqual(1, first_summary.locations_inserted)
        self.assertEqual(2, first_summary.schedules_inserted)
        self.assertEqual(1, second_summary.locations_updated)
        self.assertEqual(2, second_summary.schedules_updated)
        self.assertEqual(1, len(repository.list_opshop_locations()))
        self.assertEqual(2, len(repository.list_opshop_pickup_schedules()))
        self.assertIsNotNone(second_summary.backup_path)
        self.assertTrue(Path(second_summary.backup_path).exists())

    def test_inactive_and_on_hold_rows_are_skipped(self):
        self._save_workbook(
            [
                self._row(Op_Shop_Name="Active Shop", Status="Active", Active_Flag="1"),
                self._row(Op_Shop_Name="On Hold Shop", Status="On_Hold", Active_Flag="1"),
                self._row(Op_Shop_Name="Inactive Shop", Status="Active", Active_Flag="0"),
            ]
        )

        summary = import_sheet1_to_db(self.workbook_path, self.db_path)
        repository = SQLiteManualDispatchRepository(self.db_path)

        self.assertEqual(3, summary.rows_read)
        self.assertEqual(1, summary.rows_imported)
        self.assertEqual(2, summary.rows_skipped_inactive)
        self.assertEqual(["Active Shop"], [location.name for location in repository.list_opshop_locations()])

    def test_on_call_blank_run_day_is_accepted(self):
        self._save_workbook(
            [
                self._row(
                    Run_Day="",
                    Run_Type="On_Call",
                    Pickup_Frequency="On Call",
                )
            ]
        )

        summary = import_sheet1_to_db(self.workbook_path, self.db_path)
        schedule = SQLiteManualDispatchRepository(self.db_path).list_opshop_pickup_schedules()[0]

        self.assertEqual(0, summary.review_required_count)
        self.assertIsNone(schedule.run_day)
        self.assertEqual("ON_CALL", schedule.run_type)
        self.assertFalse(schedule.review_required)

    def test_missing_standard_run_day_is_imported_but_review_required(self):
        self._save_workbook(
            [
                self._row(
                    Run_Day="",
                    Run_Type="Standard",
                    Pickup_Frequency="Weekly",
                )
            ]
        )

        summary = import_sheet1_to_db(self.workbook_path, self.db_path)
        schedule = SQLiteManualDispatchRepository(self.db_path).list_opshop_pickup_schedules()[0]

        self.assertEqual(1, summary.review_required_count)
        self.assertTrue(schedule.review_required)
        self.assertIn("Missing run_day", schedule.review_reason)

    def test_fortnightly_and_monthly_regular_rows_are_review_required(self):
        self._save_workbook(
            [
                self._row(
                    Run_Day="Monday",
                    Run_Type="Standard",
                    Pickup_Frequency="Fortnight",
                    Time_Window="Early",
                ),
                self._row(
                    Run_Day="Monday",
                    Run_Type="Regular",
                    Pickup_Frequency="Fortnightly",
                    Time_Window="Morning",
                ),
                self._row(
                    Run_Day="Tuesday",
                    Run_Type="Regular",
                    Pickup_Frequency="Monthly",
                    Time_Window="Afternoon",
                ),
            ]
        )

        summary = import_sheet1_to_db(self.workbook_path, self.db_path)
        schedules = SQLiteManualDispatchRepository(self.db_path).list_opshop_pickup_schedules()

        self.assertEqual(3, summary.review_required_count)
        self.assertTrue(all(schedule.review_required for schedule in schedules))
        self.assertIn(
            "Fortnightly schedule missing fortnight_group",
            summary.review_required_by_reason,
        )
        self.assertIn(
            "Monthly schedule requires review",
            summary.review_required_by_reason,
        )

    def test_weekly_like_and_multi_weekly_frequencies_are_not_review_required(self):
        rows = [
            self._row(Run_Type="Standard", Pickup_Frequency="2x Weekly"),
            self._row(Run_Type="Standard", Pickup_Frequency="2 X WEEKLY (WED/FRI)"),
            self._row(Run_Type="Standard", Pickup_Frequency="2x Weekly (Wed & Fri)"),
            self._row(Run_Type="Regular", Pickup_Frequency="Weekly (Thursday only)"),
            self._row(Run_Type="Regular", Pickup_Frequency="Weekly (Friday only)"),
            self._row(Run_Type="Regular", Pickup_Frequency="Weekly (Tuesday & Thursday)"),
            self._row(Run_Type="Regular", Pickup_Frequency="Twice weekly (Wed & Fri)"),
            self._row(Run_Day="", Run_Type="On_Call", Pickup_Frequency="On Call"),
        ]
        for index, row in enumerate(rows):
            row["Time_Window"] = f"Window {index}"
        self._save_workbook(rows)

        summary = import_sheet1_to_db(self.workbook_path, self.db_path)
        schedules = SQLiteManualDispatchRepository(self.db_path).list_opshop_pickup_schedules()

        self.assertEqual(8, summary.rows_imported)
        self.assertEqual(0, summary.review_required_count)
        self.assertTrue(all(not schedule.review_required for schedule in schedules))

    def test_blank_frequency_is_review_required(self):
        self._save_workbook([self._row(Pickup_Frequency="")])

        summary = import_sheet1_to_db(self.workbook_path, self.db_path)
        schedule = SQLiteManualDispatchRepository(self.db_path).list_opshop_pickup_schedules()[0]

        self.assertEqual(1, summary.review_required_count)
        self.assertTrue(schedule.review_required)
        self.assertIn("Blank or unknown pickup_frequency", schedule.review_reason)

    def test_phone_numbers_are_preserved_as_text(self):
        self._save_workbook(
            [
                self._row(
                    Primary_Phone="0400 123 456",
                    Secondary_Phone="03 9999 8888",
                )
            ]
        )

        import_sheet1_to_db(self.workbook_path, self.db_path)
        location = SQLiteManualDispatchRepository(self.db_path).list_opshop_locations()[0]

        self.assertEqual("0400 123 456", location.primary_phone)
        self.assertEqual("03 9999 8888", location.secondary_phone)

    def test_missing_required_columns_fails_before_writing_database(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        worksheet.append(["Op_Shop_Name", "Status"])
        worksheet.append(["Missing Columns Shop", "Active"])
        workbook.save(self.workbook_path)

        with self.assertRaises(ValueError):
            import_sheet1_to_db(self.workbook_path, self.db_path)

        self.assertFalse(self.db_path.exists())

    def _save_workbook(self, rows):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        worksheet.append(REQUIRED_COLUMNS)
        for row in rows:
            worksheet.append([row.get(column, "") for column in REQUIRED_COLUMNS])
        workbook.create_sheet("Recheck_Log")
        workbook.create_sheet("Recheck_Summary")
        workbook.create_sheet("Review_Items")
        workbook.save(self.workbook_path)

    def _row(self, **overrides):
        row = {
            "Op_Shop_Name": "Northside Op Shop",
            "Run_Day": "Monday",
            "Run_Type": "Standard",
            "Active_Flag": "1",
            "Suburb": "Coburg",
            "Street_Address": "1 Sydney Road",
            "Area_Region": "North",
            "Primary_Contact": "Mary",
            "Primary_Phone": "0400 000 001",
            "Secondary_Contact": "John",
            "Secondary_Phone": "0400 000 002",
            "Pickup_Frequency": "Weekly",
            "Time_Window": "9-12",
            "Call_Before_Arrival": "Yes",
            "Call_Timing": "30 minutes",
            "Access_Type": "Rear dock",
            "Key_Required": "No",
            "Trailer_Restriction": "Small truck only",
            "Status": "Active",
            "Status_Start_Date": "2026-05-19",
            "Status_Notes": "Ring first",
        }
        row.update(overrides)
        return row


if __name__ == "__main__":
    unittest.main()
