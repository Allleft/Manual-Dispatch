from io import BytesIO
import shutil
import unittest
import uuid
from unittest.mock import patch
from pathlib import Path

from openpyxl import load_workbook

from backend.repositories.sqlite_manual_dispatch_repository import (
    SQLiteManualDispatchRepository,
)
from backend.schemas import AssignTaskRequest, UpdateOrderRequest
from backend.services.excel_export_service import build_manual_dispatch_excel
from backend.services.manual_dispatch_service import ManualDispatchService


class ManualDispatchEditOrderTest(unittest.TestCase):
    def setUp(self):
        temp_parent = Path.cwd() / "tmp"
        temp_parent.mkdir(exist_ok=True)
        self.temp_dir = temp_parent / f"edit-order-test-{uuid.uuid4().hex}"
        self.temp_dir.mkdir()
        self.db_path = self.temp_dir / "manual_dispatch.sqlite3"
        with patch.dict("os.environ", {"MANUAL_DISPATCH_SEED_DEMO_DATA": "true"}):
            self.repository = SQLiteManualDispatchRepository(self.db_path)
        self.service = ManualDispatchService(self.repository)
        self.dispatch_date = "2026-05-05"

    def tearDown(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_update_order_updates_fields(self):
        updated = self.service.update_order(
            "ORD-001",
            self._request(
                invoice_number="INV-9001",
                order_no="ORD-NO-9001",
                company_name="Updated Company",
                phone="0499 111 222",
                suburb="Mulgrave",
                pallet_quantity=5,
                loose_bags_quantity=0,
                note="Updated delivery note",
            ),
        )

        self.assertEqual("INV-9001", updated.invoice_number)
        self.assertEqual("ORD-NO-9001", updated.order_no)
        self.assertEqual("Updated Company", updated.company_name)
        self.assertEqual("0499 111 222", updated.phone)
        self.assertEqual("Mulgrave", updated.suburb)
        self.assertEqual(5, updated.pallet_quantity)
        self.assertEqual(0, updated.loose_bags_quantity)
        self.assertEqual("Updated delivery note", updated.note)

    def test_update_order_rejects_missing_suburb(self):
        with self.assertRaises(ValueError):
            self.service.update_order("ORD-001", self._request(suburb=""))

    def test_update_order_rejects_negative_pallet_quantity(self):
        with self.assertRaises(ValueError):
            self.service.update_order("ORD-001", self._request(pallet_quantity=-1))

    def test_update_order_rejects_negative_loose_bags_quantity(self):
        with self.assertRaises(ValueError):
            self.service.update_order("ORD-001", self._request(loose_bags_quantity=-1))

    def test_update_order_allows_mixed_pallet_and_bag_quantities(self):
        updated = self.service.update_order(
            "ORD-001",
            self._request(pallet_quantity=5, loose_bags_quantity=2),
        )

        self.assertEqual(5, updated.pallet_quantity)
        self.assertEqual(2, updated.loose_bags_quantity)

    def test_update_pallet_order_preserves_carton_product_detail(self):
        updated = self.service.update_order(
            "ORD-001",
            self._request(
                pallet_quantity=1,
                loose_bags_quantity=0,
                product_lines=[
                    {
                        "product_name": "COLOUR RAGS 10KG NET",
                        "quantity": 1,
                        "unit": "PALLETS",
                    },
                    {
                        "product_name": "COLOR RAGS 1.5KG BAG",
                        "quantity": 2,
                        "unit": "CARTONS",
                    },
                ],
            ),
        )

        self.assertEqual(1, updated.pallet_quantity)
        self.assertEqual(0, updated.loose_bags_quantity)
        self.assertEqual(
            ["PALLETS", "CARTONS"],
            [line.unit for line in updated.product_lines],
        )

    def test_update_order_rejects_missing_delivery_date(self):
        with self.assertRaisesRegex(ValueError, "delivery_date is required"):
            self.service.update_order("ORD-001", self._request(delivery_date=""))

    def test_update_order_rejects_invalid_delivery_date(self):
        with self.assertRaisesRegex(ValueError, "delivery_date must"):
            self.service.update_order("ORD-001", self._request(delivery_date="2026-5-7"))

    def test_assigned_order_remains_assigned_after_edit(self):
        self._assign_order("ORD-001", "D001", "trip2")

        self.service.update_order("ORD-001", self._request(suburb="Mulgrave"))

        board = self.service.get_board(self.dispatch_date)
        assignment = board.assignments[0]
        self.assertEqual("ORD-001", assignment.task_id)
        self.assertEqual("D001", assignment.driver_id)
        self.assertEqual("trip2", assignment.trip_no)

    def test_updated_assigned_order_appears_in_board_response(self):
        self._assign_order("ORD-001", "D001", "trip1")

        self.service.update_order(
            "ORD-001",
            self._request(company_name="Board Updated", suburb="Keysborough"),
        )

        board = self.service.get_board(self.dispatch_date)
        updated = next(order for order in board.orders if order.order_id == "ORD-001")
        self.assertEqual("Board Updated", updated.company_name)
        self.assertEqual("Keysborough", updated.suburb)

    def test_updated_assigned_order_appears_in_excel_export(self):
        self._assign_order("ORD-001", "D001", "trip1")

        self.service.update_order(
            "ORD-001",
            self._request(
                company_name="Excel Updated",
                suburb="Heatherton",
                pallet_quantity=7,
                note="Excel sees this note",
            ),
        )

        rows = self._export_rows()
        data_rows = rows[1:]

        self.assertEqual(1, len(data_rows))
        self.assertEqual("Excel Updated", data_rows[0][5])
        self.assertEqual("Heatherton", data_rows[0][7])
        self.assertEqual(7, data_rows[0][12])
        self.assertEqual("Excel sees this note", data_rows[0][16])

    def test_delivery_date_can_be_modified_through_edit_service(self):
        updated = self.service.update_order(
            "ORD-001",
            self._request(suburb="Date Safe", delivery_date="2026-06-06"),
        )

        reloaded = self.repository.get_order("ORD-001")

        self.assertEqual("2026-06-06", updated.delivery_date)
        self.assertEqual("2026-06-06", reloaded.delivery_date)

    def test_assigned_order_keeps_assignment_after_delivery_date_change(self):
        self._assign_order("ORD-001", "D001", "trip2")

        updated = self.service.update_order(
            "ORD-001",
            self._request(delivery_date="2026-05-06"),
        )
        board = self.service.get_board(self.dispatch_date)

        self.assertEqual("2026-05-06", updated.delivery_date)
        self.assertEqual("ORD-001", board.assignments[0].task_id)
        self.assertEqual("trip2", board.assignments[0].trip_no)
        self.assertEqual(
            "2026-05-06",
            next(order for order in board.orders if order.order_id == "ORD-001").delivery_date,
        )

    def _assign_order(self, order_id, driver_id, trip_no):
        self.service.assign_task(
            AssignTaskRequest(
                dispatch_date=self.dispatch_date,
                task_type="ORDER",
                task_id=order_id,
                driver_id=driver_id,
                trip_no=trip_no,
            )
        )

    def _request(self, **overrides):
        values = {
            "invoice_number": "INV-1001",
            "order_no": "ORD-NO-1001",
            "company_name": "Demo Customer A",
            "phone": "0400 000 001",
            "delivery_address": "1 Demo Street",
            "suburb": "Dandenong",
            "postcode": "3175",
            "delivery_date": self.dispatch_date,
            "zone": "South East",
            "urgency": "Normal",
            "preferred_driver_id": "D001",
            "pallet_quantity": 2,
            "loose_bags_quantity": 0,
            "start_time": "08:00",
            "end_time": "12:00",
            "note": "Call before delivery",
        }
        values.update(overrides)
        return UpdateOrderRequest(**values)

    def _export_rows(self):
        board = self.service.get_board(self.dispatch_date)
        excel_bytes = build_manual_dispatch_excel(board, self.dispatch_date)
        workbook = load_workbook(BytesIO(excel_bytes))
        worksheet = workbook.active
        return list(worksheet.iter_rows(values_only=True))


if __name__ == "__main__":
    unittest.main()
