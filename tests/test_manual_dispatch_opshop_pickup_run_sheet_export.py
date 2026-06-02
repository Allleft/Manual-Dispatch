from io import BytesIO
import importlib
import os
import shutil
import unittest
import uuid
from pathlib import Path

from openpyxl import load_workbook

from backend.repositories.in_memory_manual_dispatch_repository import (
    InMemoryManualDispatchRepository,
)
from backend.repositories.sqlite_manual_dispatch_repository import (
    SQLiteManualDispatchRepository,
)
from backend.schemas import (
    AssignTaskRequest,
    FinalTripSummary,
    FinalTripSummaryOrderSnapshot,
    FinalTripSummaryTrip,
    OpShopCountrysideRouteGroup,
    OpShopLocation,
    OpShopPickupSchedule,
)
from backend.services.final_summary_excel_export_service import build_final_summary_excel
from backend.services.manual_dispatch_service import ManualDispatchService
from backend.services.opshop_pickup_excel_export_service import (
    COUNTRYSIDE_SECTION_TITLE,
    ONCALL_SECTION_TITLE,
    OPSHOP_RUN_SHEET_HEADERS,
    REGULAR_SECTION_TITLE,
    UNASSIGNED_DRIVER_LABEL,
    UNASSIGNED_SECTION_TITLE,
    build_opshop_pickup_run_sheet_excel,
)

try:
    from fastapi import FastAPI
    from fastapi.testclient import TestClient
except (ImportError, ModuleNotFoundError, RuntimeError):
    FastAPI = None
    TestClient = None


class OpShopPickupRunSheetExportTest(unittest.TestCase):
    def setUp(self):
        self.repository = InMemoryManualDispatchRepository()
        self.service = ManualDispatchService(self.repository)
        self.dispatch_date = "2026-05-18"
        self._seed_opshop_data()

    def test_export_includes_regular_oncall_assigned_and_unassigned_pickups(self):
        regular = self.service.create_opshop_pickup_task(
            self._request("SCHED-REGULAR", "2026-05-18")
        )
        oncall = self.service.create_oncall_opshop_pickup_task(
            self._request("SCHED-ONCALL", "2026-05-20")
        )
        unassigned = self.service.create_opshop_pickup_task(
            self._request("SCHED-UNASSIGNED", "2026-05-19")
        )
        self._assign(regular.pickup_task_id, "D001")
        self._assign(oncall.pickup_task_id, "D002")
        self._assign_order("ORD-001")

        workbook = self._export_workbook()
        values = self._flat_values(workbook)

        self.assertIn(tuple(OPSHOP_RUN_SHEET_HEADERS), list(workbook.active.iter_rows(values_only=True)))
        self.assertIn(REGULAR_SECTION_TITLE, values)
        self.assertIn(ONCALL_SECTION_TITLE, values)
        self.assertIn(COUNTRYSIDE_SECTION_TITLE, values)
        self.assertIn("John", values)
        self.assertIn("Tony", values)
        self.assertIn(UNASSIGNED_DRIVER_LABEL, values)
        self.assertNotIn(UNASSIGNED_SECTION_TITLE, values)
        self.assertIn("Regular Aid Shop", values)
        self.assertIn("Oncall Help Shop", values)
        self.assertIn("Unassigned Goods Shop", values)
        self.assertIn("REGULAR", values)
        self.assertIn("ON_CALL", values)
        self.assertIn("ACTIVE", values)
        self.assertIn("ASSIGNED", values)
        self.assertIn("1 Charity Road", values)
        self.assertIn("0400 700 001", values)
        self.assertIn("Status: Regular status note\nTask: Regular task note", values)
        self.assertNotIn("Demo Customer A", values)
        self.assertIn(unassigned.pickup_task_id, [item.pickup_task_id for item in self.service.get_board(self.dispatch_date).scheduled_opshop_pickups])

    def test_export_groups_countryside_pickups_by_route_group(self):
        self._seed_countryside_data()
        regular = self.service.create_opshop_pickup_task(
            self._request("SCHED-REGULAR", "2026-05-18")
        )
        oncall = self.service.create_oncall_opshop_pickup_task(
            self._request("SCHED-ONCALL", "2026-05-20")
        )
        albury = self.service.create_oncall_opshop_pickup_task(
            self._request("SCHED-COUNTRY-ALBURY", "2026-05-22")
        )
        bendigo = self.service.create_oncall_opshop_pickup_task(
            self._request("SCHED-COUNTRY-BENDIGO", "2026-05-22")
        )
        self._assign(regular.pickup_task_id, "D001")
        self._assign(oncall.pickup_task_id, "D002")
        self._assign(albury.pickup_task_id, "D001")

        workbook = self._export_workbook()
        values = self._flat_values(workbook)
        oncall_values = self._section_values(workbook, ONCALL_SECTION_TITLE, COUNTRYSIDE_SECTION_TITLE)
        countryside_values = self._section_values(workbook, COUNTRYSIDE_SECTION_TITLE)

        self.assertIn("Route Group: ALBURY OPSHOP", values)
        self.assertIn("Route Group: BENDIGO & ECHUCA RUN", values)
        self.assertLess(
            values.index("Route Group: ALBURY OPSHOP"),
            values.index("Route Group: BENDIGO & ECHUCA RUN"),
        )
        self.assertIn("ALBURY OPSHOP", countryside_values)
        self.assertIn("BENDIGO & ECHUCA RUN", countryside_values)
        self.assertIn("Albury Remote Shop", countryside_values)
        self.assertIn("Bendigo Country Shop", countryside_values)
        self.assertIn("2026-05-22", countryside_values)
        self.assertIn("John", countryside_values)
        self.assertIn(UNASSIGNED_DRIVER_LABEL, countryside_values)
        self.assertNotIn("Oncall Help Shop", countryside_values)
        self.assertNotIn("Regular Aid Shop", countryside_values)
        self.assertNotIn("Albury Remote Shop", oncall_values)
        self.assertNotIn("Bendigo Country Shop", oncall_values)

    def test_export_does_not_modify_opshop_state(self):
        self._seed_countryside_data()
        task = self.service.create_oncall_opshop_pickup_task(
            self._request("SCHED-COUNTRY-ALBURY", "2026-05-22")
        )
        self._assign(task.pickup_task_id, "D001")
        board = self.service.get_board(self.dispatch_date)
        task_count = len(self.repository.list_opshop_pickup_tasks())
        assignment_count = len(self.repository.list_assignments(self.dispatch_date))

        load_workbook(
            BytesIO(
                build_opshop_pickup_run_sheet_excel(
                    board,
                    self.dispatch_date,
                )
            )
        )

        self.assertEqual(task_count, len(self.repository.list_opshop_pickup_tasks()))
        self.assertEqual(assignment_count, len(self.repository.list_assignments(self.dispatch_date)))

    def test_export_includes_assigned_countryside_task_when_route_group_inactive(self):
        self._seed_countryside_data()
        task = self.service.create_oncall_opshop_pickup_task(
            self._request("SCHED-COUNTRY-ALBURY", "2026-05-22")
        )
        self._assign(task.pickup_task_id, "D001")
        self.repository.disable_countryside_route_group("ROUTE-ALBURY")

        workbook = self._export_workbook()
        values = self._flat_values(workbook)
        countryside_values = self._section_values(workbook, COUNTRYSIDE_SECTION_TITLE)

        self.assertIn("Route Group: ALBURY OPSHOP", values)
        self.assertIn("Albury Remote Shop", countryside_values)
        self.assertIn("John", countryside_values)

    def test_final_summary_export_stays_order_only(self):
        regular = self.service.create_opshop_pickup_task(
            self._request("SCHED-REGULAR", "2026-05-18")
        )
        self._assign(regular.pickup_task_id, "D001")
        summary = FinalTripSummary(
            summary_id="FTS-TEST",
            dispatch_date=self.dispatch_date,
            delivery_date="2026-05-05",
            driver_id="D001",
            driver_name_snapshot="John",
            vehicle_id=None,
            vehicle_rego_snapshot=None,
            total_pallets=2,
            total_loose_bags=0,
            status="SAVED",
            generated_at="2026-05-18T00:00:00+00:00",
            saved_at="2026-05-18T00:00:00+00:00",
            saved_by_account_name="Mandy",
            saved_by_account_id=None,
            trips=[
                FinalTripSummaryTrip(
                    trip_no="trip1",
                    orders=[
                        FinalTripSummaryOrderSnapshot(
                            row_id="FSR-TEST",
                            trip_no="trip1",
                            row_no=1,
                            task_type="ORDER",
                            task_id="ORD-001",
                            order_id_snapshot="ORD-001",
                            invoice_number_snapshot="INV-1001",
                            company_name_snapshot="Demo Customer A",
                            suburb_snapshot="Dandenong",
                            delivery_address_snapshot="1 Demo Street",
                            product_snapshot=None,
                            pallet_quantity_snapshot=2,
                            loose_bags_quantity_snapshot=0,
                            note_snapshot=None,
                        )
                    ],
                )
            ],
        )

        workbook = load_workbook(BytesIO(build_final_summary_excel([summary], self.dispatch_date)))
        values = self._flat_values(workbook)

        self.assertIn("Demo Customer A", values)
        self.assertNotIn("Regular Aid Shop", values)
        self.assertNotIn("OP SHOP PICKUP", values)

    def _export_workbook(self):
        return load_workbook(
            BytesIO(
                build_opshop_pickup_run_sheet_excel(
                    self.service.get_board(self.dispatch_date),
                    self.dispatch_date,
                )
            )
        )

    def _seed_opshop_data(self):
        self.repository.upsert_opshop_location(
            self._location(
                "OPSHOP-REGULAR",
                "Regular Aid Shop",
                "1 Charity Road",
                "Coburg",
                "0400 700 001",
                "Regular status note",
            )
        )
        self.repository.upsert_opshop_location(
            self._location(
                "OPSHOP-ONCALL",
                "Oncall Help Shop",
                "2 Donation Lane",
                "Preston",
                "0400 700 002",
                "Oncall status note",
            )
        )
        self.repository.upsert_opshop_location(
            self._location(
                "OPSHOP-UNASSIGNED",
                "Unassigned Goods Shop",
                "3 Goods Street",
                "Brunswick",
                "0400 700 003",
                "Unassigned status note",
            )
        )
        self.repository.upsert_opshop_pickup_schedule(
            self._schedule("SCHED-REGULAR", "OPSHOP-REGULAR", "REGULAR", "MONDAY")
        )
        self.repository.upsert_opshop_pickup_schedule(
            self._schedule("SCHED-ONCALL", "OPSHOP-ONCALL", "ON_CALL", "WEDNESDAY")
        )
        self.repository.upsert_opshop_pickup_schedule(
            self._schedule("SCHED-UNASSIGNED", "OPSHOP-UNASSIGNED", "REGULAR", "TUESDAY")
        )

    def _seed_countryside_data(self):
        self.repository.upsert_countryside_route_group(
            self._route_group("ROUTE-BENDIGO", "BENDIGO & ECHUCA RUN", 2)
        )
        self.repository.upsert_countryside_route_group(
            self._route_group("ROUTE-ALBURY", "ALBURY OPSHOP", 1)
        )
        self.repository.upsert_opshop_location(
            self._location(
                "OPSHOP-COUNTRY-BENDIGO",
                "Bendigo Country Shop",
                "4 Country Road",
                "Bendigo",
                "0400 700 004",
                "Bendigo status note",
            )
        )
        self.repository.upsert_opshop_location(
            self._location(
                "OPSHOP-COUNTRY-ALBURY",
                "Albury Remote Shop",
                "5 Border Road",
                "Albury",
                "0400 700 005",
                "Albury status note",
            )
        )
        self.repository.upsert_opshop_pickup_schedule(
            self._schedule(
                "SCHED-COUNTRY-BENDIGO",
                "OPSHOP-COUNTRY-BENDIGO",
                "ON_CALL",
                None,
                pickup_category="COUNTRYSIDE",
                route_group_id="ROUTE-BENDIGO",
            )
        )
        self.repository.upsert_opshop_pickup_schedule(
            self._schedule(
                "SCHED-COUNTRY-ALBURY",
                "OPSHOP-COUNTRY-ALBURY",
                "ON_CALL",
                None,
                pickup_category="COUNTRYSIDE",
                route_group_id="ROUTE-ALBURY",
            )
        )

    def _route_group(self, route_group_id, route_group_name, display_order):
        return OpShopCountrysideRouteGroup(
            route_group_id=route_group_id,
            route_group_name=route_group_name,
            status="Active",
            active_flag=True,
            display_order=display_order,
            source_marker="TEST",
            created_at="2026-05-18T00:00:00+00:00",
            updated_at="2026-05-18T00:00:00+00:00",
        )

    def _location(self, opshop_id, name, address, suburb, phone, status_notes):
        return OpShopLocation(
            opshop_id=opshop_id,
            name=name,
            suburb=suburb,
            street_address=address,
            area_region="North",
            primary_contact="Primary Contact",
            primary_phone=phone,
            secondary_contact="Secondary Contact",
            secondary_phone="0400 800 001",
            access_type="Front door",
            key_required=True,
            trailer_restriction="No trailer",
            status_notes=status_notes,
            is_active=True,
            created_at="2026-05-18T00:00:00+00:00",
            updated_at="2026-05-18T00:00:00+00:00",
        )

    def _schedule(
        self,
        schedule_id,
        opshop_id,
        run_type,
        run_day,
        pickup_category="NORMAL",
        route_group_id=None,
    ):
        return OpShopPickupSchedule(
            schedule_id=schedule_id,
            opshop_id=opshop_id,
            run_day=run_day,
            run_type=run_type,
            pickup_frequency="Weekly",
            time_window="9am-12pm",
            call_before_arrival=True,
            call_timing="30 minutes",
            status="Active",
            active_flag=True,
            fortnight_group=None,
            review_required=False,
            review_reason=None,
            created_at="2026-05-18T00:00:00+00:00",
            updated_at="2026-05-18T00:00:00+00:00",
            pickup_category=pickup_category,
            route_group_id=route_group_id,
        )

    def _request(self, schedule_id, pickup_date):
        from backend.schemas import CreateOpShopPickupTaskRequest

        return CreateOpShopPickupTaskRequest(
            schedule_id=schedule_id,
            pickup_date=pickup_date,
            notes="Regular task note" if schedule_id == "SCHED-REGULAR" else None,
        )

    def _assign(self, pickup_task_id, driver_id):
        self.service.assign_task(
            AssignTaskRequest(
                dispatch_date=self.dispatch_date,
                task_type="OPSHOP_PICKUP",
                task_id=pickup_task_id,
                driver_id=driver_id,
                trip_no="trip1",
            )
        )

    def _assign_order(self, order_id):
        self.service.assign_task(
            AssignTaskRequest(
                dispatch_date=self.dispatch_date,
                task_type="ORDER",
                task_id=order_id,
                driver_id="D001",
                trip_no="trip1",
            )
        )

    def _flat_values(self, workbook):
        values = []
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                values.extend(cell for cell in row if cell not in (None, ""))
        return values

    def _section_values(self, workbook, start_marker, end_marker=None):
        values = self._flat_values(workbook)
        start = values.index(start_marker)
        end = values.index(end_marker) if end_marker else len(values)
        return values[start:end]


@unittest.skipIf(FastAPI is None or TestClient is None, "FastAPI TestClient is not installed")
class OpShopPickupRunSheetExportRouteTest(unittest.TestCase):
    def setUp(self):
        temp_parent = Path.cwd() / "tmp"
        temp_parent.mkdir(exist_ok=True)
        self.temp_dir = temp_parent / f"opshop-export-route-test-{uuid.uuid4().hex}"
        self.temp_dir.mkdir()
        self.db_path = self.temp_dir / "manual_dispatch.sqlite3"
        self.previous_db_path = os.environ.get("MANUAL_DISPATCH_DB_PATH")
        os.environ["MANUAL_DISPATCH_DB_PATH"] = str(self.db_path)

        self.repository = SQLiteManualDispatchRepository(self.db_path)
        self.service = ManualDispatchService(self.repository)
        self.api_module = importlib.import_module("backend.api.manual_dispatch")
        self.original_service = self.api_module.service
        self.api_module.service = self.service

        app = FastAPI()
        app.include_router(self.api_module.router)
        self.client = TestClient(app)

    def tearDown(self):
        self.api_module.service = self.original_service
        if self.previous_db_path is None:
            os.environ.pop("MANUAL_DISPATCH_DB_PATH", None)
        else:
            os.environ["MANUAL_DISPATCH_DB_PATH"] = self.previous_db_path
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_export_endpoint_returns_xlsx(self):
        response = self.client.get(
            "/api/manual-dispatch/opshop-pickups/export-excel",
            params={"dispatch_date": "2026-05-18"},
        )

        self.assertEqual(200, response.status_code)
        self.assertEqual(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.headers["content-type"],
        )
        self.assertIn(
            'attachment; filename="opshop-pickup-run-sheet-2026-05-18.xlsx"',
            response.headers["content-disposition"],
        )
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual("OP SHOP Run Sheet", workbook.active.title)


if __name__ == "__main__":
    unittest.main()
