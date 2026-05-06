from io import BytesIO
import shutil
import unittest
import uuid
from pathlib import Path

from openpyxl import load_workbook

from backend.repositories.sqlite_manual_dispatch_repository import (
    SQLiteManualDispatchRepository,
)
from backend.schemas import AssignTaskRequest, CreateOrderRequest
from backend.services.excel_export_service import build_manual_dispatch_excel
from backend.services.manual_dispatch_service import ManualDispatchService


class ManualDispatchCreateOrderTest(unittest.TestCase):
    def setUp(self):
        temp_parent = Path.cwd() / "tmp"
        temp_parent.mkdir(exist_ok=True)
        self.temp_dir = temp_parent / f"create-order-test-{uuid.uuid4().hex}"
        self.temp_dir.mkdir()
        self.db_path = self.temp_dir / "manual_dispatch.sqlite3"
        self.repository = SQLiteManualDispatchRepository(self.db_path)
        self.service = ManualDispatchService(self.repository)
        self.dispatch_date = "2026-05-05"

    def tearDown(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_create_order_saves_new_order(self):
        order = self.service.create_order(self._request())

        self.assertEqual("ORD-20260505-001", order.order_id)
        self.assertEqual("INV-2001", order.invoice_number)
        self.assertEqual("0400 000 999", order.phone)
        self.assertEqual("Noble Park", order.suburb)
        self.assertEqual("Normal", order.urgency)

    def test_created_order_appears_in_board_for_delivery_date(self):
        created = self.service.create_order(self._request())

        board = self.service.get_board(self.dispatch_date)

        self.assertIn(created.order_id, [order.order_id for order in board.orders])

    def test_created_order_is_unassigned_by_default(self):
        created = self.service.create_order(self._request())

        board = self.service.get_board(self.dispatch_date)

        self.assertNotIn(created.order_id, [assignment.task_id for assignment in board.assignments])

    def test_created_order_can_be_assigned_after_creation(self):
        created = self.service.create_order(self._request())

        assignment = self.service.assign_task(
            AssignTaskRequest(
                dispatch_date=self.dispatch_date,
                task_type="ORDER",
                task_id=created.order_id,
                driver_id="D001",
                trip_no="trip1",
            )
        )

        self.assertEqual(created.order_id, assignment.task_id)
        self.assertEqual("D001", assignment.driver_id)

    def test_negative_pallet_quantity_is_rejected(self):
        with self.assertRaises(ValueError):
            self.service.create_order(self._request(pallet_quantity=-1))

    def test_negative_loose_bags_quantity_is_rejected(self):
        with self.assertRaises(ValueError):
            self.service.create_order(self._request(loose_bags_quantity=-1))

    def test_missing_suburb_is_rejected(self):
        with self.assertRaises(ValueError):
            self.service.create_order(self._request(suburb=""))

    def test_missing_delivery_date_is_rejected(self):
        with self.assertRaises(ValueError):
            self.service.create_order(self._request(delivery_date=""))

    def test_created_order_is_included_in_excel_export_after_assignment(self):
        created = self.service.create_order(self._request())
        self.service.assign_task(
            AssignTaskRequest(
                dispatch_date=self.dispatch_date,
                task_type="ORDER",
                task_id=created.order_id,
                driver_id="D001",
                trip_no="trip2",
            )
        )

        rows = self._export_rows()
        data_rows = rows[1:]

        self.assertIn(created.order_id, [row[4] for row in data_rows])

    def test_created_unassigned_order_is_not_included_in_excel_export(self):
        created = self.service.create_order(self._request())

        rows = self._export_rows()
        data_rows = rows[1:]

        self.assertNotIn(created.order_id, [row[4] for row in data_rows])

    def _request(self, **overrides):
        values = {
            "invoice_number": "INV-2001",
            "company_name": "New Customer",
            "phone": "0400 000 999",
            "delivery_address": "10 Manual Way",
            "suburb": "Noble Park",
            "postcode": "3174",
            "delivery_date": self.dispatch_date,
            "zone": "South East",
            "urgency": "",
            "preferred_driver_id": "",
            "pallet_quantity": 2,
            "loose_bags_quantity": 1,
            "start_time": "09:00",
            "end_time": "13:00",
            "note": "Created from Add Order flow",
        }
        values.update(overrides)
        return CreateOrderRequest(**values)

    def _export_rows(self):
        board = self.service.get_board(self.dispatch_date)
        excel_bytes = build_manual_dispatch_excel(board, self.dispatch_date)
        workbook = load_workbook(BytesIO(excel_bytes))
        worksheet = workbook.active
        return list(worksheet.iter_rows(values_only=True))


if __name__ == "__main__":
    unittest.main()
