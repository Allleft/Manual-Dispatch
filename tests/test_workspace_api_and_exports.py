import importlib
import os
import shutil
import sqlite3
import unittest
import uuid
from unittest.mock import patch
from dataclasses import replace
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from backend.repositories.sqlite_manual_dispatch_repository import (
    SQLiteManualDispatchRepository,
)
from backend.schemas import (
    AssignTaskRequest,
    DeliveryWorkspaceVehicleAssignmentRequest,
    DeliveryRunSheetOrderSnapshot,
    OpShopLocation,
    OpShopPickupSchedule,
    OpShopPickupTask,
    ProductDetailLine,
    RegisterOperatorAccountRequest,
)
from backend.services.manual_dispatch_service import ManualDispatchService
from tests.manual_dispatch_api_test_helpers import authenticate_test_client
from backend.services.delivery_run_sheet_excel_export_service import (
    build_delivery_run_sheets_excel,
    delivery_run_sheet_product_display,
)

try:
    from fastapi import FastAPI
    from fastapi.testclient import TestClient
except (ImportError, ModuleNotFoundError, RuntimeError):
    FastAPI = None
    TestClient = None


@unittest.skipIf(FastAPI is None or TestClient is None, "FastAPI TestClient unavailable")
class WorkspaceApiAndExportsTest(unittest.TestCase):
    def setUp(self):
        temp_parent = Path.cwd() / "tmp"
        temp_parent.mkdir(exist_ok=True)
        self.temp_dir = temp_parent / f"workspace-api-test-{uuid.uuid4().hex}"
        self.temp_dir.mkdir()
        self.db_path = self.temp_dir / "manual_dispatch.sqlite3"
        self.previous_db_path = os.environ.get("MANUAL_DISPATCH_DB_PATH")
        os.environ["MANUAL_DISPATCH_DB_PATH"] = str(self.db_path)

        with patch.dict("os.environ", {"MANUAL_DISPATCH_SEED_DEMO_DATA": "true"}):
            self.repository = SQLiteManualDispatchRepository(self.db_path)
        self.service = ManualDispatchService(self.repository)
        self.account = self.service.register_operator_account(
            RegisterOperatorAccountRequest(
                account_name="Workspace API Tester",
                password="secret123",
                confirm_password="secret123",
            )
        )
        self.dispatch_date = "2026-05-05"
        self._assign_order()
        self._seed_and_assign_pickup()

        self.api_module = importlib.import_module("backend.api.manual_dispatch")
        self.original_service = self.api_module.service
        self.api_module.service = self.service
        app = FastAPI()
        app.include_router(self.api_module.router)
        self.client = TestClient(app)
        authenticate_test_client(self.client, self.service, self.account)

    def tearDown(self):
        self.api_module.service = self.original_service
        if self.previous_db_path is None:
            os.environ.pop("MANUAL_DISPATCH_DB_PATH", None)
        else:
            os.environ["MANUAL_DISPATCH_DB_PATH"] = self.previous_db_path
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_independent_routes_generate_list_get_save_cancel_and_export(self):
        expected_routes = {
            ("GET", "/api/manual-dispatch/delivery/specifications"),
            ("POST", "/api/manual-dispatch/delivery/drivers"),
            ("PATCH", "/api/manual-dispatch/delivery/drivers/{driver_id}"),
            ("DELETE", "/api/manual-dispatch/delivery/drivers/{driver_id}"),
            ("POST", "/api/manual-dispatch/delivery/vehicles"),
            ("PATCH", "/api/manual-dispatch/delivery/vehicles/{vehicle_id}"),
            ("DELETE", "/api/manual-dispatch/delivery/vehicles/{vehicle_id}"),
            ("POST", "/api/manual-dispatch/delivery/orders"),
            ("PATCH", "/api/manual-dispatch/delivery/orders/{order_id}"),
            ("POST", "/api/manual-dispatch/delivery/orders/{order_id}/cancel"),
            ("POST", "/api/manual-dispatch/delivery/orders/import-attache-pdf-preview"),
            ("POST", "/api/manual-dispatch/delivery/orders/import-attache-pdf-commit"),
            ("POST", "/api/manual-dispatch/delivery/run-sheets/generated"),
            ("GET", "/api/manual-dispatch/delivery/trip-summary"),
            ("GET", "/api/manual-dispatch/delivery/run-sheets"),
            ("GET", "/api/manual-dispatch/delivery/run-sheets/export-excel"),
            ("GET", "/api/manual-dispatch/delivery/run-sheets/{run_sheet_id}"),
            ("POST", "/api/manual-dispatch/delivery/run-sheets/{run_sheet_id}/save"),
            (
                "POST",
                "/api/manual-dispatch/delivery/run-sheets/{run_sheet_id}/closeout",
            ),
            (
                "POST",
                "/api/manual-dispatch/delivery/run-sheets/{run_sheet_id}/cancel-generated",
            ),
            (
                "GET",
                "/api/manual-dispatch/delivery/run-sheets/{run_sheet_id}/export-excel",
            ),
            ("POST", "/api/manual-dispatch/opshop/pickup-collections/generated"),
            ("GET", "/api/manual-dispatch/opshop/trip-summary"),
            ("GET", "/api/manual-dispatch/opshop/pickup-collections"),
            (
                "GET",
                "/api/manual-dispatch/opshop/pickup-collections/export-excel",
            ),
            (
                "GET",
                "/api/manual-dispatch/opshop/pickup-collections/{collection_id}",
            ),
            (
                "PATCH",
                "/api/manual-dispatch/opshop/pickup-collections/{collection_id}/rows",
            ),
            (
                "POST",
                "/api/manual-dispatch/opshop/pickup-collections/{collection_id}/save",
            ),
            (
                "POST",
                "/api/manual-dispatch/opshop/pickup-collections/{collection_id}/cancel-generated",
            ),
            (
                "GET",
                "/api/manual-dispatch/opshop/pickup-collections/{collection_id}/export-excel",
            ),
        }
        actual_routes = {
            (method, route.path)
            for route in self.api_module.router.routes
            for method in route.methods
        }
        self.assertFalse(expected_routes - actual_routes)

        delivery = self.client.post(
            "/api/manual-dispatch/delivery/run-sheets/generated",
            json=self._delivery_generate_payload(),
        )
        self.assertEqual(200, delivery.status_code)
        run_sheet_id = delivery.json()["run_sheet_id"]
        self.assertEqual(
            200,
            self.client.get(
                f"/api/manual-dispatch/delivery/run-sheets/{run_sheet_id}"
            ).status_code,
        )
        self.assertEqual(
            1,
            len(
                self.client.get(
                    "/api/manual-dispatch/delivery/run-sheets",
                    params={"dispatch_date": self.dispatch_date},
                ).json()
            ),
        )

        cancelled_delivery = self.client.post(
            f"/api/manual-dispatch/delivery/run-sheets/{run_sheet_id}/cancel-generated"
        )
        self.assertEqual({"cancelled": True}, cancelled_delivery.json())

        saved_run_sheet_id = self._generate_and_save_delivery()
        delivery_export = self.client.get(
            f"/api/manual-dispatch/delivery/run-sheets/{saved_run_sheet_id}/export-excel"
        )
        self.assertEqual(200, delivery_export.status_code)
        self.assertIn("Delivery_Run_Sheet_", delivery_export.headers["content-disposition"])

        collection = self.client.post(
            "/api/manual-dispatch/opshop/pickup-collections/generated",
            json=self._collection_generate_payload(),
        )
        self.assertEqual(200, collection.status_code)
        collection_id = collection.json()["collection_id"]
        cancelled_collection = self.client.post(
            f"/api/manual-dispatch/opshop/pickup-collections/{collection_id}/cancel-generated"
        )
        self.assertEqual({"cancelled": True}, cancelled_collection.json())

        saved_collection_id = self._generate_and_save_collection(
            self._weight_sheet_entry_payload()
        )
        collection_export = self.client.get(
            f"/api/manual-dispatch/opshop/pickup-collections/{saved_collection_id}/export-excel"
        )
        self.assertEqual(200, collection_export.status_code)
        self.assertIn(
            "OPSHOP_Pickup_Collection_",
            collection_export.headers["content-disposition"],
        )

        self._seed_assigned_pickup(
            task_id="PICKUP-GENERATED-SINGLE",
            driver_id="D001",
            pickup_date="2026-05-06",
            opshop_id="OPSHOP-GENERATED-SINGLE",
            schedule_id="SCHEDULE-GENERATED-SINGLE",
            opshop_name="Generated Single Op Shop",
            suburb="Northcote",
        )
        generated_collection = self.client.post(
            "/api/manual-dispatch/opshop/pickup-collections/generated",
            json=self._collection_generate_payload("D001", "2026-05-06"),
        )
        self.assertEqual(200, generated_collection.status_code)
        generated_export = self.client.get(
            "/api/manual-dispatch/opshop/pickup-collections/"
            f"{generated_collection.json()['collection_id']}/export-excel"
        )
        self.assertEqual(200, generated_export.status_code)
        self.assertIn(
            "OPSHOP_Pickup_Collection_2026-05-06_John.xlsx",
            generated_export.headers["content-disposition"],
        )


    def test_snapshot_saves_ignore_spoofed_client_actor(self):
        delivery = self.client.post(
            "/api/manual-dispatch/delivery/run-sheets/generated",
            json=self._delivery_generate_payload(),
        )
        collection = self.client.post(
            "/api/manual-dispatch/opshop/pickup-collections/generated",
            json=self._collection_generate_payload(),
        )
        spoofed_actor = {
            "saved_by_account_name": "Spoofed Operator",
            "saved_by_account_id": -999,
        }

        saved_delivery = self.client.post(
            f"/api/manual-dispatch/delivery/run-sheets/{delivery.json()['run_sheet_id']}/save",
            json=spoofed_actor,
        )
        saved_collection = self.client.post(
            f"/api/manual-dispatch/opshop/pickup-collections/{collection.json()['collection_id']}/save",
            json=spoofed_actor,
        )

        self.assertEqual(200, saved_delivery.status_code)
        self.assertEqual(200, saved_collection.status_code)
        for response in (saved_delivery, saved_collection):
            self.assertEqual(
                self.account.account_name,
                response.json()["saved_by_account_name"],
            )
            self.assertEqual(
                self.account.account_id,
                response.json()["saved_by_account_id"],
            )

    def test_opshop_collection_weight_sheet_patch_is_typed_atomic_and_generated_only(self):
        self._seed_assigned_pickup(
            task_id="PICKUP-WORKSPACE-2",
            driver_id="D001",
            pickup_date=self.dispatch_date,
            opshop_id="OPSHOP-WORKSPACE-2",
            schedule_id="SCHEDULE-WORKSPACE-2",
            opshop_name="Southside Op Shop",
            suburb="Richmond",
        )
        generated = self.client.post(
            "/api/manual-dispatch/opshop/pickup-collections/generated",
            json=self._collection_generate_payload(),
        )
        self.assertEqual(200, generated.status_code)
        collection_id = generated.json()["collection_id"]
        rows = generated.json()["pickups"]
        self.assertEqual(2, len(rows))

        first_payload = {
            "row_id": rows[0]["row_id"],
            "clothing_kg": 12.25,
            "shoes_kg": 3,
            "time_in": "09:05",
            "time_out": "09:45",
            "trolleys_out_to_opshops": 0,
            "trolleys_in_to_mcc": 2,
            "hard_toys": 4,
            "soft_toys": 5,
            "black_bags": 6,
            "shoe_bags": 7,
        }
        second_payload = {
            "row_id": rows[1]["row_id"],
            "clothing_kg": None,
            "shoes_kg": 0,
            "time_in": "",
            "time_out": None,
            "trolleys_out_to_opshops": 1,
            "trolleys_in_to_mcc": 0,
            "hard_toys": 0,
            "soft_toys": 0,
            "black_bags": 0,
            "shoe_bags": 0,
        }
        updated = self.client.patch(
            f"/api/manual-dispatch/opshop/pickup-collections/{collection_id}/rows",
            json={"rows": [first_payload, second_payload]},
        )
        self.assertEqual(200, updated.status_code)
        by_id = {
            row["row_id"]: row
            for row in updated.json()["pickups"]
        }
        first = by_id[first_payload["row_id"]]
        self.assertEqual(12.25, first["clothing_kg_snapshot"])
        self.assertEqual(3.0, first["shoes_kg_snapshot"])
        self.assertEqual("09:05", first["time_in_snapshot"])
        self.assertEqual("09:45", first["time_out_snapshot"])
        self.assertEqual(0, first["trolleys_out_to_opshops_snapshot"])
        self.assertEqual(2, first["trolleys_in_to_mcc_snapshot"])
        self.assertEqual(4, first["hard_toys_snapshot"])
        self.assertEqual(5, first["soft_toys_snapshot"])
        self.assertEqual(6, first["black_bags_snapshot"])
        self.assertEqual(7, first["shoe_bags_snapshot"])
        second = by_id[second_payload["row_id"]]
        self.assertIsNone(second["clothing_kg_snapshot"])
        self.assertIsNone(second["time_in_snapshot"])
        self.assertEqual(0, second["shoes_kg_snapshot"])

        invalid_atomic = self.client.patch(
            f"/api/manual-dispatch/opshop/pickup-collections/{collection_id}/rows",
            json={
                "rows": [
                    {**first_payload, "clothing_kg": 99},
                    {**second_payload, "hard_toys": 1.5},
                ]
            },
        )
        self.assertEqual(400, invalid_atomic.status_code)
        unchanged = self.client.get(
            f"/api/manual-dispatch/opshop/pickup-collections/{collection_id}"
        ).json()
        unchanged_by_id = {
            row["row_id"]: row
            for row in unchanged["pickups"]
        }
        self.assertEqual(
            12.25,
            unchanged_by_id[first_payload["row_id"]]["clothing_kg_snapshot"],
        )

        invalid_rows = (
            [{**first_payload, "clothing_kg": -1}],
            [{**first_payload, "clothing_kg": "1.5"}],
            [{**first_payload, "time_in": "24:00"}],
            [first_payload, first_payload],
            [{**first_payload, "row_id": "UNKNOWN-ROW"}],
        )
        for invalid in invalid_rows:
            with self.subTest(invalid=invalid):
                response = self.client.patch(
                    f"/api/manual-dispatch/opshop/pickup-collections/{collection_id}/rows",
                    json={"rows": invalid},
                )
                self.assertEqual(400, response.status_code)

        self._seed_assigned_pickup(
            task_id="PICKUP-OTHER-COLLECTION",
            driver_id="D002",
            pickup_date=self.dispatch_date,
            opshop_id="OPSHOP-OTHER-COLLECTION",
            schedule_id="SCHEDULE-OTHER-COLLECTION",
            opshop_name="Other Driver Op Shop",
            suburb="Preston",
        )
        other = self.client.post(
            "/api/manual-dispatch/opshop/pickup-collections/generated",
            json=self._collection_generate_payload("D002"),
        )
        self.assertEqual(200, other.status_code)
        other_row_id = other.json()["pickups"][0]["row_id"]
        wrong_collection = self.client.patch(
            f"/api/manual-dispatch/opshop/pickup-collections/{collection_id}/rows",
            json={"rows": [{**first_payload, "row_id": other_row_id}]},
        )
        self.assertEqual(400, wrong_collection.status_code)

        cleared = self.client.patch(
            f"/api/manual-dispatch/opshop/pickup-collections/{collection_id}/rows",
            json={
                "rows": [
                    {
                        "row_id": first_payload["row_id"],
                        "clothing_kg": "",
                        "shoes_kg": "",
                        "time_in": " ",
                        "time_out": "",
                        "trolleys_out_to_opshops": "",
                        "trolleys_in_to_mcc": "",
                        "hard_toys": "",
                        "soft_toys": "",
                        "black_bags": "",
                        "shoe_bags": "",
                    }
                ]
            },
        )
        self.assertEqual(200, cleared.status_code)
        cleared_row = next(
            row
            for row in cleared.json()["pickups"]
            if row["row_id"] == first_payload["row_id"]
        )
        for field_name in (
            "clothing_kg_snapshot",
            "shoes_kg_snapshot",
            "time_in_snapshot",
            "time_out_snapshot",
            "trolleys_out_to_opshops_snapshot",
            "trolleys_in_to_mcc_snapshot",
            "hard_toys_snapshot",
            "soft_toys_snapshot",
            "black_bags_snapshot",
            "shoe_bags_snapshot",
        ):
            self.assertIsNone(cleared_row[field_name])

        saved = self.client.post(
            f"/api/manual-dispatch/opshop/pickup-collections/{collection_id}/save",
            json=self._save_payload(),
        )
        self.assertEqual(200, saved.status_code)
        immutable = self.client.patch(
            f"/api/manual-dispatch/opshop/pickup-collections/{collection_id}/rows",
            json={"rows": [first_payload]},
        )
        self.assertEqual(400, immutable.status_code)

    def test_opshop_collection_partial_patch_preserves_omitted_fields_and_supports_explicit_clears(self):
        generated = self.client.post(
            "/api/manual-dispatch/opshop/pickup-collections/generated",
            json=self._collection_generate_payload(),
        )
        self.assertEqual(200, generated.status_code)
        collection_id = generated.json()["collection_id"]
        row_id = generated.json()["pickups"][0]["row_id"]
        endpoint = (
            f"/api/manual-dispatch/opshop/pickup-collections/{collection_id}/rows"
        )

        seeded = self.client.patch(
            endpoint,
            json={
                "rows": [{
                    "row_id": row_id,
                    "clothing_kg": 1,
                    "shoes_kg": 2,
                    "time_in": "09:00",
                    "black_bags": 4,
                }],
            },
        )
        self.assertEqual(200, seeded.status_code)

        partial = self.client.patch(
            endpoint,
            json={"rows": [{"row_id": row_id, "clothing_kg": 9.5}]},
        )
        self.assertEqual(200, partial.status_code)
        row = partial.json()["pickups"][0]
        self.assertEqual(9.5, row["clothing_kg_snapshot"])
        self.assertEqual(2, row["shoes_kg_snapshot"])
        self.assertEqual("09:00", row["time_in_snapshot"])
        self.assertEqual(4, row["black_bags_snapshot"])

        cleared_null = self.client.patch(
            endpoint,
            json={"rows": [{"row_id": row_id, "shoes_kg": None}]},
        )
        self.assertIsNone(cleared_null.json()["pickups"][0]["shoes_kg_snapshot"])
        self.assertEqual(
            "09:00",
            cleared_null.json()["pickups"][0]["time_in_snapshot"],
        )

        cleared_blank = self.client.patch(
            endpoint,
            json={"rows": [{"row_id": row_id, "time_in": ""}]},
        )
        self.assertIsNone(cleared_blank.json()["pickups"][0]["time_in_snapshot"])

        explicit_zero = self.client.patch(
            endpoint,
            json={"rows": [{"row_id": row_id, "black_bags": 0}]},
        )
        self.assertEqual(0, explicit_zero.json()["pickups"][0]["black_bags_snapshot"])

        multiple = self.client.patch(
            endpoint,
            json={
                "rows": [{
                    "row_id": row_id,
                    "time_out": "10:15",
                    "soft_toys": 3,
                }],
            },
        )
        self.assertEqual("10:15", multiple.json()["pickups"][0]["time_out_snapshot"])
        self.assertEqual(3, multiple.json()["pickups"][0]["soft_toys_snapshot"])
        reloaded = self.client.get(
            f"/api/manual-dispatch/opshop/pickup-collections/{collection_id}"
        ).json()["pickups"][0]
        self.assertEqual(9.5, reloaded["clothing_kg_snapshot"])
        self.assertEqual(0, reloaded["black_bags_snapshot"])
        self.assertEqual("10:15", reloaded["time_out_snapshot"])

    def test_opshop_collection_count_bound_is_validated_before_sqlite_binding(self):
        self._seed_assigned_pickup(
            task_id="PICKUP-OVERFLOW-2",
            driver_id="D001",
            pickup_date=self.dispatch_date,
            opshop_id="OPSHOP-OVERFLOW-2",
            schedule_id="SCHEDULE-OVERFLOW-2",
            opshop_name="Overflow Second Shop",
            suburb="Richmond",
        )
        generated = self.client.post(
            "/api/manual-dispatch/opshop/pickup-collections/generated",
            json=self._collection_generate_payload(),
        )
        collection_id = generated.json()["collection_id"]
        row_ids = [row["row_id"] for row in generated.json()["pickups"]]
        endpoint = (
            f"/api/manual-dispatch/opshop/pickup-collections/{collection_id}/rows"
        )

        maximum = self.client.patch(
            endpoint,
            json={"rows": [{"row_id": row_ids[0], "hard_toys": 2**63 - 1}]},
        )
        self.assertEqual(200, maximum.status_code)
        self.assertEqual(
            2**63 - 1,
            maximum.json()["pickups"][0]["hard_toys_snapshot"],
        )

        for overflow in (2**63, 10**100):
            with self.subTest(overflow=overflow):
                rejected = self.client.patch(
                    endpoint,
                    json={"rows": [{"row_id": row_ids[0], "hard_toys": overflow}]},
                )
                self.assertEqual(400, rejected.status_code)

        atomic = self.client.patch(
            endpoint,
            json={
                "rows": [
                    {"row_id": row_ids[0], "black_bags": 7},
                    {"row_id": row_ids[1], "shoe_bags": 2**63},
                ],
            },
        )
        self.assertEqual(400, atomic.status_code)
        reloaded = self.client.get(
            f"/api/manual-dispatch/opshop/pickup-collections/{collection_id}"
        ).json()
        by_id = {row["row_id"]: row for row in reloaded["pickups"]}
        self.assertIsNone(by_id[row_ids[0]]["black_bags_snapshot"])
        self.assertIsNone(by_id[row_ids[1]]["shoe_bags_snapshot"])

    def test_opshop_collection_count_rejects_booleans_without_persisting(self):
        generated = self.client.post(
            "/api/manual-dispatch/opshop/pickup-collections/generated",
            json=self._collection_generate_payload(),
        )
        self.assertEqual(200, generated.status_code)
        collection_id = generated.json()["collection_id"]
        row_id = generated.json()["pickups"][0]["row_id"]
        endpoint = (
            f"/api/manual-dispatch/opshop/pickup-collections/{collection_id}/rows"
        )

        for value in (True, False):
            with self.subTest(value=value):
                rejected = self.client.patch(
                    endpoint,
                    json={"rows": [{"row_id": row_id, "black_bags": value}]},
                )
                self.assertGreaterEqual(rejected.status_code, 400)
                self.assertLess(rejected.status_code, 500)
                reloaded = self.client.get(
                    f"/api/manual-dispatch/opshop/pickup-collections/{collection_id}"
                )
                self.assertEqual(200, reloaded.status_code)
                self.assertIsNone(
                    reloaded.json()["pickups"][0]["black_bags_snapshot"]
                )

    def test_api_rejects_saved_snapshot_overwrite(self):
        run_sheet_id = self._generate_and_save_delivery()
        collection_id = self._generate_and_save_collection()

        duplicate_delivery_save = self.client.post(
            f"/api/manual-dispatch/delivery/run-sheets/{run_sheet_id}/save",
            json=self._save_payload(),
        )
        duplicate_collection_save = self.client.post(
            f"/api/manual-dispatch/opshop/pickup-collections/{collection_id}/save",
            json=self._save_payload(),
        )
        regenerate_delivery = self.client.post(
            "/api/manual-dispatch/delivery/run-sheets/generated",
            json=self._delivery_generate_payload(),
        )
        regenerate_collection = self.client.post(
            "/api/manual-dispatch/opshop/pickup-collections/generated",
            json=self._collection_generate_payload(),
        )

        self.assertEqual(400, duplicate_delivery_save.status_code)
        self.assertEqual(400, duplicate_collection_save.status_code)
        self.assertEqual(409, regenerate_delivery.status_code)
        self.assertEqual(409, regenerate_collection.status_code)

    def test_exports_use_saved_snapshots_and_keep_domains_separate(self):
        run_sheet_id = self._generate_and_save_delivery()
        collection_id = self._generate_and_save_collection()

        order = self.repository.get_order("ORD-001")
        order.company_name = "Edited Live Customer"
        self.repository.update_order(order)
        location = self.repository.get_opshop_location("OPSHOP-WORKSPACE")
        location.name = "Edited Live OP SHOP"
        self.repository.upsert_opshop_location(location)

        delivery_values = self._workbook_values(
            self.client.get(
                f"/api/manual-dispatch/delivery/run-sheets/{run_sheet_id}/export-excel"
            ).content
        )
        collection_values = self._workbook_values(
            self.client.get(
                f"/api/manual-dispatch/opshop/pickup-collections/{collection_id}/export-excel"
            ).content
        )

        self.assertIn("Demo Customer A", delivery_values)
        self.assertNotIn("Edited Live Customer", delivery_values)
        self.assertNotIn("Northside Op Shop", delivery_values)
        self.assertNotIn("OP SHOP Name", delivery_values)
        self.assertIn("Northside Op Shop", collection_values)
        self.assertNotIn("Edited Live OP SHOP", collection_values)
        self.assertIn("DAILY OP SHOP COLLECTIONS - WEIGHT SHEET", collection_values)
        self.assertIn("PLEASE RECORD WEIGHT OF BAGS FOR EACH OP SHOP ", collection_values)
        self.assertIn("CLOTHING KG", collection_values)
        self.assertIn("SHOES KG", collection_values)
        self.assertNotIn("KG", collection_values)
        self.assertNotIn("Call Before Arrival", collection_values)
        self.assertNotIn("30 minutes", collection_values)
        self.assertNotIn("Trip 1", collection_values)
        self.assertNotIn("Total Pallets", collection_values)
        self.assertNotIn("Demo Customer A", collection_values)

    def test_opshop_collection_export_uses_daily_weight_sheet_layout(self):
        collection_id = self._generate_and_save_collection(
            self._weight_sheet_entry_payload()
        )
        response = self.client.get(
            f"/api/manual-dispatch/opshop/pickup-collections/{collection_id}/export-excel"
        )
        self.assertEqual(200, response.status_code)
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        self.assertEqual("WEIGHT SHEET", worksheet.title)
        self.assertEqual("landscape", worksheet.page_setup.orientation)
        self.assertEqual(str(worksheet.PAPERSIZE_A4), str(worksheet.page_setup.paperSize))
        self.assertEqual(1, worksheet.page_setup.fitToWidth)
        self.assertEqual(0, worksheet.page_setup.fitToHeight)
        self.assertEqual("DAILY OP SHOP COLLECTIONS - WEIGHT SHEET", worksheet["A1"].value)
        self.assertIn("NO BOARD GAMES/ PUZZLES", worksheet["A3"].value)
        self.assertIn("HARD & SOFT TOYS", worksheet["A4"].value)
        self.assertIn("DRIVER NAME: John", worksheet["A5"].value)
        self.assertIn("PICK UP DATE: 05/05/2026", worksheet["A5"].value)
        self.assertIn("DAY: TUESDAY", worksheet["A5"].value)
        self.assertEqual("REGO # ________________________", worksheet["A6"].value)
        self.assertEqual("PLEASE RECORD WEIGHT OF BAGS FOR EACH OP SHOP ", worksheet["A8"].value)
        self.assertEqual(
            [
                "OPSHOP NAME",
                "Suburb",
                "CLOTHING KG",
                "SHOES KG",
                "TIME IN",
                "TIME OUT",
                "TROLLEYS OUT TO OPSHOPS",
                "TROLLEYS IN TO MCC ",
                "HARD TOYS",
                "SOFT TOYS",
                "BLACK BAGS",
                "SHOE BAGS",
            ],
            [cell.value for cell in worksheet[10]],
        )
        self.assertEqual("Northside Op Shop", worksheet["A12"].value)
        self.assertEqual("Coburg", worksheet["B12"].value)
        self.assertEqual(12.5, worksheet["C12"].value)
        self.assertEqual(3.25, worksheet["D12"].value)
        self.assertEqual('0.## "KG"', worksheet["C12"].number_format)
        self.assertEqual('0.## "KG"', worksheet["D12"].number_format)
        self.assertEqual("09:15", worksheet["E12"].value)
        self.assertEqual("09:45", worksheet["F12"].value)
        self.assertEqual([2, 1, 4, 5, 6, 0], [
            worksheet.cell(row=12, column=column).value
            for column in range(7, 13)
        ])
        self.assertIsNone(worksheet["C13"].value)
        self.assertEqual('0.## "KG"', worksheet["C13"].number_format)
        self.assertGreater(worksheet.column_dimensions["A"].width, 28)
        self.assertIn("$A$1:$L$22", worksheet.print_area)

    def test_opshop_daily_collection_export_uses_one_sheet_per_driver(self):
        saved_collection_id = self._generate_and_save_collection(
            self._weight_sheet_entry_payload()
        )
        self._seed_assigned_pickup(
            task_id="PICKUP-TONY",
            driver_id="D002",
            pickup_date=self.dispatch_date,
            opshop_id="OPSHOP-TONY",
            schedule_id="SCHEDULE-TONY",
            opshop_name="Tony North Op Shop",
            suburb="Preston",
        )
        generated_tony = self.client.post(
            "/api/manual-dispatch/opshop/pickup-collections/generated",
            json=self._collection_generate_payload("D002", self.dispatch_date),
        )
        self.assertEqual(200, generated_tony.status_code)
        self._seed_assigned_pickup(
            task_id="PICKUP-OTHER-DATE",
            driver_id="D001",
            pickup_date="2026-05-06",
            opshop_id="OPSHOP-OTHER-DATE",
            schedule_id="SCHEDULE-OTHER-DATE",
            opshop_name="Other Date Op Shop",
            suburb="Brunswick",
        )
        other_date = self.client.post(
            "/api/manual-dispatch/opshop/pickup-collections/generated",
            json=self._collection_generate_payload("D001", "2026-05-06"),
        )
        self.assertEqual(200, other_date.status_code)

        location = self.repository.get_opshop_location("OPSHOP-WORKSPACE")
        location.name = "Edited Live OP SHOP"
        self.repository.upsert_opshop_location(location)

        response = self.client.get(
            "/api/manual-dispatch/opshop/pickup-collections/export-excel",
            params={
                "dispatch_date": self.dispatch_date,
                "pickup_date": self.dispatch_date,
            },
        )
        self.assertEqual(200, response.status_code)
        self.assertEqual(
            'attachment; filename="Daily_OPSHOP_Collections_2026-05-05.xlsx"',
            response.headers["content-disposition"],
        )
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual({"John", "Tony"}, set(workbook.sheetnames))
        self.assertNotIn("WEIGHT SHEET", workbook.sheetnames)
        self.assertNotIn("Sheet", workbook.sheetnames)

        john_values = [
            cell.value
            for row in workbook["John"].iter_rows()
            for cell in row
            if cell.value is not None
        ]
        tony_values = [
            cell.value
            for row in workbook["Tony"].iter_rows()
            for cell in row
            if cell.value is not None
        ]
        self.assertIn("Northside Op Shop", john_values)
        self.assertIn("Tony North Op Shop", tony_values)
        self.assertNotIn("Tony North Op Shop", john_values)
        self.assertNotIn("Northside Op Shop", tony_values)
        self.assertNotIn("Edited Live OP SHOP", john_values + tony_values)
        self.assertNotIn("Other Date Op Shop", john_values + tony_values)
        for worksheet in workbook.worksheets:
            self.assertEqual("landscape", worksheet.page_setup.orientation)
            self.assertEqual(1, worksheet.page_setup.fitToWidth)
            self.assertEqual(0, worksheet.page_setup.fitToHeight)
            self.assertEqual("DAILY OP SHOP COLLECTIONS - WEIGHT SHEET", worksheet["A1"].value)
            self.assertIn("PICK UP DATE: 05/05/2026", worksheet["A5"].value)
            self.assertEqual("PLEASE RECORD WEIGHT OF BAGS FOR EACH OP SHOP ", worksheet["A8"].value)
        self.assertEqual(12.5, workbook["John"]["C12"].value)
        self.assertEqual(3.25, workbook["John"]["D12"].value)
        self.assertEqual("09:15", workbook["John"]["E12"].value)
        self.assertEqual(0, workbook["John"]["L12"].value)
        self.assertEqual('0.## "KG"', workbook["John"]["C12"].number_format)
        self.assertIsNone(workbook["Tony"]["C12"].value)
        self.assertEqual('0.## "KG"', workbook["Tony"]["C12"].number_format)

        statuses = {
            collection.collection_id: collection.status
            for collection in self.service.list_opshop_pickup_collections(
                pickup_date=self.dispatch_date
            )
        }
        self.assertEqual("SAVED", statuses[saved_collection_id])
        self.assertEqual("GENERATED", statuses[generated_tony.json()["collection_id"]])

    def test_opshop_daily_collection_export_rejects_empty_pickup_date(self):
        response = self.client.get(
            "/api/manual-dispatch/opshop/pickup-collections/export-excel",
            params={"pickup_date": "2026-05-07"},
        )
        self.assertEqual(400, response.status_code)
        self.assertEqual(
            "No OP SHOP Pickup Collections available for this pickup date.",
            response.json()["detail"],
        )

    def test_delivery_export_uses_daily_run_sheet_form_layout(self):
        self._assign_delivery_vehicle("D001", "V001")
        self._set_product_lines(
            "ORD-001",
            [
                ProductDetailLine(
                    "Snapshot Alpha Product",
                    450,
                    "KG",
                    product_code="RWIND",
                    package_quantity=45,
                    package_unit="BAG10",
                ),
                ProductDetailLine("Snapshot Beta Product", 3, "BAGS"),
                ProductDetailLine("Snapshot Alpha Product", 1, "PALLETS"),
            ],
        )
        order = self.repository.get_order("ORD-001")
        order.loose_bags_quantity = 5
        order.carton_quantity = 4
        order.note = "Use rear delivery gate"
        self.repository.update_order(order)
        self.service.assign_task(
            AssignTaskRequest(
                dispatch_date=self.dispatch_date,
                task_type="ORDER",
                task_id="ORD-002",
                driver_id="D001",
                trip_no="trip2",
            )
        )
        run_sheet_id = self._generate_and_save_delivery()
        with sqlite3.connect(self.db_path) as connection:
            connection.execute(
                "UPDATE manual_vehicles SET rego = ? WHERE vehicle_id = ?",
                ("EDITED-LIVE-REGO", "V001"),
            )
        workbook = load_workbook(
            BytesIO(
                self.client.get(
                    f"/api/manual-dispatch/delivery/run-sheets/{run_sheet_id}/export-excel"
                ).content
            )
        )
        worksheet = workbook.active
        self.assertEqual("Daily Run Sheet", worksheet.title)
        self.assertEqual("landscape", worksheet.page_setup.orientation)
        self.assertEqual(str(worksheet.PAPERSIZE_A4), str(worksheet.page_setup.paperSize))

        values = [
            cell.value
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        ]
        self.assertIn("DAILY RUN SHEET", values)
        self.assertIn("DATE  05/05/2026", values)
        self.assertIn("DRIVER: John", values)
        self.assertIn("REGO #", values)
        self.assertIn("ABC123", values)
        self.assertNotIn("EDITED-LIVE-REGO", values)
        self.assertIn("START TIME: ____________________________________", values)
        self.assertIn(
            "TIME LOADING STARTED(TO BE FILLED IN BY STOREMAN)___________",
            values,
        )
        self.assertIn(
            "TIME LOADING COMPLETED(TO BE FILLED IN BY STOREMAN)_____________",
            values,
        )
        self.assertIn("FINISH TIME:____________________________________", values)
        self.assertNotIn("Final Trip Summary", values)
        self.assertNotIn("Address", values)
        self.assertNotIn("Product Details", values)
        self.assertNotIn("Order #", values)
        self.assertEqual("DRIVER: John", worksheet["F1"].value)
        self.assertEqual("REGO #", worksheet["L1"].value)
        self.assertEqual("ABC123", worksheet["M1"].value)

        header_row = [cell.value for cell in worksheet[8]]
        self.assertEqual(
            [
                None,
                "Customer Name",
                "Suburb",
                "Invoice #",
                "PRODUCT",
                "KG'S",
                "Pallets",
                "Loose Bags",
                "Cartons",
                "Notes",
                "COD",
                "CQ",
                "Time In",
                "Time Out",
                "PRINT NAME",
                "SIGNATURE",
                "NO. # PALLETS RETND",
            ],
            header_row,
        )

        rows = list(worksheet.iter_rows(values_only=True))
        customer_a_index = next(
            index for index, row in enumerate(rows) if row[1] == "Demo Customer A"
        )
        customer_b_index = next(
            index for index, row in enumerate(rows) if row[1] == "Demo Customer B"
        )
        self.assertLess(customer_a_index, customer_b_index)
        customer_a_row = rows[customer_a_index]
        self.assertEqual(1, customer_a_row[0])
        self.assertEqual("INV-1001", customer_a_row[3])
        self.assertEqual(
            "1. [RWIND] Snapshot Alpha Product - 450 KG | Packaging: 45 BAG10\n"
            "2. Snapshot Beta Product - 3 BAGS\n"
            "3. Snapshot Alpha Product - 1 PALLETS",
            customer_a_row[4],
        )
        self.assertEqual(450, customer_a_row[5])
        self.assertEqual(2, customer_a_row[6])
        self.assertEqual(5, customer_a_row[7])
        self.assertEqual(4, customer_a_row[8])
        self.assertEqual("Use rear delivery gate", customer_a_row[9])
        self.assertTrue(worksheet.cell(row=customer_a_index + 1, column=5).alignment.wrap_text)
        self.assertGreaterEqual(worksheet.row_dimensions[customer_a_index + 1].height, 30)
        for manual_column in range(10, 17):
            self.assertIsNone(customer_a_row[manual_column])

    def test_delivery_date_export_uses_snapshot_rows_and_one_sheet_per_driver(self):
        self._assign_delivery_vehicle("D001", "V001")
        self._assign_delivery_vehicle("D002", "V002")
        self._set_product_lines(
            "ORD-001",
            [
                ProductDetailLine(
                    "Snapshot Alpha Product",
                    450,
                    "KG",
                    product_code="RWIND",
                    package_quantity=45,
                    package_unit="BAG10",
                ),
                ProductDetailLine("Snapshot Beta Product", 3, "BAGS"),
                ProductDetailLine("Snapshot Alpha Product", 9, "PALLETS"),
            ],
        )
        snapshot_order = self.repository.get_order("ORD-001")
        snapshot_order.loose_bags_quantity = 5
        snapshot_order.carton_quantity = 4
        snapshot_order.note = "Use rear delivery gate"
        self.repository.update_order(snapshot_order)
        self.service.assign_task(
            AssignTaskRequest(
                dispatch_date=self.dispatch_date,
                task_type="ORDER",
                task_id="ORD-003",
                driver_id="D001",
                trip_no="trip2",
            )
        )
        self.service.assign_task(
            AssignTaskRequest(
                dispatch_date=self.dispatch_date,
                task_type="ORDER",
                task_id="ORD-002",
                driver_id="D002",
                trip_no="trip1",
            )
        )
        john = self.client.post(
            "/api/manual-dispatch/delivery/run-sheets/generated",
            json=self._delivery_generate_payload(),
        ).json()
        self.assertEqual(
            200,
            self.client.post(
                f"/api/manual-dispatch/delivery/run-sheets/{john['run_sheet_id']}/save",
                json=self._save_payload(),
            ).status_code,
        )
        tony = self.client.post(
            "/api/manual-dispatch/delivery/run-sheets/generated",
            json={
                "dispatch_date": self.dispatch_date,
                "delivery_date": self.dispatch_date,
                "driver_id": "D002",
            },
        ).json()

        for order_id in ("ORD-001", "ORD-002", "ORD-003"):
            order = self.repository.get_order(order_id)
            order.company_name = f"Edited Live {order_id}"
            order.suburb = "Edited Live Suburb"
            order.product_lines = [
                ProductDetailLine(f"Edited Live Product {order_id}", 1, "PALLETS")
            ]
            order.loose_bags_quantity = 99
            order.carton_quantity = 99
            order.note = "Edited live note"
            self.repository.update_order(order)
        with sqlite3.connect(self.db_path) as connection:
            connection.execute(
                "UPDATE manual_vehicles SET rego = ? WHERE vehicle_id IN (?, ?)",
                ("EDITED-LIVE-REGO", "V001", "V002"),
            )

        response = self.client.get(
            "/api/manual-dispatch/delivery/run-sheets/export-excel",
            params={"delivery_date": self.dispatch_date},
        )
        self.assertEqual(200, response.status_code)
        self.assertEqual(
            'attachment; filename="Daily_Run_Sheets_2026-05-05.xlsx"',
            response.headers["content-disposition"],
        )
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(2, len(workbook.worksheets))
        self.assertEqual({"John", "Tony"}, set(workbook.sheetnames))
        self.assertNotIn("Sheet", workbook.sheetnames)

        expected_headers = [
            None,
            "Customer Name",
            "Suburb",
            "Invoice #",
            "PRODUCT",
            "KG'S",
            "Pallets",
            "Loose Bags",
            "Cartons",
            "Notes",
            "COD",
            "CQ",
            "Time In",
            "Time Out",
            "PRINT NAME",
            "SIGNATURE",
            "NO. # PALLETS RETND",
        ]
        for worksheet in workbook.worksheets:
            self.assertEqual("landscape", worksheet.page_setup.orientation)
            self.assertEqual(
                str(worksheet.PAPERSIZE_A4),
                str(worksheet.page_setup.paperSize),
            )
            self.assertEqual(1, worksheet.page_setup.fitToWidth)
            self.assertEqual(0, worksheet.page_setup.fitToHeight)
            self.assertEqual("$8:$8", worksheet.print_title_rows)
            self.assertEqual(expected_headers, [cell.value for cell in worksheet[8]])
            self.assertEqual(17, worksheet.max_column)
            self.assertEqual("DAILY RUN SHEET", worksheet["A1"].value)
            self.assertEqual("DATE  05/05/2026", worksheet["C1"].value)
            self.assertTrue(str(worksheet["F1"].value).startswith("DRIVER: "))
            self.assertEqual("REGO #", worksheet["L1"].value)
            values = [
                cell.value
                for row in worksheet.iter_rows()
                for cell in row
                if cell.value is not None
            ]
            for label in (
                "START TIME: ____________________________________",
                "TIME LOADING STARTED(TO BE FILLED IN BY STOREMAN)___________",
                "TIME LOADING COMPLETED(TO BE FILLED IN BY STOREMAN)_____________",
                "FINISH TIME:____________________________________",
            ):
                self.assertIn(label, values)
            self.assertIn("COD", values)
            self.assertIn("CQ", values)
            self.assertIn("Time In", values)
            self.assertIn("Time Out", values)
            self.assertIn("PRINT NAME", values)
            self.assertIn("SIGNATURE", values)
            self.assertIn("NO. # PALLETS RETND", values)
            self.assertFalse(any(str(value).startswith("Edited Live") for value in values))

        john_rows = list(workbook["John"].iter_rows(values_only=True))
        self.assertEqual(1, john_rows[8][0])
        self.assertEqual("Demo Customer A", john_rows[8][1])
        self.assertEqual("Demo Customer C", john_rows[9][1])
        self.assertEqual("INV-1001", john_rows[8][3])
        self.assertEqual(
            "1. [RWIND] Snapshot Alpha Product - 450 KG | Packaging: 45 BAG10\n"
            "2. Snapshot Beta Product - 3 BAGS\n"
            "3. Snapshot Alpha Product - 9 PALLETS",
            john_rows[8][4],
        )
        self.assertEqual(450, john_rows[8][5])
        self.assertEqual(2, john_rows[8][6])
        self.assertEqual(5, john_rows[8][7])
        self.assertEqual(4, john_rows[8][8])
        self.assertEqual("Use rear delivery gate", john_rows[8][9])
        self.assertEqual(450, workbook["John"]["F9"].value)
        self.assertEqual("General", workbook["John"]["G9"].number_format)
        self.assertTrue(workbook["John"]["E9"].alignment.wrap_text)
        self.assertGreaterEqual(workbook["John"].row_dimensions[9].height, 30)
        self.assertGreater(workbook["John"].column_dimensions["B"].width, 25)
        self.assertNotIn("Edited Live Product ORD-001", self._workbook_values(response.content))
        self.assertEqual("DRIVER: John", workbook["John"]["F1"].value)
        self.assertEqual("ABC123", workbook["John"]["M1"].value)
        self.assertEqual("DRIVER: Tony", workbook["Tony"]["F1"].value)
        self.assertEqual("XYZ888", workbook["Tony"]["M1"].value)
        self.assertNotIn("EDITED-LIVE-REGO", self._workbook_values(response.content))

        statuses = {
            run_sheet.run_sheet_id: run_sheet.status
            for run_sheet in self.service.list_delivery_run_sheets(
                delivery_date=self.dispatch_date
            )
        }
        self.assertEqual("SAVED", statuses[john["run_sheet_id"]])
        self.assertEqual("GENERATED", statuses[tony["run_sheet_id"]])

        snapshots = self.service.list_delivery_run_sheets_for_date_export(
            self.dispatch_date
        )
        duplicate_name_bytes = build_delivery_run_sheets_excel(
            [
                replace(
                    snapshots[0],
                    driver_name_snapshot="Driver / North: Very Long Shared Name 123456789",
                ),
                replace(
                    snapshots[1],
                    driver_name_snapshot="Driver / North: Very Long Shared Name 123456789",
                ),
            ],
            self.dispatch_date,
        )
        duplicate_name_workbook = load_workbook(BytesIO(duplicate_name_bytes))
        self.assertEqual(2, len(set(duplicate_name_workbook.sheetnames)))
        self.assertTrue(all(len(name) <= 31 for name in duplicate_name_workbook.sheetnames))
        self.assertTrue(
            all(not any(character in name for character in "[]:*?/\\")
                for name in duplicate_name_workbook.sheetnames)
        )
        no_rego_bytes = build_delivery_run_sheets_excel(
            [replace(snapshots[0], vehicle_rego_snapshot=None)],
            self.dispatch_date,
        )
        no_rego_workbook = load_workbook(BytesIO(no_rego_bytes))
        self.assertEqual("Not selected", no_rego_workbook.active["M1"].value)

    def test_delivery_run_sheet_product_display_prefers_snapshot_lines(self):
        order = DeliveryRunSheetOrderSnapshot(
            row_id="ROW-1",
            trip_no="trip1",
            row_no=1,
            task_type="ORDER",
            task_id="ORDER-1",
            order_id_snapshot="ORDER-1",
            invoice_number_snapshot="INV-1",
            order_no_snapshot=None,
            company_name_snapshot="Customer",
            suburb_snapshot="Suburb",
            delivery_address_snapshot="Address",
            product_snapshot="Legacy Product",
            pallet_quantity_snapshot=1,
            loose_bags_quantity_snapshot=7,
            note_snapshot=None,
            product_lines_snapshot=[
                ProductDetailLine(
                    "Product A",
                    450,
                    "KG",
                    product_code="RWIND",
                    package_quantity=45,
                    package_unit="BAG10",
                ),
                ProductDetailLine("Product B", 2, "BAGS"),
                ProductDetailLine("Product A", 3, "BAGS"),
            ],
        )
        self.assertEqual(
            "1. [RWIND] Product A - 450 KG | Packaging: 45 BAG10\n"
            "2. Product B - 2 BAGS\n"
            "3. Product A - 3 BAGS",
            delivery_run_sheet_product_display(order),
        )

        order.product_lines_snapshot = []
        self.assertEqual("Legacy Product", delivery_run_sheet_product_display(order))

        order.product_snapshot = "  "
        self.assertEqual("", delivery_run_sheet_product_display(order))

    def test_delivery_date_export_rejects_empty_scope_without_mutation(self):
        response = self.client.get(
            "/api/manual-dispatch/delivery/run-sheets/export-excel",
            params={"delivery_date": "2026-05-06"},
        )
        self.assertEqual(400, response.status_code)
        self.assertEqual(
            "No Generated or Saved Delivery Run Sheets are available for this Delivery Date.",
            response.json()["detail"],
        )
        self.assertEqual(
            [],
            self.service.list_delivery_run_sheets(delivery_date="2026-05-06"),
        )

    def test_legacy_final_summary_schema_and_routes_remain_available(self):
        route_pairs = {
            (method, route.path)
            for route in self.api_module.router.routes
            for method in route.methods
        }
        self.assertIn(("POST", "/api/manual-dispatch/final-summaries"), route_pairs)
        self.assertIn(("GET", "/api/manual-dispatch/final-summaries"), route_pairs)
        self.assertIn(("GET", "/api/manual-dispatch/board"), route_pairs)

        with sqlite3.connect(self.db_path) as connection:
            legacy_tables = {
                row[0]
                for row in connection.execute(
                    "SELECT name FROM sqlite_master WHERE type = 'table'"
                ).fetchall()
            }
        self.assertTrue(
            {
                "final_trip_summaries",
                "final_trip_summary_rows",
                "final_trip_summary_opshop_pickup_rows",
            }.issubset(legacy_tables)
        )

    def _assign_order(self):
        self.service.assign_task(
            AssignTaskRequest(
                dispatch_date=self.dispatch_date,
                task_type="ORDER",
                task_id="ORD-001",
                driver_id="D001",
                trip_no="trip1",
            )
        )

    def _assign_delivery_vehicle(self, driver_id, vehicle_id):
        self.service.assign_delivery_workspace_vehicle(
            DeliveryWorkspaceVehicleAssignmentRequest(
                dispatch_date=self.dispatch_date,
                delivery_date=self.dispatch_date,
                driver_id=driver_id,
                vehicle_id=vehicle_id,
            )
        )

    def _set_product_lines(self, order_id, product_lines):
        order = self.repository.get_order(order_id)
        order.product_lines = product_lines
        self.repository.update_order(order)

    def _seed_and_assign_pickup(self):
        self._seed_assigned_pickup(
            task_id="PICKUP-WORKSPACE",
            driver_id="D001",
            pickup_date=self.dispatch_date,
            opshop_id="OPSHOP-WORKSPACE",
            schedule_id="SCHEDULE-WORKSPACE",
            opshop_name="Northside Op Shop",
            suburb="Coburg",
        )

    def _seed_assigned_pickup(
        self,
        task_id,
        driver_id,
        pickup_date,
        opshop_id,
        schedule_id,
        opshop_name,
        suburb,
    ):
        self.repository.upsert_opshop_location(
            OpShopLocation(
                opshop_id=opshop_id,
                name=opshop_name,
                suburb=suburb,
                street_address="1 Sydney Road",
                area_region="North",
                primary_contact="Mary",
                primary_phone="0400 000 001",
                secondary_contact=None,
                secondary_phone=None,
                access_type="Rear dock",
                key_required=True,
                trailer_restriction="Small truck only",
                status_notes="Ring first",
                is_active=True,
                created_at="2026-05-05T00:00:00+00:00",
                updated_at="2026-05-05T00:00:00+00:00",
            )
        )
        self.repository.upsert_opshop_pickup_schedule(
            OpShopPickupSchedule(
                schedule_id=schedule_id,
                opshop_id=opshop_id,
                run_day="TUESDAY",
                run_type="ON_CALL",
                pickup_frequency="Weekly",
                time_window="09:00-12:00",
                call_before_arrival=True,
                call_timing="30 minutes",
                status="Active",
                active_flag=True,
                fortnight_group=None,
                review_required=False,
                review_reason=None,
                created_at="2026-05-05T00:00:00+00:00",
                updated_at="2026-05-05T00:00:00+00:00",
            )
        )
        self.repository.upsert_opshop_pickup_task(
            OpShopPickupTask(
                pickup_task_id=task_id,
                schedule_id=schedule_id,
                opshop_id=opshop_id,
                pickup_date=pickup_date,
                task_type="OPSHOP_PICKUP",
                generated_from="ON_CALL",
                status="ACTIVE",
                dispatch_date=pickup_date,
                driver_id=None,
                trip_no=None,
                notes="Leave at rear door",
                created_at="2026-05-05T00:00:00+00:00",
                updated_at="2026-05-05T00:00:00+00:00",
            )
        )
        self.service.assign_task(
            AssignTaskRequest(
                dispatch_date=pickup_date,
                task_type="OPSHOP_PICKUP",
                task_id=task_id,
                driver_id=driver_id,
                trip_no="trip1",
            )
        )

    def _generate_and_save_delivery(self):
        generated = self.client.post(
            "/api/manual-dispatch/delivery/run-sheets/generated",
            json=self._delivery_generate_payload(),
        )
        run_sheet_id = generated.json()["run_sheet_id"]
        saved = self.client.post(
            f"/api/manual-dispatch/delivery/run-sheets/{run_sheet_id}/save",
            json=self._save_payload(),
        )
        self.assertEqual(200, saved.status_code)
        return run_sheet_id

    def _generate_and_save_collection(self, row_values=None):
        generated = self.client.post(
            "/api/manual-dispatch/opshop/pickup-collections/generated",
            json=self._collection_generate_payload(),
        )
        self.assertEqual(200, generated.status_code)
        generated_payload = generated.json()
        collection_id = generated_payload["collection_id"]
        if row_values is not None:
            updated = self.client.patch(
                f"/api/manual-dispatch/opshop/pickup-collections/{collection_id}/rows",
                json={
                    "rows": [{
                        "row_id": generated_payload["pickups"][0]["row_id"],
                        **row_values,
                    }],
                },
            )
            self.assertEqual(200, updated.status_code)
        saved = self.client.post(
            f"/api/manual-dispatch/opshop/pickup-collections/{collection_id}/save",
            json=self._save_payload(),
        )
        self.assertEqual(200, saved.status_code)
        return collection_id

    def _delivery_generate_payload(self):
        return {
            "dispatch_date": self.dispatch_date,
            "delivery_date": self.dispatch_date,
            "driver_id": "D001",
        }

    def _collection_generate_payload(self, driver_id="D001", pickup_date=None):
        return {
            "dispatch_date": self.dispatch_date,
            "pickup_date": pickup_date or self.dispatch_date,
            "driver_id": driver_id,
        }

    @staticmethod
    def _weight_sheet_entry_payload():
        return {
            "clothing_kg": 12.5,
            "shoes_kg": 3.25,
            "time_in": "09:15",
            "time_out": "09:45",
            "trolleys_out_to_opshops": 2,
            "trolleys_in_to_mcc": 1,
            "hard_toys": 4,
            "soft_toys": 5,
            "black_bags": 6,
            "shoe_bags": 0,
        }

    def _save_payload(self):
        return {
            "saved_by_account_name": self.account.account_name,
            "saved_by_account_id": self.account.account_id,
        }

    @staticmethod
    def _workbook_values(workbook_bytes):
        workbook = load_workbook(BytesIO(workbook_bytes))
        return [
            cell.value
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        ]


if __name__ == "__main__":
    unittest.main()
