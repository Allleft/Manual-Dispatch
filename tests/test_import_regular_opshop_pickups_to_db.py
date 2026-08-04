import shutil
import sqlite3
import unittest
import uuid
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from backend.repositories.sqlite_manual_dispatch_repository import SQLiteManualDispatchRepository
from backend.schemas import (
    CreateDriverRequest,
    CreateOpShopTemplateRequest,
    OpShopLocation,
    OpShopPickupSchedule,
    UpdateDriverRequest,
)
from backend.services.manual_dispatch_service import ManualDispatchService
from tools.import_regular_opshop_pickups_to_db import (
    REQUIRED_COLUMNS,
    deactivate_regular_schedules,
    import_regular_opshop_pickups_to_db,
)


class ImportRegularOpShopPickupsToDbTest(unittest.TestCase):
    def setUp(self):
        self.temp_dir = Path.cwd() / "tmp" / f"regular-opshop-import-test-{uuid.uuid4().hex}"
        self.temp_dir.mkdir(parents=True)
        self.workbook_path = self.temp_dir / "Opshop reuglar pickup.xlsx"
        self.db_path = self.temp_dir / "manual_dispatch.sqlite3"
        with patch.dict("os.environ", {"MANUAL_DISPATCH_SEED_DEMO_DATA": "true"}):
            self.repository = SQLiteManualDispatchRepository(self.db_path)
        self.service = ManualDispatchService(self.repository)
        self.service.update_driver("D001", UpdateDriverRequest(name="John Georgiadis"))
        self.service.update_driver("D002", UpdateDriverRequest(name="Gavin Fynn"))
        self.service.update_driver("D003", UpdateDriverRequest(name="Epaminondas Tsatsoulis"))
        self.lee = self.service.create_driver(CreateDriverRequest(name="Guanlin Li"))

    def tearDown(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_import_reads_weekday_sheets_and_assigned_to_aliases(self):
        self._save_workbook(
            {
                "MON": [self._row(Op_Shop_Name="Monday Shop", Assigned_to="John G")],
                "TUE": [self._row(Op_Shop_Name="Tuesday Shop", Assigned_to="Gavin")],
                "WED": [self._row(Op_Shop_Name="Wednesday Shop", Assigned_to="Nonda")],
                "THU": [self._row(Op_Shop_Name="Thursday Shop", Assigned_to="LEE")],
                "FRI": [self._row(Op_Shop_Name="Friday Shop", Assigned_to="Mystery")],
            }
        )

        summary = import_regular_opshop_pickups_to_db(self.workbook_path, self.db_path)
        schedules = self.repository.list_opshop_pickup_schedules()
        by_day = {schedule.run_day: schedule for schedule in schedules}

        self.assertEqual(5, summary.rows_read)
        self.assertEqual(5, summary.rows_imported)
        self.assertEqual(5, summary.schedules_inserted)
        self.assertEqual(1, summary.unresolved_assigned_to["Mystery"])
        self.assertEqual("REGULAR", by_day["MONDAY"].run_type)
        self.assertEqual("D001", by_day["MONDAY"].default_driver_id)
        self.assertEqual("John G", by_day["MONDAY"].default_driver_alias)
        self.assertEqual("John Georgiadis", by_day["MONDAY"].default_driver_name_snapshot)
        self.assertEqual("D002", by_day["TUESDAY"].default_driver_id)
        self.assertEqual("D003", by_day["WEDNESDAY"].default_driver_id)
        self.assertEqual(self.lee.driver_id, by_day["THURSDAY"].default_driver_id)
        self.assertIsNone(by_day["FRIDAY"].default_driver_id)
        self.assertEqual("Mystery", by_day["FRIDAY"].default_driver_name_snapshot)
        self.assertEqual([], self.repository.list_opshop_pickup_tasks())

    def test_blank_active_flag_with_active_status_is_imported_and_inactive_rows_skip(self):
        self._save_workbook(
            {
                "MON": [
                    self._row(Op_Shop_Name="Blank Active Flag", Active_Flag=""),
                    self._row(Op_Shop_Name="On Hold", Status="On_Hold", Active_Flag="1"),
                    self._row(Op_Shop_Name="Inactive", Status="Active", Active_Flag="0"),
                ]
            }
        )

        summary = import_regular_opshop_pickups_to_db(self.workbook_path, self.db_path)
        locations = self.repository.list_opshop_locations()
        schedules = self.repository.list_opshop_pickup_schedules()

        self.assertEqual(3, summary.rows_read)
        self.assertEqual(1, summary.rows_imported)
        self.assertEqual(2, summary.rows_skipped_inactive)
        self.assertEqual(["Blank Active Flag"], [location.name for location in locations])
        self.assertEqual("REGULAR", schedules[0].run_type)
        self.assertEqual("MONDAY", schedules[0].run_day)

    def test_same_location_multiple_weekday_rows_create_one_location_and_multiple_schedules(self):
        self._save_workbook(
            {
                "MON": [self._row(Op_Shop_Name="Shared Shop", Assigned_to="John G")],
                "WED": [self._row(Op_Shop_Name="Shared Shop", Assigned_to="John G")],
            }
        )

        summary = import_regular_opshop_pickups_to_db(self.workbook_path, self.db_path)

        self.assertEqual(1, summary.locations_inserted)
        self.assertEqual(2, summary.schedules_inserted)
        self.assertEqual(1, len(self.repository.list_opshop_locations()))
        self.assertEqual(2, len(self.repository.list_opshop_pickup_schedules()))

    def test_multi_weekly_frequency_text_is_weekly_like_and_uses_sheet_weekday_only(self):
        for frequency in ["2x Weekly", "2 x Weekly", "Twice weekly", "two times weekly"]:
            with self.subTest(frequency=frequency):
                shutil.rmtree(self.temp_dir, ignore_errors=True)
                self.setUp()
                self._save_workbook(
                    {
                        "THU": [
                            self._row(
                                Op_Shop_Name=f"Thursday {frequency}",
                                Pickup_Frequency=frequency,
                            )
                        ],
                    }
                )

                summary = import_regular_opshop_pickups_to_db(self.workbook_path, self.db_path)
                schedules = self.repository.list_opshop_pickup_schedules()

                self.assertEqual(1, summary.rows_imported)
                self.assertEqual(1, summary.schedules_inserted)
                self.assertEqual(1, len(schedules))
                self.assertEqual("REGULAR", schedules[0].run_type)
                self.assertEqual("THURSDAY", schedules[0].run_day)
                self.assertEqual(frequency, schedules[0].pickup_frequency)

    def test_multi_weekly_frequency_variants_update_same_regular_schedule_key(self):
        self._save_workbook(
            {
                "THU": [
                    self._row(
                        Op_Shop_Name="Northside Op Shop",
                        Pickup_Frequency="2x Weekly",
                        Time_Window="9-12",
                    )
                ],
            }
        )
        first_summary = import_regular_opshop_pickups_to_db(self.workbook_path, self.db_path)
        original_schedule = self.repository.list_opshop_pickup_schedules()[0]
        self.service.opshop_pickup_service.ensure_regular_opshop_pickup_tasks_for_week("2026-05-18")
        original_task = self.repository.list_opshop_pickup_tasks()[0]
        original_assignment = self.repository.find_assignment_for_task(
            "OPSHOP_PICKUP",
            original_task.pickup_task_id,
        )

        self._save_workbook(
            {
                "THU": [
                    self._row(
                        Op_Shop_Name="Northside Op Shop",
                        Pickup_Frequency="Twice weekly",
                        Time_Window="10-1",
                    )
                ],
            }
        )
        second_summary = import_regular_opshop_pickups_to_db(self.workbook_path, self.db_path)
        updated_schedule = self.repository.list_opshop_pickup_schedules()[0]
        retained_task = self.repository.get_opshop_pickup_task(
            original_task.pickup_task_id
        )
        retained_assignment = self.repository.find_assignment_for_task(
            "OPSHOP_PICKUP",
            original_task.pickup_task_id,
        )

        self.assertEqual(1, first_summary.schedules_inserted)
        self.assertEqual(1, second_summary.schedules_updated)
        self.assertEqual(0, second_summary.schedules_inserted)
        self.assertEqual(original_schedule.schedule_id, updated_schedule.schedule_id)
        self.assertEqual("Twice weekly", updated_schedule.pickup_frequency)
        self.assertEqual("10-1", updated_schedule.time_window)
        self.assertEqual(original_schedule.schedule_id, retained_task.schedule_id)
        self.assertEqual(original_assignment.assignment_id, retained_assignment.assignment_id)
        self.assertEqual("D001", retained_assignment.driver_id)
        self.assertEqual(1, len(self.repository.list_opshop_pickup_schedules()))

    def test_real_workbook_address_variants_share_physical_opshop_and_slots(self):
        self._save_workbook(
            {
                "MON": [
                    self._row(
                        Op_Shop_Name="OUR VILLAGE NETWORK (St Kilda mums)",
                        Suburb="CLAYTON",
                        Street_Address="14 Winterton Road",
                        Pickup_Frequency="2x Weekly",
                    )
                ],
                "WED": [
                    self._row(
                        Op_Shop_Name="OUR VILLAGE NETWORK (St Kilda mums)",
                        Suburb="CLAYTON",
                        Street_Address="14 WINTERTON RD",
                        Pickup_Frequency="2x Weekly",
                    )
                ],
            }
        )

        summary = import_regular_opshop_pickups_to_db(
            self.workbook_path,
            self.db_path,
        )
        schedules = sorted(
            self.repository.list_opshop_pickup_schedules(),
            key=lambda schedule: schedule.run_day,
        )

        self.assertEqual(1, summary.locations_inserted)
        self.assertEqual(2, summary.schedules_inserted)
        self.assertEqual(1, len(self.repository.list_opshop_locations()))
        self.assertEqual(["MONDAY", "WEDNESDAY"], [item.run_day for item in schedules])
        self.assertEqual(1, len({item.opshop_id for item in schedules}))
        self.assertEqual(
            ["2x Weekly", "2x Weekly"],
            [item.pickup_frequency for item in schedules],
        )

    def test_rerun_updates_existing_rows_without_duplicates_and_deactivates_missing_regular(self):
        self._save_workbook(
            {
                "MON": [self._row(Op_Shop_Name="Monday Shop", Assigned_to="John G")],
                "TUE": [self._row(Op_Shop_Name="Tuesday Shop", Assigned_to="Gavin")],
            }
        )
        first_summary = import_regular_opshop_pickups_to_db(self.workbook_path, self.db_path)

        self._save_workbook(
            {
                "MON": [self._row(Op_Shop_Name="Monday Shop", Assigned_to="John G")],
            }
        )
        second_summary = import_regular_opshop_pickups_to_db(self.workbook_path, self.db_path)
        schedules = self.repository.list_opshop_pickup_schedules()
        active_schedules = [schedule for schedule in schedules if schedule.active_flag]

        self.assertEqual(2, first_summary.schedules_inserted)
        self.assertEqual(1, second_summary.schedules_updated)
        self.assertEqual(1, second_summary.schedules_deactivated)
        self.assertEqual(2, len(schedules))
        self.assertEqual(["MONDAY"], [schedule.run_day for schedule in active_schedules])
        self.assertIsNotNone(second_summary.backup_path)
        self.assertTrue(Path(second_summary.backup_path).exists())

    def test_rerun_deactivates_missing_workbook_regular_without_disabling_ui_template(self):
        ui_template = self.service.create_opshop_template(
            CreateOpShopTemplateRequest(
                run_type="REGULAR",
                run_day="FRIDAY",
                name="UI Created Regular",
                suburb="Richmond",
                street_address="9 Bridge Road",
                pickup_frequency="Weekly",
                time_window="10-12",
            )
        )
        self._save_workbook(
            {
                "MON": [self._row(Op_Shop_Name="Monday Shop", Assigned_to="John G")],
                "TUE": [self._row(Op_Shop_Name="Tuesday Shop", Assigned_to="Gavin")],
            }
        )
        import_regular_opshop_pickups_to_db(self.workbook_path, self.db_path)

        self._save_workbook(
            {
                "MON": [self._row(Op_Shop_Name="Monday Shop", Assigned_to="John G")],
            }
        )
        second_summary = import_regular_opshop_pickups_to_db(self.workbook_path, self.db_path)
        schedules_by_id = {
            schedule.schedule_id: schedule
            for schedule in self.repository.list_opshop_pickup_schedules()
        }

        self.assertEqual(1, second_summary.schedules_deactivated)
        self.assertTrue(schedules_by_id[ui_template.schedule_id].active_flag)
        self.assertEqual("Active", schedules_by_id[ui_template.schedule_id].status)

    def test_missing_required_sheet_or_column_fails_before_writing_database(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "MON"
        worksheet.append(["Op_Shop_Name", "Assigned to"])
        workbook.save(self.workbook_path)

        with self.assertRaises(ValueError):
            import_regular_opshop_pickups_to_db(self.workbook_path, self.db_path)

        self.assertEqual([], self.repository.list_opshop_pickup_schedules())

    def test_late_duplicate_schedule_conflict_rolls_back_all_source_mutations(self):
        self.repository.upsert_opshop_location(
            OpShopLocation(
                opshop_id="OPSHOP-CONFLICT",
                name="Conflict Shop",
                suburb="Coburg",
                street_address="99 Sydney Road",
                area_region="North",
                primary_contact="Existing Contact",
                primary_phone="0400 100 100",
                secondary_contact=None,
                secondary_phone=None,
                access_type="Front door",
                key_required=False,
                trailer_restriction=None,
                status_notes="Existing note",
                is_active=True,
                created_at="2026-05-01T00:00:00+00:00",
                updated_at="2026-05-01T00:00:00+00:00",
            )
        )
        for schedule_id in ("CONFLICT-A", "CONFLICT-B"):
            self.repository.upsert_opshop_pickup_schedule(
                OpShopPickupSchedule(
                    schedule_id=schedule_id,
                    opshop_id="OPSHOP-CONFLICT",
                    run_day="WEDNESDAY",
                    run_type="REGULAR",
                    pickup_frequency="Weekly",
                    time_window="09:00-12:00",
                    call_before_arrival=False,
                    call_timing=None,
                    status="Active",
                    active_flag=True,
                    fortnight_group=None,
                    review_required=False,
                    review_reason="WORKBOOK_IMPORT",
                    created_at="2026-05-01T00:00:00+00:00",
                    updated_at="2026-05-01T00:00:00+00:00",
                    default_driver_alias="John G",
                )
            )
        before = self._logical_source_tables()
        self._save_workbook(
            {
                "MON": [
                    self._row(
                        Op_Shop_Name="Earlier New Shop",
                        Street_Address="7 New Street",
                        Pickup_Frequency="Weekly",
                    )
                ],
                "WED": [
                    self._row(
                        Op_Shop_Name="Conflict Shop",
                        Street_Address="99 Sydney Road",
                        Pickup_Frequency="2x Weekly",
                    )
                ],
            }
        )

        with self.assertRaisesRegex(ValueError, "Duplicate OP SHOP schedule slot"):
            import_regular_opshop_pickups_to_db(self.workbook_path, self.db_path)

        self.assertEqual(before, self._logical_source_tables())
        self.assertIsNone(
            next(
                (
                    location
                    for location in self.repository.list_opshop_locations()
                    if location.name == "Earlier New Shop"
                ),
                None,
            )
        )

    def test_import_reuses_only_active_schedule_when_inactive_legacy_slot_exists(self):
        self.repository.upsert_opshop_location(
            OpShopLocation(
                opshop_id="OPSHOP-LEGACY-SLOT",
                name="Legacy Slot Shop",
                suburb="Coburg",
                street_address="10 Sydney Road",
                area_region="North",
                primary_contact=None,
                primary_phone=None,
                secondary_contact=None,
                secondary_phone=None,
                access_type=None,
                key_required=False,
                trailer_restriction=None,
                status_notes=None,
                is_active=True,
                created_at="2026-05-01T00:00:00+00:00",
                updated_at="2026-05-01T00:00:00+00:00",
            )
        )
        for schedule_id, active_flag, status, frequency in (
            ("LEGACY-INACTIVE", False, "On_Hold", "Fortnightly"),
            ("CURRENT-ACTIVE", True, "Active", "Weekly"),
        ):
            self.repository.upsert_opshop_pickup_schedule(
                OpShopPickupSchedule(
                    schedule_id=schedule_id,
                    opshop_id="OPSHOP-LEGACY-SLOT",
                    run_day="WEDNESDAY",
                    run_type="REGULAR",
                    pickup_frequency=frequency,
                    time_window="09:00-12:00",
                    call_before_arrival=False,
                    call_timing=None,
                    status=status,
                    active_flag=active_flag,
                    fortnight_group=None,
                    review_required=not active_flag,
                    review_reason="WORKBOOK_IMPORT",
                    created_at="2026-05-01T00:00:00+00:00",
                    updated_at="2026-05-01T00:00:00+00:00",
                    default_driver_alias=None,
                )
            )
        self._save_workbook(
            {
                "WED": [
                    self._row(
                        Op_Shop_Name="Legacy Slot Shop",
                        Street_Address="10 Sydney Road",
                        Pickup_Frequency="2x Weekly",
                    )
                ]
            }
        )

        summary = import_regular_opshop_pickups_to_db(
            self.workbook_path, self.db_path
        )

        schedules = {
            schedule.schedule_id: schedule
            for schedule in self.repository.list_opshop_pickup_schedules()
        }
        self.assertEqual(2, len(schedules))
        self.assertEqual(1, summary.schedules_updated)
        self.assertEqual("2x Weekly", schedules["CURRENT-ACTIVE"].pickup_frequency)
        self.assertTrue(schedules["CURRENT-ACTIVE"].active_flag)
        self.assertEqual("Fortnightly", schedules["LEGACY-INACTIVE"].pickup_frequency)
        self.assertFalse(schedules["LEGACY-INACTIVE"].active_flag)

    def test_duplicate_location_error_includes_normalized_key_workbook_row_and_ids(self):
        for opshop_id, address, is_active in (
            ("OPSHOP-A", "1 Sydney Road", True),
            ("OPSHOP-B", "1 Sydney Rd", False),
        ):
            self.repository.upsert_opshop_location(
                OpShopLocation(
                    opshop_id=opshop_id,
                    name="Shared Shop",
                    suburb="Coburg",
                    street_address=address,
                    area_region="North",
                    primary_contact=None,
                    primary_phone=None,
                    secondary_contact=None,
                    secondary_phone=None,
                    access_type=None,
                    key_required=False,
                    trailer_restriction=None,
                    status_notes=None,
                    is_active=is_active,
                    created_at="2026-05-01T00:00:00+00:00",
                    updated_at="2026-05-01T00:00:00+00:00",
                )
            )
        self._save_workbook(
            {
                "WED": [
                    self._row(
                        Op_Shop_Name="Shared Shop",
                        Suburb="Coburg",
                        Street_Address="1 Sydney Road",
                    )
                ]
            }
        )

        with self.assertRaises(ValueError) as raised:
            import_regular_opshop_pickups_to_db(self.workbook_path, self.db_path)

        message = str(raised.exception)
        self.assertIn("normalized_key=shared shop|coburg|1 sydney rd", message)
        self.assertIn("workbook=WED row=2", message)
        self.assertIn("workbook_location=Shared Shop|Coburg|1 Sydney Road", message)
        self.assertIn("matched_ids=OPSHOP-A, OPSHOP-B", message)
        self.assertIn("OPSHOP-A=Shared Shop|Coburg|1 Sydney Road|active=1", message)
        self.assertIn("OPSHOP-B=Shared Shop|Coburg|1 Sydney Rd|active=0", message)

    def test_exception_after_deactivation_rolls_back_entire_import_transaction(self):
        self.repository.upsert_opshop_location(
            OpShopLocation(
                opshop_id="OPSHOP-MISSING",
                name="Missing From Workbook",
                suburb="Coburg",
                street_address="8 Old Street",
                area_region="North",
                primary_contact=None,
                primary_phone=None,
                secondary_contact=None,
                secondary_phone=None,
                access_type=None,
                key_required=False,
                trailer_restriction=None,
                status_notes=None,
                is_active=True,
                created_at="2026-05-01T00:00:00+00:00",
                updated_at="2026-05-01T00:00:00+00:00",
            )
        )
        self.repository.upsert_opshop_pickup_schedule(
            OpShopPickupSchedule(
                schedule_id="SCHEDULE-MISSING",
                opshop_id="OPSHOP-MISSING",
                run_day="FRIDAY",
                run_type="REGULAR",
                pickup_frequency="Weekly",
                time_window="09:00-12:00",
                call_before_arrival=False,
                call_timing=None,
                status="Active",
                active_flag=True,
                fortnight_group=None,
                review_required=False,
                review_reason="WORKBOOK_IMPORT",
                created_at="2026-05-01T00:00:00+00:00",
                updated_at="2026-05-01T00:00:00+00:00",
                default_driver_alias="John G",
            )
        )
        before = self._logical_source_tables()
        self._save_workbook(
            {
                "MON": [
                    self._row(
                        Op_Shop_Name="Transactional New Shop",
                        Street_Address="10 New Street",
                        Pickup_Frequency="Weekly",
                    )
                ]
            }
        )

        def fail_after_deactivation(connection, schedule_ids):
            deactivate_regular_schedules(connection, schedule_ids)
            raise RuntimeError("forced failure after deactivation")

        with patch(
            "tools.import_regular_opshop_pickups_to_db.deactivate_regular_schedules",
            side_effect=fail_after_deactivation,
        ):
            with self.assertRaisesRegex(
                RuntimeError,
                "forced failure after deactivation",
            ):
                import_regular_opshop_pickups_to_db(
                    self.workbook_path,
                    self.db_path,
                )

        self.assertEqual(before, self._logical_source_tables())
        retained = self.repository.get_opshop_pickup_schedule("SCHEDULE-MISSING")
        self.assertTrue(retained.active_flag)
        self.assertEqual("Active", retained.status)

    def _save_workbook(self, rows_by_sheet):
        workbook = Workbook()
        workbook.remove(workbook.active)
        for sheet_name in ["MON", "TUE", "WED", "THU", "FRI"]:
            worksheet = workbook.create_sheet(sheet_name)
            worksheet.append(REQUIRED_COLUMNS)
            for row in rows_by_sheet.get(sheet_name, []):
                worksheet.append([row.get(column, "") for column in REQUIRED_COLUMNS])
        workbook.save(self.workbook_path)

    def _logical_source_tables(self):
        with sqlite3.connect(self.db_path) as connection:
            locations = connection.execute(
                "SELECT * FROM opshop_locations ORDER BY opshop_id"
            ).fetchall()
            schedules = connection.execute(
                "SELECT * FROM opshop_pickup_schedules ORDER BY schedule_id"
            ).fetchall()
        return {
            "opshop_locations": [tuple(row) for row in locations],
            "opshop_pickup_schedules": [tuple(row) for row in schedules],
        }

    def _row(self, **overrides):
        row = {
            "Op_Shop_Name": "Northside Op Shop",
            "Run_Day": "Ignored Workbook Day",
            "Run_Type": "Standard",
            "Active_Flag": "1",
            "Suburb": "Coburg",
            "Street_Address": "1 Sydney Road",
            "Area_Region": "North",
            "Primary_Contact": "Mary",
            "Primary_Phone": "0400 000 001",
            "Secondary_Contact": "John",
            "Secondary_Phone": "0400 000 002",
            "Pickup_Frequency": "Twice weekly (Mon & Thu)",
            "Time_Window": "9-12",
            "Call_Before_Arrival": "Yes",
            "Call_Timing": "30 minutes",
            "Access_Type": "Rear dock",
            "Key_Required": "No",
            "Trailer_Restriction": "Small truck only",
            "Status": "Active",
            "Status_Start_Date": "2026-05-19",
            "Status_Notes": "Ring first",
            "Assigned to": "John G",
        }
        normalized_overrides = {
            ("Assigned to" if key == "Assigned_to" else key): value
            for key, value in overrides.items()
        }
        row.update(normalized_overrides)
        return row


if __name__ == "__main__":
    unittest.main()
