from io import BytesIO
import shutil
import sqlite3
import unittest
import uuid
from pathlib import Path

from openpyxl import load_workbook

from backend.repositories.sqlite_manual_dispatch_repository import (
    SQLiteManualDispatchRepository,
)
from backend.schemas import AssignTaskRequest, UnassignTaskRequest
from backend.services.excel_export_service import build_manual_dispatch_excel
from backend.services.manual_dispatch_service import ManualDispatchService


class ManualDispatchCancelOrderTest(unittest.TestCase):
    def setUp(self):
        temp_parent = Path.cwd() / "tmp"
        temp_parent.mkdir(exist_ok=True)
        self.temp_dir = temp_parent / f"cancel-order-test-{uuid.uuid4().hex}"
        self.temp_dir.mkdir()
        self.db_path = self.temp_dir / "manual_dispatch.sqlite3"
        self.repository = SQLiteManualDispatchRepository(self.db_path)
        self.service = ManualDispatchService(self.repository)
        self.dispatch_date = "2026-05-05"

    def tearDown(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_cancel_unassigned_order_succeeds(self):
        cancelled = self.service.cancel_order("ORD-001")

        self.assertEqual("CANCELLED", cancelled.status)

    def test_cancelled_order_disappears_from_board(self):
        self.service.cancel_order("ORD-001")

        board = self.service.get_board(self.dispatch_date)

        self.assertNotIn("ORD-001", [order.order_id for order in board.orders])

    def test_cancelled_order_cannot_be_assigned(self):
        self.service.cancel_order("ORD-001")

        with self.assertRaises(ValueError):
            self._assign_order("ORD-001", "D001", "trip1")

    def test_assigned_order_cannot_be_cancelled_until_unassigned(self):
        self._assign_order("ORD-001", "D001", "trip1")

        with self.assertRaises(ValueError) as context:
            self.service.cancel_order("ORD-001")

        self.assertIn("unassigned", str(context.exception))

        self.service.unassign_task(
            UnassignTaskRequest(
                dispatch_date=self.dispatch_date,
                task_type="ORDER",
                task_id="ORD-001",
            )
        )
        cancelled = self.service.cancel_order("ORD-001")
        self.assertEqual("CANCELLED", cancelled.status)

    def test_cancelled_order_is_excluded_from_excel_export(self):
        self._assign_order("ORD-001", "D001", "trip1")
        with sqlite3.connect(self.db_path) as connection:
            connection.execute(
                "UPDATE manual_orders SET status = 'CANCELLED' WHERE order_id = ?",
                ("ORD-001",),
            )
            connection.commit()

        rows = self._export_rows()

        self.assertEqual(1, len(rows))
        self.assertEqual("Order ID", rows[0][4])

    def test_status_safe_migration_works_for_old_database(self):
        legacy_path = self.temp_dir / "legacy_manual_dispatch.sqlite3"
        self._create_legacy_database_without_status(legacy_path)

        repository = SQLiteManualDispatchRepository(legacy_path)
        board = ManualDispatchService(repository).get_board(self.dispatch_date)

        legacy_order = next(order for order in board.orders if order.order_id == "ORD-OLD")
        self.assertEqual("ACTIVE", legacy_order.status)

        with sqlite3.connect(legacy_path) as connection:
            columns = {
                row[1]
                for row in connection.execute("PRAGMA table_info(manual_orders)").fetchall()
            }
        self.assertIn("status", columns)

    def test_existing_active_orders_remain_visible(self):
        board = self.service.get_board(self.dispatch_date)

        self.assertEqual(
            ["ORD-001", "ORD-002", "ORD-003"],
            [order.order_id for order in board.orders],
        )
        self.assertEqual(["ACTIVE", "ACTIVE", "ACTIVE"], [order.status for order in board.orders])

    def _assign_order(self, order_id, driver_id, trip_no):
        return self.service.assign_task(
            AssignTaskRequest(
                dispatch_date=self.dispatch_date,
                task_type="ORDER",
                task_id=order_id,
                driver_id=driver_id,
                trip_no=trip_no,
            )
        )

    def _export_rows(self):
        board = self.service.get_board(self.dispatch_date)
        excel_bytes = build_manual_dispatch_excel(board, self.dispatch_date)
        workbook = load_workbook(BytesIO(excel_bytes))
        worksheet = workbook.active
        return list(worksheet.iter_rows(values_only=True))

    def _create_legacy_database_without_status(self, db_path):
        with sqlite3.connect(db_path) as connection:
            connection.executescript(
                """
                CREATE TABLE manual_orders (
                    order_id TEXT PRIMARY KEY,
                    company_name TEXT,
                    delivery_address TEXT,
                    suburb TEXT NOT NULL,
                    postcode TEXT,
                    delivery_date TEXT,
                    zone TEXT,
                    urgency TEXT,
                    preferred_driver_id TEXT,
                    pallet_quantity INTEGER NOT NULL DEFAULT 0,
                    loose_bags_quantity INTEGER NOT NULL DEFAULT 0,
                    start_time TEXT,
                    end_time TEXT,
                    note TEXT
                );

                INSERT INTO manual_orders (
                    order_id,
                    company_name,
                    delivery_address,
                    suburb,
                    postcode,
                    delivery_date,
                    zone,
                    urgency,
                    preferred_driver_id,
                    pallet_quantity,
                    loose_bags_quantity,
                    start_time,
                    end_time,
                    note
                ) VALUES (
                    'ORD-OLD',
                    'Legacy Customer',
                    '9 Legacy Street',
                    'Moorabbin',
                    '3189',
                    '2026-05-05',
                    'South East',
                    'Normal',
                    NULL,
                    1,
                    0,
                    '08:00',
                    '10:00',
                    'Legacy row'
                );
                """
            )
            connection.commit()


if __name__ == "__main__":
    unittest.main()
