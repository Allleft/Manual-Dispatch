from io import BytesIO
import shutil
import unittest
import uuid
from unittest.mock import patch
from pathlib import Path

from openpyxl import load_workbook

from backend.repositories.sqlite_manual_dispatch_repository import (
    SQLiteManualDispatchRepository,
)
from backend.schemas import AssignDriverVehicleRequest, AssignTaskRequest
from backend.services.excel_export_service import (
    EXPORT_HEADERS,
    NO_VEHICLE_SELECTED,
    build_manual_dispatch_excel,
)
from backend.services.manual_dispatch_service import ManualDispatchService


class ManualDispatchExcelExportTest(unittest.TestCase):
    def setUp(self):
        temp_parent = Path.cwd() / "tmp"
        temp_parent.mkdir(exist_ok=True)
        self.temp_dir = temp_parent / f"excel-export-test-{uuid.uuid4().hex}"
        self.temp_dir.mkdir()
        self.db_path = self.temp_dir / "manual_dispatch.sqlite3"
        with patch.dict("os.environ", {"MANUAL_DISPATCH_SEED_DEMO_DATA": "true"}):
            self.repository = SQLiteManualDispatchRepository(self.db_path)
        self.service = ManualDispatchService(self.repository)
        self.dispatch_date = "2026-05-05"

    def tearDown(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_export_workbook_includes_expected_headers_without_location(self):
        rows = self._export_rows()

        self.assertEqual(tuple(EXPORT_HEADERS), rows[0])
        self.assertNotIn("Location", rows[0])

    def test_export_includes_only_assigned_orders(self):
        self._assign_order("ORD-001", "D001", "trip1")
        self._assign_vehicle("D001", "V002")

        rows = self._export_rows()
        data_rows = rows[1:]

        self.assertEqual(1, len(data_rows))
        self.assertEqual("ORD-001", data_rows[0][4])
        self.assertEqual("John", data_rows[0][1])
        self.assertEqual("XYZ888", data_rows[0][2])
        self.assertEqual("trip1", data_rows[0][3])
        self.assertNotIn("ORD-002", [row[4] for row in data_rows])

    def test_export_includes_trip_rows_only_when_orders_are_assigned(self):
        self._assign_order("ORD-001", "D001", "trip1")
        self._assign_order("ORD-003", "D001", "trip2")

        rows = self._export_rows()
        data_rows = rows[1:]

        self.assertEqual(["trip1", "trip2"], [row[3] for row in data_rows])
        self.assertEqual(["ORD-001", "ORD-003"], [row[4] for row in data_rows])
        self.assertEqual({"John"}, {row[1] for row in data_rows})
        self.assertNotIn("Tony", {row[1] for row in data_rows})

    def test_export_empty_board_has_header_only(self):
        rows = self._export_rows()

        self.assertEqual(1, len(rows))
        self.assertEqual(tuple(EXPORT_HEADERS), rows[0])

    def test_export_uses_no_vehicle_selected_when_driver_has_no_vehicle(self):
        self._assign_order("ORD-002", "D002", "trip1")

        rows = self._export_rows()
        data_rows = rows[1:]

        self.assertEqual(1, len(data_rows))
        self.assertEqual(NO_VEHICLE_SELECTED, data_rows[0][2])

    def test_export_sorts_by_driver_trip_and_order_id(self):
        self._assign_order("ORD-003", "D003", "trip2")
        self._assign_order("ORD-001", "D001", "trip2")
        self._assign_order("ORD-002", "D001", "trip1")

        rows = self._export_rows()
        data_rows = rows[1:]

        self.assertEqual(
            [
                ("David", "trip2", "ORD-003"),
                ("John", "trip1", "ORD-002"),
                ("John", "trip2", "ORD-001"),
            ],
            [(row[1], row[3], row[4]) for row in data_rows],
        )

    def test_export_preferred_driver_displays_name_when_known(self):
        self._assign_order("ORD-001", "D002", "trip1")

        rows = self._export_rows()
        data_rows = rows[1:]

        self.assertEqual("John", data_rows[0][11])

    def _assign_order(self, order_id, driver_id, trip_no):
        self.service.assign_task(
            AssignTaskRequest(
                dispatch_date=self.dispatch_date,
                task_type="ORDER",
                task_id=order_id,
                driver_id=driver_id,
                trip_no=trip_no,
            )
        )

    def _assign_vehicle(self, driver_id, vehicle_id):
        self.service.assign_vehicle_to_driver(
            AssignDriverVehicleRequest(
                dispatch_date=self.dispatch_date,
                driver_id=driver_id,
                vehicle_id=vehicle_id,
            )
        )

    def _export_rows(self):
        board = self.service.get_board(self.dispatch_date)
        excel_bytes = build_manual_dispatch_excel(board, self.dispatch_date)
        workbook = load_workbook(BytesIO(excel_bytes))
        worksheet = workbook.active
        return list(worksheet.iter_rows(values_only=True))


if __name__ == "__main__":
    unittest.main()
