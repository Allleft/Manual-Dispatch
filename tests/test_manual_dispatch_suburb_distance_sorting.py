from io import BytesIO
import shutil
import unittest
import uuid
from pathlib import Path

from openpyxl import load_workbook

from backend.repositories.sqlite_manual_dispatch_repository import (
    SQLiteManualDispatchRepository,
)
from backend.schemas import (
    AssignTaskRequest,
    CreateOrderRequest,
    RegisterOperatorAccountRequest,
    SaveFinalTripSummaryRequest,
)
from backend.services.final_summary_excel_export_service import (
    build_final_summary_excel,
)
from backend.services.manual_dispatch.suburb_distance_service import (
    WAREHOUSE_ORIGIN,
    get_estimated_distance_km,
    normalize_suburb_name,
)
from backend.services.manual_dispatch_service import ManualDispatchService


class ManualDispatchSuburbDistanceSortingTest(unittest.TestCase):
    def setUp(self):
        temp_parent = Path.cwd() / "tmp"
        temp_parent.mkdir(exist_ok=True)
        self.temp_dir = temp_parent / f"suburb-distance-test-{uuid.uuid4().hex}"
        self.temp_dir.mkdir()
        self.db_path = self.temp_dir / "manual_dispatch.sqlite3"
        self.repository = SQLiteManualDispatchRepository(self.db_path)
        self.service = ManualDispatchService(self.repository)
        self.dispatch_date = "2026-05-05"
        self.account = self.service.register_operator_account(
            RegisterOperatorAccountRequest(
                account_name="Mandy",
                password="secret123",
                confirm_password="secret123",
            )
        )

    def tearDown(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_suburb_name_normalization_and_lookup_are_stable(self):
        self.assertEqual("clayton", normalize_suburb_name("clayton"))
        self.assertEqual("clayton", normalize_suburb_name(" Clayton "))
        self.assertEqual("clayton", normalize_suburb_name("CLAYTON"))
        self.assertEqual(42.0, get_estimated_distance_km("Clayton"))
        self.assertIsNone(get_estimated_distance_km("Unmapped Ridge"))
        self.assertEqual(
            "98-102 Hume Hwy, Somerton, VIC, 3062",
            WAREHOUSE_ORIGIN,
        )

    def test_expanded_suburbs_and_aliases_resolve_to_known_distances(self):
        expanded_suburbs = [
            "Dandenong South",
            "Pakenham",
            "Ballarat Central",
            "Bendigo",
            "Footscray",
            "Northcote",
            "Box Hill",
            "Essendon",
            "Shepparton",
        ]
        for suburb in expanded_suburbs:
            with self.subTest(suburb=suburb):
                self.assertIsNotNone(get_estimated_distance_km(suburb))

        self.assertEqual(
            get_estimated_distance_km("Dandenong South"),
            get_estimated_distance_km("  dandenong south  "),
        )
        self.assertEqual(
            get_estimated_distance_km("Dandenong South"),
            get_estimated_distance_km("DANDENONG SOUTH"),
        )
        self.assertEqual(
            get_estimated_distance_km("Ballarat Central"),
            get_estimated_distance_km("Ballarat   Central"),
        )
        self.assertEqual(
            get_estimated_distance_km("Dandenong South"),
            get_estimated_distance_km("Dandenong Sth"),
        )
        self.assertEqual(
            get_estimated_distance_km("Melbourne"),
            get_estimated_distance_km("Melbourne CBD"),
        )
        self.assertEqual(
            get_estimated_distance_km("Melbourne"),
            get_estimated_distance_km("CBD"),
        )

    def test_board_orders_expose_estimated_suburb_distance(self):
        created = self._create_order("BOARD-DISTANCE", "Clayton", "09:00")
        board_order = next(
            order
            for order in self.service.get_board(self.dispatch_date).orders
            if order.order_id == created.order_id
        )
        self.assertEqual(42.0, board_order.estimated_distance_km_from_warehouse)

    def test_final_summary_sorts_by_distance_same_suburb_start_time_and_unknown_last(self):
        unknown = self._create_order("INV-UNKNOWN", "Unmapped Ridge", "07:00")
        coburg_late = self._create_order("INV-COBURG-LATE", "Coburg", "10:30")
        campbellfield = self._create_order(
            "INV-CAMPBELLFIELD",
            "Campbellfield",
            "09:00",
        )
        coburg_missing = self._create_order("INV-COBURG-MISSING", "Coburg", "")
        coburg_early = self._create_order("INV-COBURG-EARLY", "Coburg", "08:30")

        for order in (
            unknown,
            coburg_late,
            campbellfield,
            coburg_missing,
            coburg_early,
        ):
            self.service.assign_task(
                AssignTaskRequest(
                    dispatch_date=self.dispatch_date,
                    task_type="ORDER",
                    task_id=order.order_id,
                    driver_id="D001",
                    trip_no="trip1",
                )
            )

        saved = self.service.save_final_trip_summary(
            self._summary_request(
                [
                    unknown,
                    coburg_late,
                    campbellfield,
                    coburg_missing,
                    coburg_early,
                ]
            )
        )
        saved_orders = saved.trips[0].orders

        self.assertEqual(
            [
                campbellfield.order_id,
                coburg_early.order_id,
                coburg_late.order_id,
                coburg_missing.order_id,
                unknown.order_id,
            ],
            [order.task_id for order in saved_orders],
        )
        self.assertEqual(
            [4.0, 14.0, 14.0, 14.0, None],
            [
                order.estimated_distance_km_from_warehouse_snapshot
                for order in saved_orders
            ],
        )

        history = self.service.list_final_trip_summaries(self.dispatch_date)
        self.assertEqual(
            4.0,
            history[0].trips[0].orders[0].estimated_distance_km_from_warehouse_snapshot,
        )

    def test_excel_export_keeps_sorted_order_and_distance_column(self):
        clayton = self._create_order("INV-CLAYTON", "Clayton", "09:30")
        coburg = self._create_order("INV-COBURG", "Coburg", "08:30")
        campbellfield = self._create_order("INV-CAMP", "Campbellfield", "09:00")

        saved = self.service.save_final_trip_summary(
            self._summary_request([clayton, coburg, campbellfield])
        )
        workbook = load_workbook(
            BytesIO(build_final_summary_excel([saved], self.dispatch_date)),
            data_only=True,
        )
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        flattened = [cell for row in rows for cell in row if cell is not None]

        self.assertIn("Estimated Distance From Warehouse (km)", flattened)

        data_rows = [
            row
            for row in rows
            if isinstance(row[0], int)
        ]
        self.assertEqual(
            ["Campbellfield", "Coburg", "Clayton"],
            [row[2] for row in data_rows],
        )
        self.assertEqual([4.0, 14.0, 42.0], [row[3] for row in data_rows])

    def test_expanded_suburbs_keep_distances_in_saved_summary_and_excel(self):
        orders = [
            self._create_order("INV-DANDY-SOUTH", "Dandenong South", "09:00"),
            self._create_order("INV-PAKENHAM", "Pakenham", "10:00"),
            self._create_order("INV-BALLARAT-CENTRAL", "Ballarat Central", "11:00"),
        ]
        saved = self.service.save_final_trip_summary(self._summary_request(orders))
        saved_orders = saved.trips[0].orders

        self.assertEqual(3, len(saved_orders))
        self.assertTrue(
            all(
                order.estimated_distance_km_from_warehouse_snapshot is not None
                for order in saved_orders
            )
        )

        workbook = load_workbook(
            BytesIO(build_final_summary_excel([saved], self.dispatch_date)),
            data_only=True,
        )
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        data_rows = [row for row in rows if isinstance(row[0], int)]
        suburb_to_distance = {row[2]: row[3] for row in data_rows}

        self.assertIsInstance(suburb_to_distance["Dandenong South"], (int, float))
        self.assertIsInstance(suburb_to_distance["Pakenham"], (int, float))
        self.assertIsInstance(suburb_to_distance["Ballarat Central"], (int, float))

    def _create_order(self, invoice_number, suburb, start_time):
        return self.service.create_order(
            CreateOrderRequest(
                invoice_number=invoice_number,
                company_name=f"{suburb} Customer",
                phone="0399991800",
                delivery_address=f"1 {suburb} Street",
                suburb=suburb,
                postcode="3000",
                delivery_date=self.dispatch_date,
                zone="Metro",
                urgency="Normal",
                preferred_driver_id="",
                pallet_quantity=1,
                loose_bags_quantity=0,
                start_time=start_time,
                end_time="12:00",
                note="distance sorting fixture",
            )
        )

    def _summary_request(self, orders):
        return SaveFinalTripSummaryRequest(
            dispatch_date=self.dispatch_date,
            delivery_date=self.dispatch_date,
            driver_id="D001",
            driver_name_snapshot="John",
            vehicle_id=None,
            vehicle_rego_snapshot="No vehicle selected",
            total_pallets=0,
            total_loose_bags=0,
            generated_at="2026-05-05T00:00:00Z",
            saved_by_account_name=self.account.account_name,
            saved_by_account_id=self.account.account_id,
            trips=[
                {
                    "trip_no": "trip1",
                    "orders": [self._order_payload(order) for order in orders],
                }
            ],
        )

    def _order_payload(self, order):
        return {
            "task_type": "ORDER",
            "task_id": order.order_id,
            "order_id_snapshot": order.order_id,
            "invoice_number_snapshot": order.invoice_number,
            "company_name_snapshot": order.company_name,
            "suburb_snapshot": order.suburb,
            "delivery_address_snapshot": order.delivery_address,
            "product_snapshot": "",
            "pallet_quantity_snapshot": order.pallet_quantity,
            "loose_bags_quantity_snapshot": order.loose_bags_quantity,
            "note_snapshot": order.note,
            "start_time": order.start_time,
        }
