import importlib
import os
import shutil
import unittest
import uuid
from io import BytesIO
from pathlib import Path

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.repositories.in_memory_manual_dispatch_repository import (
    InMemoryManualDispatchRepository,
)
from backend.repositories.sqlite_manual_dispatch_repository import (
    SQLiteManualDispatchRepository,
)
from backend.schemas import (
    Driver,
    GenerateOpShopPickupCollectionRequest,
    OpShopLocation,
    OpShopPickupSchedule,
    OpShopPickupTask,
    RegisterOperatorAccountRequest,
)
from backend.services.manual_dispatch_service import ManualDispatchService
from backend.services.opshop_pickup_collection_excel_export_service import (
    build_opshop_pickup_collection_excel,
)
from tests.manual_dispatch_api_test_helpers import authenticate_test_client


class OpShopPickupCollectionGenerationTest(unittest.TestCase):
    def setUp(self):
        temp_parent = Path.cwd() / "tmp"
        temp_parent.mkdir(exist_ok=True)
        self.temp_dir = temp_parent / f"opshop-collection-generation-{uuid.uuid4().hex}"
        self.temp_dir.mkdir()
        self.db_path = self.temp_dir / "manual_dispatch.sqlite3"
        self.previous_db_path = os.environ.get("MANUAL_DISPATCH_DB_PATH")
        self.previous_seed_flag = os.environ.get("MANUAL_DISPATCH_SEED_DEMO_DATA")
        os.environ["MANUAL_DISPATCH_DB_PATH"] = str(self.db_path)
        os.environ["MANUAL_DISPATCH_SEED_DEMO_DATA"] = "0"

        self.repository = SQLiteManualDispatchRepository(self.db_path)
        self.repository.create_driver(
            Driver(
                driver_id="D001",
                name="John Georgiadis",
                start_time="07:00",
                end_time="15:00",
                is_available=True,
                preferred_zone=None,
                pallet_only=False,
            )
        )
        self.repository.create_driver(
            Driver(
                driver_id="D002",
                name="Second Driver",
                start_time="07:00",
                end_time="15:00",
                is_available=True,
                preferred_zone=None,
                pallet_only=False,
            )
        )
        self.service = ManualDispatchService(self.repository)
        self.dispatch_date = "2026-07-05"
        self.pickup_date = "2026-07-06"
        self.account = self.service.register_operator_account(
            RegisterOperatorAccountRequest(
                account_name="Collection QA",
                password="secret123",
                confirm_password="secret123",
            )
        )

        self.api_module = importlib.import_module("backend.api.manual_dispatch")
        self.original_service = self.api_module.service
        self.api_module.service = self.service
        app = FastAPI()
        app.include_router(self.api_module.router)
        self.client = TestClient(app)
        authenticate_test_client(self.client, self.service, self.account)

    def tearDown(self):
        self.api_module.service = self.original_service
        self._restore_environment("MANUAL_DISPATCH_DB_PATH", self.previous_db_path)
        self._restore_environment(
            "MANUAL_DISPATCH_SEED_DEMO_DATA",
            self.previous_seed_flag,
        )
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_five_source_backed_regular_pickups_generate_cancel_and_save(self):
        for number in range(1, 6):
            self._seed_template(
                f"REGRESSION-{number}",
                run_type="REGULAR",
                pickup_category="NORMAL",
                default_driver_id="D001",
            )

        board_response = self.client.get(
            "/api/manual-dispatch/opshop/board",
            params={"dispatch_date": self.dispatch_date},
        )
        self.assertEqual(200, board_response.status_code)
        visible = self._visible_pickups(board_response.json(), "D001")
        self.assertEqual(5, len(visible))
        self.assertEqual({"REGULAR"}, {pickup["run_type"] for pickup in visible})
        for pickup in visible:
            assignment = self.repository.get_assignment(
                self.pickup_date,
                "OPSHOP_PICKUP",
                pickup["pickup_task_id"],
            )
            self.assertIsNotNone(assignment)
            self.assertEqual("D001", assignment.driver_id)

        generated = self._post_generate("D001")
        self.assertEqual(200, generated.status_code)
        generated_body = generated.json()
        self.assertEqual("John Georgiadis", generated_body["driver_name_snapshot"])
        self.assertEqual(5, len(generated_body["pickups"]))
        self.assertEqual(
            {pickup["pickup_task_id"] for pickup in visible},
            {
                pickup["pickup_task_id_snapshot"]
                for pickup in generated_body["pickups"]
            },
        )

        first_location = self.repository.get_opshop_location(
            self.repository.get_opshop_pickup_task(
                visible[0]["pickup_task_id"]
            ).opshop_id
        )
        snapshot_name = generated_body["pickups"][0]["opshop_name_snapshot"]
        first_location.name = "Edited after generation"
        self.repository.upsert_opshop_location(first_location)
        persisted = self.client.get(
            f"/api/manual-dispatch/opshop/pickup-collections/{generated_body['collection_id']}"
        ).json()
        self.assertEqual(
            snapshot_name,
            persisted["pickups"][0]["opshop_name_snapshot"],
        )

        duplicate = self._post_generate("D001")
        self.assertEqual(409, duplicate.status_code)
        cancelled = self.client.post(
            "/api/manual-dispatch/opshop/pickup-collections/"
            f"{generated_body['collection_id']}/cancel-generated"
        )
        self.assertEqual({"cancelled": True}, cancelled.json())
        self.assertEqual(5, len(self._visible_pickups(self._get_board(), "D001")))

        regenerated = self._post_generate("D001")
        self.assertEqual(200, regenerated.status_code)
        saved = self.client.post(
            "/api/manual-dispatch/opshop/pickup-collections/"
            f"{regenerated.json()['collection_id']}/save",
            json={
                "saved_by_account_name": self.account.account_name,
                "saved_by_account_id": self.account.account_id,
            },
        )
        self.assertEqual(200, saved.status_code)
        self.assertEqual("SAVED", saved.json()["status"])
        self.assertEqual(409, self._post_generate("D001").status_code)

    def test_generated_reservation_uses_task_identity_across_dispatch_dates(self):
        self.dispatch_date = "2026-08-17"
        self.pickup_date = "2026-08-18"
        task_id = "CROSS-DISPATCH-PICKUP"
        self._seed_pickup(
            task_id,
            driver_id="D001",
            assignment_dispatch_date=self.dispatch_date,
        )

        generated = self._post_generate("D001")
        self.assertEqual(200, generated.status_code, generated.text)
        generated_body = generated.json()
        self.assertEqual(self.dispatch_date, generated_body["dispatch_date"])
        self.assertEqual(self.pickup_date, generated_body["pickup_date"])
        self.assertEqual(
            [generated_body["collection_id"]],
            [
                collection.collection_id
                for collection in self.repository.list_opshop_pickup_collection_reservations_for_task_ids(
                    {task_id}
                )
            ],
        )

        for board_date in (self.dispatch_date, self.pickup_date):
            with self.subTest(board_date=board_date):
                item = self._board_pickup(task_id, board_date)
                self.assertEqual("D001", item["driver_id"])
                self.assertEqual("D001", item["assigned_driver_id"])
                self.assertEqual("John Georgiadis", item["assigned_driver_name"])
                self.assertTrue(item["assigned_to_locked"])
                self.assertEqual(
                    "Already generated to John Georgiadis",
                    item["assignment_lock_reason"],
                )

        reloaded_service = ManualDispatchService(
            SQLiteManualDispatchRepository(self.db_path)
        )
        reloaded_item = next(
            item
            for item in reloaded_service.get_opshop_workspace_board(
                self.pickup_date
            ).opshop_pickups
            if item.pickup_task_id == task_id
        )
        self.assertTrue(reloaded_item.assigned_to_locked)
        self.assertEqual("John Georgiadis", reloaded_item.assigned_driver_name)

        cancelled = self.client.post(
            "/api/manual-dispatch/opshop/pickup-collections/"
            f"{generated_body['collection_id']}/cancel-generated"
        )
        self.assertEqual({"cancelled": True}, cancelled.json())
        for board_date in (self.dispatch_date, self.pickup_date):
            released = self._board_pickup(task_id, board_date)
            self.assertFalse(released["assigned_to_locked"])
            self.assertIsNone(released["assignment_lock_reason"])
            self.assertEqual("D001", released["assigned_driver_id"])

        regenerated = self._post_generate("D001")
        self.assertEqual(200, regenerated.status_code, regenerated.text)
        saved = self.client.post(
            "/api/manual-dispatch/opshop/pickup-collections/"
            f"{regenerated.json()['collection_id']}/save",
            json={
                "saved_by_account_name": self.account.account_name,
                "saved_by_account_id": self.account.account_id,
            },
        )
        self.assertEqual(200, saved.status_code, saved.text)
        for board_date in (self.dispatch_date, self.pickup_date):
            self.assertNotIn(
                task_id,
                {
                    pickup["pickup_task_id"]
                    for pickup in self._get_board(board_date)["opshop_pickups"]
                },
            )

    def test_missing_assignment_is_not_rendered_as_assigned_or_collectable(self):
        self._seed_pickup(
            "MISSING-ASSIGNMENT",
            driver_id="D001",
            assignment_dispatch_date=None,
        )

        item = next(
            pickup
            for pickup in self._get_board()["opshop_pickups"]
            if pickup["pickup_task_id"] == "MISSING-ASSIGNMENT"
        )
        self.assertIsNone(item["driver_id"])
        self.assertIsNone(item["assigned_driver_id"])
        self.assertFalse(item["is_assigned"])
        response = self._post_generate("D001")
        self.assertEqual(400, response.status_code)
        self.assertIn("assigned OP SHOP pickup", response.json()["detail"])

    def test_manual_unassign_and_reassign_follow_persisted_task_driver(self):
        self._seed_pickup(
            "REASSIGN-PICKUP",
            driver_id="D001",
            assignment_dispatch_date=self.dispatch_date,
        )

        unassigned = self.client.post(
            "/api/manual-dispatch/opshop/pickups/assignments/unassign",
            json={
                "dispatch_date": self.dispatch_date,
                "pickup_task_id": "REASSIGN-PICKUP",
            },
        )
        self.assertEqual(200, unassigned.status_code)
        self.assertEqual(400, self._post_generate("D001").status_code)

        reassigned = self.client.post(
            "/api/manual-dispatch/opshop/pickups/assignments/apply",
            json={
                "dispatch_date": self.dispatch_date,
                "assignments": [
                    {"pickup_task_id": "REASSIGN-PICKUP", "driver_id": "D002"}
                ],
            },
        )
        self.assertEqual(200, reassigned.status_code)
        item = next(
            pickup
            for pickup in reassigned.json()["opshop_pickups"]
            if pickup["pickup_task_id"] == "REASSIGN-PICKUP"
        )
        self.assertEqual("D002", item["driver_id"])
        self.assertEqual(400, self._post_generate("D001").status_code)
        collection = self._post_generate("D002")
        self.assertEqual(200, collection.status_code)
        self.assertEqual(
            ["REASSIGN-PICKUP"],
            [
                pickup["pickup_task_id_snapshot"]
                for pickup in collection.json()["pickups"]
            ],
        )

    def test_oncall_and_countryside_pickups_generate_across_assignment_dates(self):
        self._seed_pickup(
            "ONCALL-PICKUP",
            driver_id="D001",
            assignment_dispatch_date=self.pickup_date,
            run_type="ON_CALL",
            pickup_category="NORMAL",
        )
        self._seed_pickup(
            "COUNTRYSIDE-PICKUP",
            driver_id="D002",
            assignment_dispatch_date=self.dispatch_date,
            run_type="ON_CALL",
            pickup_category="COUNTRYSIDE",
        )

        oncall = self._post_generate("D001")
        countryside = self._post_generate("D002")
        self.assertEqual(200, oncall.status_code)
        self.assertEqual(200, countryside.status_code)
        self.assertEqual("ON_CALL", oncall.json()["pickups"][0]["run_type_snapshot"])
        self.assertEqual(
            "COUNTRYSIDE",
            countryside.json()["pickups"][0]["pickup_category_snapshot"],
        )

    def test_in_memory_repository_uses_same_collectable_rule(self):
        repository = InMemoryManualDispatchRepository()
        service = ManualDispatchService(repository)
        self._seed_pickup(
            "MEMORY-PICKUP",
            driver_id="D001",
            assignment_dispatch_date=self.pickup_date,
            repository=repository,
        )

        self.assertEqual(
            ["MEMORY-PICKUP"],
            [
                pickup.pickup_task_id
                for pickup in repository.list_collectable_opshop_pickup_board_items(
                    self.pickup_date,
                    "D001",
                )
            ],
        )
        service.create_generated_opshop_pickup_collection(
            GenerateOpShopPickupCollectionRequest(
                dispatch_date=self.dispatch_date,
                pickup_date=self.pickup_date,
                driver_id="D001",
            )
        )
        self.assertEqual(
            [],
            repository.list_collectable_opshop_pickup_board_items(
                self.pickup_date,
                "D001",
            ),
        )
        self.assertEqual(
            ["MEMORY-PICKUP"],
            [
                row.pickup_task_id_snapshot
                for collection in repository.list_opshop_pickup_collection_reservations_for_task_ids(
                    {"MEMORY-PICKUP", "NOT-CAPTURED"}
                )
                for row in collection.pickups
            ],
        )

    def test_sqlite_and_in_memory_collection_route_order_match(self):
        in_memory = InMemoryManualDispatchRepository()
        if not in_memory.get_driver("D001"):
            in_memory.create_driver(
                Driver(
                    driver_id="D001",
                    name="John Georgiadis",
                    start_time="07:00",
                    end_time="15:00",
                    is_available=True,
                    preferred_zone=None,
                    pallet_only=False,
                )
            )
        expected = ["ROUTE-1", "ROUTE-2", "ROUTE-NULL", "ONCALL"]
        for repository in (self.repository, in_memory):
            self._seed_pickup(
                "ROUTE-2",
                "D001",
                self.dispatch_date,
                repository=repository,
                suburb="AAAA",
                regular_route_sequence=2,
            )
            self._seed_pickup(
                "ROUTE-1",
                "D001",
                self.dispatch_date,
                repository=repository,
                suburb="ZZZZ",
                regular_route_sequence=1,
            )
            self._seed_pickup(
                "ROUTE-NULL",
                "D001",
                self.dispatch_date,
                repository=repository,
                suburb="AAAA",
            )
            self._seed_pickup(
                "ONCALL",
                "D001",
                self.dispatch_date,
                run_type="ON_CALL",
                repository=repository,
                suburb="AAAA",
            )
            collection = ManualDispatchService(
                repository
            ).create_generated_opshop_pickup_collection(
                GenerateOpShopPickupCollectionRequest(
                    dispatch_date=self.dispatch_date,
                    pickup_date=self.pickup_date,
                    driver_id="D001",
                )
            )
            self.assertEqual(
                expected,
                [row.pickup_task_id_snapshot for row in collection.pickups],
            )

    def test_route_sequence_becomes_immutable_collection_and_excel_row_order(self):
        first_schedule = self._seed_pickup(
            "POSH-FIRST",
            driver_id="D001",
            assignment_dispatch_date=self.dispatch_date,
            opshop_name="POSH OPP SHOPPE",
            suburb="GLENHUNTLY",
            regular_route_sequence=1,
        )
        second_schedule = self._seed_pickup(
            "POSH-SECOND",
            driver_id="D001",
            assignment_dispatch_date=self.dispatch_date,
            opshop_name="POSH OPP SHOPPE",
            suburb="ELSTERNWICK",
            regular_route_sequence=2,
        )
        st_james_schedule = self._seed_pickup(
            "ST-JAMES",
            driver_id="D001",
            assignment_dispatch_date=self.dispatch_date,
            opshop_name="ST JAMES OP SHOP",
            suburb="MALVERN",
        )
        self._seed_pickup(
            "ONCALL-FIRST-ALPHABETICALLY",
            driver_id="D001",
            assignment_dispatch_date=self.dispatch_date,
            run_type="ON_CALL",
            opshop_name="AAA ONCALL OP SHOP",
            suburb="AAAA",
        )

        generated = self._post_generate("D001")
        self.assertEqual(200, generated.status_code, generated.text)
        expected_task_ids = [
            "POSH-FIRST",
            "POSH-SECOND",
            "ST-JAMES",
            "ONCALL-FIRST-ALPHABETICALLY",
        ]
        self.assertEqual(
            expected_task_ids,
            [row["pickup_task_id_snapshot"] for row in generated.json()["pickups"]],
        )
        self.assertEqual([1, 2, 3, 4], [row["row_no"] for row in generated.json()["pickups"]])

        for schedule_id, sequence in (
            (first_schedule, 20),
            (second_schedule, 10),
            (st_james_schedule, 5),
        ):
            schedule = self.repository.get_opshop_pickup_schedule(schedule_id)
            schedule.regular_route_sequence = sequence
            self.repository.upsert_opshop_pickup_schedule(schedule)

        collection_id = generated.json()["collection_id"]
        persisted = self.client.get(
            f"/api/manual-dispatch/opshop/pickup-collections/{collection_id}"
        ).json()
        self.assertEqual(
            expected_task_ids,
            [row["pickup_task_id_snapshot"] for row in persisted["pickups"]],
        )

        saved = self.client.post(
            f"/api/manual-dispatch/opshop/pickup-collections/{collection_id}/save",
            json={
                "saved_by_account_name": self.account.account_name,
                "saved_by_account_id": self.account.account_id,
            },
        )
        self.assertEqual(200, saved.status_code, saved.text)
        saved_collection = self.service.get_opshop_pickup_collection(collection_id)
        self.assertEqual(
            expected_task_ids,
            [row.pickup_task_id_snapshot for row in saved_collection.pickups],
        )
        worksheet = load_workbook(
            BytesIO(build_opshop_pickup_collection_excel(saved_collection))
        ).active
        self.assertEqual(
            [
                "POSH OPP SHOPPE",
                "POSH OPP SHOPPE",
                "ST JAMES OP SHOP",
                "AAA ONCALL OP SHOP",
            ],
            [worksheet.cell(row=row, column=1).value for row in range(12, 16)],
        )

    def test_countryside_trip_sequence_becomes_immutable_row_and_excel_order(self):
        self._seed_pickup(
            "REGULAR-FIRST",
            "D001",
            self.dispatch_date,
            opshop_name="Regular First",
            suburb="ZZZZ",
            regular_route_sequence=1,
        )
        self._seed_pickup(
            "COUNTRY-A",
            "D001",
            self.dispatch_date,
            run_type="ON_CALL",
            pickup_category="COUNTRYSIDE",
            opshop_name="Country A",
            suburb="ZZZZ",
            trip_sequence=2,
        )
        self._seed_pickup(
            "COUNTRY-B",
            "D001",
            self.dispatch_date,
            run_type="ON_CALL",
            pickup_category="COUNTRYSIDE",
            opshop_name="Country B",
            suburb="AAAA",
            trip_sequence=3,
        )
        self._seed_pickup(
            "COUNTRY-C",
            "D001",
            self.dispatch_date,
            run_type="ON_CALL",
            pickup_category="COUNTRYSIDE",
            opshop_name="Country C",
            suburb="MMMM",
            trip_sequence=1,
        )
        self._seed_pickup(
            "ONCALL-MIDDLE",
            "D001",
            self.dispatch_date,
            run_type="ON_CALL",
            opshop_name="Oncall Middle",
            suburb="HHHH",
        )

        trip_summary = self.service.get_opshop_trip_summary_board(self.pickup_date)
        self.assertEqual(
            ["COUNTRY-C", "COUNTRY-A", "COUNTRY-B"],
            [
                pickup.pickup_task_id
                for pickup in trip_summary.opshop_pickups
                if pickup.pickup_category == "COUNTRYSIDE"
            ],
        )

        generated = self._post_generate("D001")
        self.assertEqual(200, generated.status_code, generated.text)
        expected_task_ids = [
            "REGULAR-FIRST",
            "COUNTRY-C",
            "ONCALL-MIDDLE",
            "COUNTRY-A",
            "COUNTRY-B",
        ]
        self.assertEqual(
            expected_task_ids,
            [row["pickup_task_id_snapshot"] for row in generated.json()["pickups"]],
        )
        self.assertEqual(
            [1, 2, 3, 4, 5],
            [row["row_no"] for row in generated.json()["pickups"]],
        )

        for task_id, next_sequence in (
            ("COUNTRY-A", 1),
            ("COUNTRY-B", 2),
            ("COUNTRY-C", 3),
        ):
            task = self.repository.get_opshop_pickup_task(task_id)
            task.trip_sequence = next_sequence
            self.repository.upsert_opshop_pickup_task(task)

        collection_id = generated.json()["collection_id"]
        persisted = self.service.get_opshop_pickup_collection(collection_id)
        self.assertEqual(
            expected_task_ids,
            [row.pickup_task_id_snapshot for row in persisted.pickups],
        )
        worksheet = load_workbook(
            BytesIO(build_opshop_pickup_collection_excel(persisted))
        ).active
        self.assertEqual(
            [
                "Regular First",
                "Country C",
                "Oncall Middle",
                "Country A",
                "Country B",
            ],
            [worksheet.cell(row=row, column=1).value for row in range(12, 17)],
        )

    def _seed_template(
        self,
        suffix,
        run_type,
        pickup_category,
        default_driver_id=None,
        repository=None,
        opshop_name=None,
        suburb="MELBOURNE",
        regular_route_sequence=None,
    ):
        repository = repository or self.repository
        opshop_id = f"OPSHOP-{suffix}"
        schedule_id = f"SCHEDULE-{suffix}"
        repository.upsert_opshop_location(
            OpShopLocation(
                opshop_id=opshop_id,
                name=opshop_name or f"QA OP SHOP {suffix}",
                suburb=suburb,
                street_address=f"{suffix} TEST STREET",
                area_region="QA",
                primary_contact="QA Contact",
                primary_phone="0400 000 000",
                secondary_contact=None,
                secondary_phone=None,
                access_type="Front door",
                key_required=False,
                trailer_restriction="No",
                status_notes="QA only",
                is_active=True,
                created_at="2026-07-05T00:00:00+00:00",
                updated_at="2026-07-05T00:00:00+00:00",
            )
        )
        repository.upsert_opshop_pickup_schedule(
            OpShopPickupSchedule(
                schedule_id=schedule_id,
                opshop_id=opshop_id,
                run_day="MONDAY",
                run_type=run_type,
                pickup_frequency="Weekly",
                time_window="09:00-12:00",
                call_before_arrival=False,
                call_timing=None,
                status="Active",
                active_flag=True,
                fortnight_group=None,
                review_required=False,
                review_reason=None,
                created_at="2026-07-05T00:00:00+00:00",
                updated_at="2026-07-05T00:00:00+00:00",
                default_driver_id=default_driver_id,
                default_driver_name_snapshot=(
                    repository.get_driver(default_driver_id).name
                    if default_driver_id
                    else None
                ),
                pickup_category=pickup_category,
                route_group_id=None,
                regular_route_sequence=regular_route_sequence,
            )
        )
        return schedule_id, opshop_id

    def _seed_pickup(
        self,
        task_id,
        driver_id,
        assignment_dispatch_date,
        run_type="REGULAR",
        pickup_category="NORMAL",
        repository=None,
        opshop_name=None,
        suburb="MELBOURNE",
        regular_route_sequence=None,
        trip_sequence=None,
    ):
        repository = repository or self.repository
        schedule_id, opshop_id = self._seed_template(
            task_id,
            run_type,
            pickup_category,
            repository=repository,
            opshop_name=opshop_name,
            suburb=suburb,
            regular_route_sequence=regular_route_sequence,
        )
        repository.upsert_opshop_pickup_task(
            OpShopPickupTask(
                pickup_task_id=task_id,
                schedule_id=schedule_id,
                opshop_id=opshop_id,
                pickup_date=self.pickup_date,
                task_type="OPSHOP_PICKUP",
                generated_from="ON_CALL" if run_type == "ON_CALL" else "REGULAR",
                status="ASSIGNED" if driver_id else "ACTIVE",
                dispatch_date=self.pickup_date,
                driver_id=driver_id,
                trip_no="trip1" if driver_id else None,
                notes="Collection QA",
                created_at="2026-07-05T00:00:00+00:00",
                updated_at="2026-07-05T00:00:00+00:00",
                trip_sequence=trip_sequence,
            )
        )
        if assignment_dispatch_date:
            repository.upsert_assignment(
                assignment_dispatch_date,
                "OPSHOP_PICKUP",
                task_id,
                driver_id,
                "trip1",
            )
        return schedule_id

    def _post_generate(self, driver_id):
        return self.client.post(
            "/api/manual-dispatch/opshop/pickup-collections/generated",
            json={
                "dispatch_date": self.dispatch_date,
                "pickup_date": self.pickup_date,
                "driver_id": driver_id,
            },
        )

    def _get_board(self, dispatch_date=None):
        response = self.client.get(
            "/api/manual-dispatch/opshop/board",
            params={"dispatch_date": dispatch_date or self.dispatch_date},
        )
        self.assertEqual(200, response.status_code)
        return response.json()

    def _board_pickup(self, task_id, dispatch_date):
        return next(
            pickup
            for pickup in self._get_board(dispatch_date)["opshop_pickups"]
            if pickup["pickup_task_id"] == task_id
        )

    def _visible_pickups(self, board, driver_id):
        return [
            pickup
            for pickup in board["opshop_pickups"]
            if pickup["pickup_date"] == self.pickup_date
            and pickup["driver_id"] == driver_id
        ]

    @staticmethod
    def _restore_environment(name, value):
        if value is None:
            os.environ.pop(name, None)
        else:
            os.environ[name] = value


if __name__ == "__main__":
    unittest.main()
