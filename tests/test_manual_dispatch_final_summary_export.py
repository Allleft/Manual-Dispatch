from io import BytesIO
import importlib
import os
import shutil
import sqlite3
import unittest
import uuid
from pathlib import Path

from openpyxl import load_workbook

from backend.repositories.sqlite_manual_dispatch_repository import (
    SQLiteManualDispatchRepository,
)
from backend.schemas import (
    AssignTaskRequest,
    OpShopCountrysideRouteGroup,
    OpShopLocation,
    OpShopPickupSchedule,
    OpShopPickupTask,
    RegisterOperatorAccountRequest,
    SaveFinalTripSummaryRequest,
)
from backend.services.final_summary_excel_export_service import build_final_summary_excel
from backend.services.manual_dispatch_service import ManualDispatchService

try:
    from fastapi import FastAPI
    from fastapi.testclient import TestClient
except (ImportError, ModuleNotFoundError, RuntimeError):
    FastAPI = None
    TestClient = None


class ManualDispatchFinalSummaryExportTest(unittest.TestCase):
    def setUp(self):
        temp_parent = Path.cwd() / "tmp"
        temp_parent.mkdir(exist_ok=True)
        self.temp_dir = temp_parent / f"final-summary-export-test-{uuid.uuid4().hex}"
        self.temp_dir.mkdir()
        self.db_path = self.temp_dir / "manual_dispatch.sqlite3"
        self.repository = SQLiteManualDispatchRepository(self.db_path)
        self.service = ManualDispatchService(self.repository)
        self.dispatch_date = "2026-05-05"
        self.account = self.service.register_operator_account(
            RegisterOperatorAccountRequest(
                account_name="Mandy",
                password="secret123",
                confirm_password="secret123",
            )
        )

    def tearDown(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_export_workbook_contains_saved_snapshot_values(self):
        order = self.repository.get_order("ORD-001")
        order.order_no = "002848"
        self.repository.update_order(order)
        self._assign_order("ORD-001", "D001", "trip1")
        self._assign_order("ORD-003", "D001", "trip2")
        self.service.save_final_trip_summary(
            self._summary_request(
                driver_name="Snapshot Driver",
                vehicle_rego="SNAP123",
                trips=[
                    self._trip_payload("trip1", [self._order_payload("ORD-001")]),
                    self._trip_payload("trip2", [self._order_payload("ORD-003")]),
                ],
            )
        )

        workbook = self._load_export_workbook()
        values = self._flat_values(workbook)

        self.assertIn("Dispatch Date", values)
        self.assertIn(self.dispatch_date, values)
        self.assertIn("Delivery Date", values)
        self.assertIn("Driver", values)
        self.assertIn("Snapshot Driver", values)
        self.assertIn("Rego #", values)
        self.assertIn("SNAP123", values)
        self.assertIn("Saved By", values)
        self.assertIn("Mandy", values)
        self.assertIn("Trip 1", values)
        self.assertIn("Trip 2", values)
        self.assertIn("No.", values)
        self.assertIn("Customer Name", values)
        self.assertIn("Suburb", values)
        self.assertIn("Estimated Distance From Warehouse (km)", values)
        self.assertIn("Invoice #", values)
        self.assertIn("Order #", values)
        self.assertIn("002848", values)
        self.assertIn("Product Details", values)
        self.assertIn("Load", values)

    def test_export_skips_empty_trips(self):
        self._assign_order("ORD-001", "D001", "trip1")
        self.service.save_final_trip_summary(
            self._summary_request(
                trips=[
                    self._trip_payload("trip1", [self._order_payload("ORD-001")]),
                    self._trip_payload("trip2", []),
                ],
            )
        )

        values = self._flat_values(self._load_export_workbook())

        self.assertIn("Trip 1", values)
        self.assertNotIn("Trip 2", values)

    def test_export_does_not_include_generated_or_saved_timestamps(self):
        self._assign_order("ORD-001", "D001", "trip1")
        self.service.save_final_trip_summary(self._summary_request())

        values = self._flat_values(self._load_export_workbook())

        self.assertNotIn("Generated At", values)
        self.assertNotIn("Saved At", values)

    def test_export_does_not_include_password_data(self):
        self._assign_order("ORD-001", "D001", "trip1")
        self.service.save_final_trip_summary(self._summary_request())

        values = self._flat_values(self._load_export_workbook())
        normalized_values = [str(value).lower() for value in values]

        self.assertFalse(any("password" in value for value in normalized_values))
        self.assertFalse(any("password_hash" in value for value in normalized_values))
        self.assertFalse(any("password_salt" in value for value in normalized_values))
        self.assertNotIn("secret123", values)

    def test_export_uses_snapshots_after_live_records_change(self):
        self._assign_order("ORD-001", "D001", "trip1")
        self.service.save_final_trip_summary(
            self._summary_request(driver_name="Original Driver", vehicle_rego="ORIG123")
        )

        with sqlite3.connect(self.db_path) as connection:
            connection.execute(
                "UPDATE manual_orders SET company_name = ? WHERE order_id = ?",
                ("Edited Customer", "ORD-001"),
            )
            connection.execute(
                "UPDATE manual_drivers SET name = ? WHERE driver_id = ?",
                ("Edited Driver", "D001"),
            )
            connection.execute(
                "UPDATE manual_vehicles SET rego = ? WHERE vehicle_id = ?",
                ("EDIT123", "V002"),
            )
            connection.commit()

        values = self._flat_values(self._load_export_workbook())

        self.assertIn("Original Driver", values)
        self.assertIn("ORIG123", values)
        self.assertIn("Demo Customer A", values)
        self.assertNotIn("Edited Driver", values)
        self.assertNotIn("EDIT123", values)
        self.assertNotIn("Edited Customer", values)

    def test_empty_export_returns_useful_workbook(self):
        workbook_bytes = build_final_summary_excel([], self.dispatch_date)
        workbook = load_workbook(BytesIO(workbook_bytes))
        worksheet = workbook.active

        self.assertEqual("Final Summaries", worksheet.title)
        self.assertEqual(
            f"No saved Final Trip Summaries for {self.dispatch_date}",
            worksheet["A1"].value,
        )

    def test_export_handles_saved_summary_with_no_delivery_orders(self):
        self.service.save_final_trip_summary(self._summary_request(trips=[]))

        values = self._flat_values(self._load_export_workbook())

        self.assertIn("Driver", values)
        self.assertIn("John", values)
        self.assertIn("Total Pallets", values)
        self.assertIn("Total Loose Bags", values)
        self.assertIn("No Delivery Orders included.", values)
        self.assertNotIn("OPSHOP_PICKUP", values)

    def test_opshop_only_export_writes_separate_opshop_pickups_section(self):
        self._seed_opshop_pickup("TASK-OPSHOP-001")
        self.service.save_final_trip_summary(
            self._summary_request(
                trips=[],
                opshop_pickups=[self._opshop_payload("TASK-OPSHOP-001")],
            )
        )

        values = self._flat_values(self._load_export_workbook())

        self.assertIn("No Delivery Orders included.", values)
        self.assertIn("OP SHOP PICKUPS", values)
        self.assertIn("Northside Op Shop", values)
        self.assertIn("Coburg", values)
        self.assertIn("1 Sydney Road", values)
        self.assertNotIn("Trip 1", values)

    def test_countryside_export_writes_category_and_route_group_in_opshop_section(self):
        self._seed_opshop_pickup(
            "TASK-COUNTRYSIDE-001",
            run_type="ON_CALL",
            pickup_category="COUNTRYSIDE",
            route_group_id="CRG-EXPORT",
            route_group_name="Gippsland Route",
        )
        self.service.save_final_trip_summary(
            self._summary_request(
                trips=[],
                opshop_pickups=[
                    self._opshop_payload("TASK-COUNTRYSIDE-001", run_type="ON_CALL")
                ],
            )
        )

        worksheet_rows = list(self._load_export_workbook().active.iter_rows(values_only=True))
        opshop_heading_index = next(index for index, row in enumerate(worksheet_rows) if row[0] == "OP SHOP PICKUPS")
        header_row = worksheet_rows[opshop_heading_index + 1]
        data_row = worksheet_rows[opshop_heading_index + 2]

        self.assertEqual("Category", header_row[1])
        self.assertEqual("Route Group", header_row[2])
        self.assertEqual("Countryside", data_row[1])
        self.assertEqual("Gippsland Route", data_row[2])
        self.assertIn("Northside Op Shop", data_row)
        self.assertNotIn("OPSHOP_PICKUP", [value for row in worksheet_rows for value in row if value is not None])

    def test_mixed_export_keeps_opshop_section_separate_from_delivery_trip(self):
        self._assign_order("ORD-001", "D001", "trip1")
        self._seed_opshop_pickup("TASK-OPSHOP-001")
        self.service.save_final_trip_summary(
            self._summary_request(opshop_pickups=[self._opshop_payload("TASK-OPSHOP-001")])
        )

        worksheet_rows = list(self._load_export_workbook().active.iter_rows(values_only=True))
        values = [value for row in worksheet_rows for value in row if value is not None]
        trip_heading_index = next(index for index, row in enumerate(worksheet_rows) if row[0] == "Trip 1")
        opshop_heading_index = next(index for index, row in enumerate(worksheet_rows) if row[0] == "OP SHOP PICKUPS")

        self.assertLess(trip_heading_index, opshop_heading_index)
        self.assertIn("Demo Customer A", values)
        self.assertIn("Northside Op Shop", values)

    def _load_export_workbook(self):
        workbook_bytes = build_final_summary_excel(
            self.service.list_final_trip_summaries(self.dispatch_date),
            self.dispatch_date,
        )
        return load_workbook(BytesIO(workbook_bytes), data_only=True)

    def _flat_values(self, workbook):
        values = []
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                values.extend(cell for cell in row if cell is not None)
        return values

    def _assign_order(self, order_id, driver_id, trip_no):
        return self.service.assign_task(
            AssignTaskRequest(
                dispatch_date=self.dispatch_date,
                task_type="ORDER",
                task_id=order_id,
                driver_id=driver_id,
                trip_no=trip_no,
            )
        )

    def _summary_request(self, driver_name="John", vehicle_rego="XYZ888", trips=None, opshop_pickups=None):
        return SaveFinalTripSummaryRequest(
            dispatch_date=self.dispatch_date,
            delivery_date=self.dispatch_date,
            driver_id="D001",
            driver_name_snapshot=driver_name,
            vehicle_id="V002",
            vehicle_rego_snapshot=vehicle_rego,
            total_pallets=0,
            total_loose_bags=0,
            generated_at="2026-05-05T00:00:00Z",
            saved_by_account_name=self.account.account_name,
            saved_by_account_id=self.account.account_id,
            trips=trips
            if trips is not None
            else [self._trip_payload("trip1", [self._order_payload("ORD-001")])],
            opshop_pickups=opshop_pickups or [],
        )

    def _trip_payload(self, trip_no, orders):
        return {"trip_no": trip_no, "orders": orders}

    def _order_payload(self, order_id):
        order = self.repository.get_order(order_id)
        return {
            "task_type": "ORDER",
            "task_id": order_id,
            "order_id": order_id,
            "invoice_number": order.invoice_number,
            "order_no": order.order_no,
            "company_name": order.company_name,
            "suburb": order.suburb,
            "delivery_address": order.delivery_address,
            "product": "",
            "pallet_quantity": order.pallet_quantity,
            "loose_bags_quantity": order.loose_bags_quantity,
            "note": order.note,
        }

    def _seed_opshop_pickup(
        self,
        task_id,
        run_type="REGULAR",
        pickup_category="NORMAL",
        route_group_id=None,
        route_group_name=None,
    ):
        if route_group_id:
            self.repository.upsert_countryside_route_group(
                OpShopCountrysideRouteGroup(
                    route_group_id=route_group_id,
                    route_group_name=route_group_name or "Unknown Route Group",
                    status="Active",
                    active_flag=True,
                    display_order=1,
                    source_marker="TEST",
                    created_at="2026-05-05T00:00:00+00:00",
                    updated_at="2026-05-05T00:00:00+00:00",
                )
            )
        self.repository.upsert_opshop_location(
            OpShopLocation(
                opshop_id="OPSHOP-FINAL",
                name="Northside Op Shop",
                suburb="Coburg",
                street_address="1 Sydney Road",
                area_region="North",
                primary_contact="Mary",
                primary_phone="0400 700 001",
                secondary_contact=None,
                secondary_phone=None,
                access_type="Rear dock",
                key_required=True,
                trailer_restriction="Small truck only",
                status_notes="Ring first",
                is_active=True,
                created_at="2026-05-05T00:00:00+00:00",
                updated_at="2026-05-05T00:00:00+00:00",
            )
        )
        schedule_id = f"SCHED-{task_id}"
        self.repository.upsert_opshop_pickup_schedule(
            OpShopPickupSchedule(
                schedule_id=schedule_id,
                opshop_id="OPSHOP-FINAL",
                run_day="TUESDAY",
                run_type=run_type,
                pickup_frequency="Weekly",
                time_window="09:00-12:00",
                call_before_arrival=False,
                call_timing=None,
                status="Active",
                active_flag=True,
                fortnight_group=None,
                review_required=False,
                review_reason=None,
                created_at="2026-05-05T00:00:00+00:00",
                updated_at="2026-05-05T00:00:00+00:00",
                pickup_category=pickup_category,
                route_group_id=route_group_id,
            )
        )
        self.repository.upsert_opshop_pickup_task(
            OpShopPickupTask(
                pickup_task_id=task_id,
                schedule_id=schedule_id,
                opshop_id="OPSHOP-FINAL",
                pickup_date=self.dispatch_date,
                task_type="OPSHOP_PICKUP",
                generated_from=run_type,
                status="ASSIGNED",
                dispatch_date=self.dispatch_date,
                driver_id="D001",
                trip_no="trip1",
                notes="Leave at rear door",
                created_at="2026-05-05T00:00:00+00:00",
                updated_at="2026-05-05T00:00:00+00:00",
            )
        )

    def _opshop_payload(self, task_id, run_type="REGULAR"):
        return {
            "task_type": "OPSHOP_PICKUP",
            "pickup_task_id": task_id,
            "opshop_name": "Northside Op Shop",
            "suburb": "Coburg",
            "street_address": "1 Sydney Road",
            "pickup_date": self.dispatch_date,
            "run_type": run_type,
            "pickup_frequency": "Weekly",
            "time_window": "09:00-12:00",
            "primary_contact": "Mary",
            "primary_phone": "0400 700 001",
            "access_type": "Rear dock",
            "key_required": True,
            "trailer_restriction": "Small truck only",
            "task_notes": "Leave at rear door",
            "status": "ASSIGNED",
        }


@unittest.skipIf(FastAPI is None or TestClient is None, "FastAPI TestClient is not installed")
class ManualDispatchFinalSummaryExportRouteTest(unittest.TestCase):
    def setUp(self):
        temp_parent = Path.cwd() / "tmp"
        temp_parent.mkdir(exist_ok=True)
        self.temp_dir = temp_parent / f"final-summary-export-route-test-{uuid.uuid4().hex}"
        self.temp_dir.mkdir()
        self.db_path = self.temp_dir / "manual_dispatch.sqlite3"
        self.previous_db_path = os.environ.get("MANUAL_DISPATCH_DB_PATH")
        os.environ["MANUAL_DISPATCH_DB_PATH"] = str(self.db_path)

        self.repository = SQLiteManualDispatchRepository(self.db_path)
        self.service = ManualDispatchService(self.repository)
        self.account = self.service.register_operator_account(
            RegisterOperatorAccountRequest(
                account_name="Mandy",
                password="secret123",
                confirm_password="secret123",
            )
        )
        self.api_module = importlib.import_module("backend.api.manual_dispatch")
        self.original_service = self.api_module.service
        self.api_module.service = self.service

        app = FastAPI()
        app.include_router(self.api_module.router)
        self.client = TestClient(app)
        self.dispatch_date = "2026-05-05"

    def tearDown(self):
        self.api_module.service = self.original_service
        if self.previous_db_path is None:
            os.environ.pop("MANUAL_DISPATCH_DB_PATH", None)
        else:
            os.environ["MANUAL_DISPATCH_DB_PATH"] = self.previous_db_path
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_export_endpoint_returns_xlsx(self):
        self._assign_order("ORD-001", "D001", "trip1")
        self.service.save_final_trip_summary(self._summary_request())

        response = self.client.get(
            "/api/manual-dispatch/final-summaries/export-excel",
            params={"dispatch_date": self.dispatch_date},
        )

        self.assertEqual(200, response.status_code)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.headers["content-type"],
        )
        self.assertIn(
            f"final-trip-summary-{self.dispatch_date}.xlsx",
            response.headers["content-disposition"],
        )

    def _assign_order(self, order_id, driver_id, trip_no):
        return self.service.assign_task(
            AssignTaskRequest(
                dispatch_date=self.dispatch_date,
                task_type="ORDER",
                task_id=order_id,
                driver_id=driver_id,
                trip_no=trip_no,
            )
        )

    def _summary_request(self):
        order = self.repository.get_order("ORD-001")
        return SaveFinalTripSummaryRequest(
            dispatch_date=self.dispatch_date,
            delivery_date=self.dispatch_date,
            driver_id="D001",
            driver_name_snapshot="John",
            vehicle_id=None,
            vehicle_rego_snapshot="No vehicle selected",
            total_pallets=order.pallet_quantity,
            total_loose_bags=order.loose_bags_quantity,
            generated_at="2026-05-05T00:00:00Z",
            saved_by_account_name=self.account.account_name,
            saved_by_account_id=self.account.account_id,
            trips=[
                {
                    "trip_no": "trip1",
                    "orders": [
                        {
                            "task_type": "ORDER",
                            "task_id": order.order_id,
                            "order_id_snapshot": order.order_id,
                            "invoice_number_snapshot": order.invoice_number,
                            "company_name_snapshot": order.company_name,
                            "suburb_snapshot": order.suburb,
                            "delivery_address_snapshot": order.delivery_address,
                            "product_snapshot": "",
                            "pallet_quantity_snapshot": order.pallet_quantity,
                            "loose_bags_quantity_snapshot": order.loose_bags_quantity,
                            "note_snapshot": order.note,
                        }
                    ],
                }
            ],
        )


if __name__ == "__main__":
    unittest.main()
