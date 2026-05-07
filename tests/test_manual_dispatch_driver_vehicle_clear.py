from io import BytesIO
import shutil
import unittest
import uuid
from pathlib import Path

from openpyxl import load_workbook

from backend.repositories.sqlite_manual_dispatch_repository import (
    SQLiteManualDispatchRepository,
)
from backend.schemas import AssignDriverVehicleRequest, AssignTaskRequest
from backend.services.excel_export_service import (
    NO_VEHICLE_SELECTED,
    build_manual_dispatch_excel,
)
from backend.services.manual_dispatch_service import ManualDispatchService


class ManualDispatchDriverVehicleClearTest(unittest.TestCase):
    def setUp(self):
        temp_parent = Path.cwd() / "tmp"
        temp_parent.mkdir(exist_ok=True)
        self.temp_dir = temp_parent / f"driver-vehicle-clear-test-{uuid.uuid4().hex}"
        self.temp_dir.mkdir()
        self.db_path = self.temp_dir / "manual_dispatch.sqlite3"
        self.repository = SQLiteManualDispatchRepository(self.db_path)
        self.service = ManualDispatchService(self.repository)
        self.dispatch_date = "2026-05-05"

    def tearDown(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_assign_vehicle_to_driver_stores_vehicle_assignment(self):
        self._assign_vehicle("D001", "V002")

        board = self.service.get_board(self.dispatch_date)

        self.assertEqual(1, len(board.driver_vehicle_assignments))
        self.assertEqual("V002", board.driver_vehicle_assignments[0].vehicle_id)

    def test_clearing_vehicle_removes_driver_vehicle_assignment(self):
        self._assign_vehicle("D001", "V002")

        self._clear_vehicle("D001")

        board = self.service.get_board(self.dispatch_date)
        self.assertEqual([], board.driver_vehicle_assignments)

    def test_clearing_vehicle_does_not_remove_task_assignments(self):
        self._assign_order("ORD-001", "D001", "trip1")
        self._assign_vehicle("D001", "V002")

        self._clear_vehicle("D001")

        board = self.service.get_board(self.dispatch_date)
        self.assertEqual(1, len(board.assignments))
        self.assertEqual("ORD-001", board.assignments[0].task_id)
        self.assertEqual([], board.driver_vehicle_assignments)

    def test_clearing_vehicle_with_invalid_driver_id_is_rejected(self):
        with self.assertRaises(ValueError):
            self.service.assign_vehicle_to_driver(
                AssignDriverVehicleRequest(
                    dispatch_date=self.dispatch_date,
                    driver_id="D999",
                    vehicle_id=None,
                )
            )

    def test_assigning_invalid_vehicle_id_is_still_rejected(self):
        with self.assertRaises(ValueError):
            self._assign_vehicle("D001", "V999")

    def test_board_response_no_longer_includes_cleared_driver_vehicle_assignment(self):
        self._assign_vehicle("D001", "V002")
        self._assign_vehicle("D002", "V003")

        self._clear_vehicle("D001")

        board = self.service.get_board(self.dispatch_date)
        self.assertEqual(["D002"], [assignment.driver_id for assignment in board.driver_vehicle_assignments])

    def test_excel_export_uses_no_vehicle_selected_after_clearing_vehicle(self):
        self._assign_order("ORD-001", "D001", "trip1")
        self._assign_vehicle("D001", "V002")
        self._clear_vehicle("D001")

        rows = self._export_rows()
        data_rows = rows[1:]

        self.assertEqual(1, len(data_rows))
        self.assertEqual(NO_VEHICLE_SELECTED, data_rows[0][2])

    def _assign_order(self, order_id, driver_id, trip_no):
        return self.service.assign_task(
            AssignTaskRequest(
                dispatch_date=self.dispatch_date,
                task_type="ORDER",
                task_id=order_id,
                driver_id=driver_id,
                trip_no=trip_no,
            )
        )

    def _assign_vehicle(self, driver_id, vehicle_id):
        return self.service.assign_vehicle_to_driver(
            AssignDriverVehicleRequest(
                dispatch_date=self.dispatch_date,
                driver_id=driver_id,
                vehicle_id=vehicle_id,
            )
        )

    def _clear_vehicle(self, driver_id):
        return self.service.assign_vehicle_to_driver(
            AssignDriverVehicleRequest(
                dispatch_date=self.dispatch_date,
                driver_id=driver_id,
                vehicle_id="",
            )
        )

    def _export_rows(self):
        board = self.service.get_board(self.dispatch_date)
        excel_bytes = build_manual_dispatch_excel(board, self.dispatch_date)
        workbook = load_workbook(BytesIO(excel_bytes))
        worksheet = workbook.active
        return list(worksheet.iter_rows(values_only=True))


if __name__ == "__main__":
    unittest.main()
