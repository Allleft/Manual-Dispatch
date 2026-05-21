import shutil
import unittest
import uuid
from pathlib import Path

from openpyxl import Workbook

from backend.repositories.sqlite_manual_dispatch_repository import SQLiteManualDispatchRepository
from backend.schemas import CreateDriverRequest, UpdateDriverRequest
from backend.services.manual_dispatch_service import ManualDispatchService
from tools.import_oncall_opshop_pickups_to_db import (
    REQUIRED_COLUMNS,
    import_oncall_opshop_pickups_to_db,
)


class ImportOncallOpShopPickupsToDbTest(unittest.TestCase):
    def setUp(self):
        self.temp_dir = Path.cwd() / "tmp" / f"oncall-opshop-import-test-{uuid.uuid4().hex}"
        self.temp_dir.mkdir(parents=True)
        self.workbook_path = self.temp_dir / "Opshop oncall pickup.xlsx"
        self.db_path = self.temp_dir / "manual_dispatch.sqlite3"
        self.repository = SQLiteManualDispatchRepository(self.db_path)
        self.service = ManualDispatchService(self.repository)
        self.service.update_driver("D001", UpdateDriverRequest(name="John Georgiadis"))
        self.service.update_driver("D002", UpdateDriverRequest(name="Gavin Fynn"))
        self.service.update_driver("D003", UpdateDriverRequest(name="Epaminondas Tsatsoulis"))
        self.lee = self.service.create_driver(CreateDriverRequest(name="Guanlin Li"))

    def tearDown(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_import_reads_oncall_sheets_and_assigned_to_aliases(self):
        self._save_workbook(
            {
                "MON": [self._row(Op_Shop_Name="Monday Oncall", Assigned_to="John G")],
                "TUE": [self._row(Op_Shop_Name="Tuesday Oncall", Assigned_to="Gavin")],
                "WED": [self._row(Op_Shop_Name="Wednesday Oncall", Assigned_to="Nonda")],
                "THU": [self._row(Op_Shop_Name="Thursday Oncall", Assigned_to="LEE")],
                "FRI": [self._row(Op_Shop_Name="Friday Oncall", Assigned_to="Mystery")],
                "Gavin": [self._row(Op_Shop_Name="Gavin Sheet Oncall", Assigned_to="Gavin")],
            }
        )

        summary = import_oncall_opshop_pickups_to_db(self.workbook_path, self.db_path)
        schedules = self.repository.list_opshop_pickup_schedules()
        by_name = {
            self.repository.get_opshop_location(schedule.opshop_id).name: schedule
            for schedule in schedules
        }

        self.assertEqual(6, summary.rows_read)
        self.assertEqual(6, summary.rows_imported)
        self.assertEqual(6, summary.schedules_inserted)
        self.assertEqual(1, summary.unresolved_assigned_to["Mystery"])
        self.assertTrue(all(schedule.run_type == "ON_CALL" for schedule in schedules))
        self.assertEqual("MONDAY", by_name["Monday Oncall"].run_day)
        self.assertEqual("TUESDAY", by_name["Tuesday Oncall"].run_day)
        self.assertEqual("WEDNESDAY", by_name["Wednesday Oncall"].run_day)
        self.assertEqual("THURSDAY", by_name["Thursday Oncall"].run_day)
        self.assertEqual("FRIDAY", by_name["Friday Oncall"].run_day)
        self.assertIsNone(by_name["Gavin Sheet Oncall"].run_day)
        self.assertEqual("D001", by_name["Monday Oncall"].default_driver_id)
        self.assertEqual("D002", by_name["Tuesday Oncall"].default_driver_id)
        self.assertEqual("D003", by_name["Wednesday Oncall"].default_driver_id)
        self.assertEqual(self.lee.driver_id, by_name["Thursday Oncall"].default_driver_id)
        self.assertIsNone(by_name["Friday Oncall"].default_driver_id)
        self.assertEqual("Mystery", by_name["Friday Oncall"].default_driver_name_snapshot)
        self.assertEqual([], self.repository.list_opshop_pickup_tasks())

    def test_blank_active_flag_with_active_status_is_imported_and_inactive_rows_skip(self):
        self._save_workbook(
            {
                "MON": [
                    self._row(Op_Shop_Name="Blank Active Flag", Active_Flag=""),
                    self._row(Op_Shop_Name="On Hold", Status="On_Hold", Active_Flag="1"),
                    self._row(Op_Shop_Name="Inactive", Status="Active", Active_Flag="0"),
                ]
            }
        )

        summary = import_oncall_opshop_pickups_to_db(self.workbook_path, self.db_path)
        locations = self.repository.list_opshop_locations()
        schedules = self.repository.list_opshop_pickup_schedules()

        self.assertEqual(3, summary.rows_read)
        self.assertEqual(1, summary.rows_imported)
        self.assertEqual(2, summary.rows_skipped_inactive)
        self.assertEqual(["Blank Active Flag"], [location.name for location in locations])
        self.assertEqual("ON_CALL", schedules[0].run_type)
        self.assertEqual("MONDAY", schedules[0].run_day)

    def test_rerun_updates_existing_rows_without_duplicates_and_deactivates_missing_oncall(self):
        self._save_workbook(
            {
                "MON": [self._row(Op_Shop_Name="Monday Oncall", Assigned_to="John G")],
                "Gavin": [self._row(Op_Shop_Name="Gavin Sheet Oncall", Assigned_to="Gavin")],
            }
        )
        first_summary = import_oncall_opshop_pickups_to_db(self.workbook_path, self.db_path)

        self._save_workbook(
            {
                "MON": [self._row(Op_Shop_Name="Monday Oncall", Assigned_to="John G")],
            }
        )
        second_summary = import_oncall_opshop_pickups_to_db(self.workbook_path, self.db_path)
        schedules = self.repository.list_opshop_pickup_schedules()
        active_schedules = [schedule for schedule in schedules if schedule.active_flag]

        self.assertEqual(2, first_summary.schedules_inserted)
        self.assertEqual(1, second_summary.schedules_updated)
        self.assertEqual(1, second_summary.schedules_deactivated)
        self.assertEqual(2, len(schedules))
        self.assertEqual(["MONDAY"], [schedule.run_day for schedule in active_schedules])
        self.assertIsNotNone(second_summary.backup_path)
        self.assertTrue(Path(second_summary.backup_path).exists())

    def test_missing_required_sheet_or_column_fails_before_writing_database(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "MON"
        worksheet.append(["Op_Shop_Name", "Assigned to"])
        workbook.save(self.workbook_path)

        with self.assertRaises(ValueError):
            import_oncall_opshop_pickups_to_db(self.workbook_path, self.db_path)

        self.assertEqual([], self.repository.list_opshop_pickup_schedules())

    def _save_workbook(self, rows_by_sheet):
        workbook = Workbook()
        workbook.remove(workbook.active)
        for sheet_name in ["MON", "TUE", "WED", "THU", "FRI", "Gavin"]:
            worksheet = workbook.create_sheet(sheet_name)
            worksheet.append(REQUIRED_COLUMNS)
            for row in rows_by_sheet.get(sheet_name, []):
                worksheet.append([row.get(column, "") for column in REQUIRED_COLUMNS])
        workbook.save(self.workbook_path)

    def _row(self, **overrides):
        row = {
            "Op_Shop_Name": "Northside Oncall Op Shop",
            "Run_Day": "Ignored Workbook Day",
            "Run_Type": "On_Call",
            "Active_Flag": "1",
            "Suburb": "Coburg",
            "Street_Address": "1 Sydney Road",
            "Area_Region": "North",
            "Primary_Contact": "Mary",
            "Primary_Phone": "0400 000 001",
            "Secondary_Contact": "John",
            "Secondary_Phone": "0400 000 002",
            "Pickup_Frequency": "On Call",
            "Time_Window": "9-12",
            "Call_Before_Arrival": "Yes",
            "Call_Timing": "30 minutes",
            "Access_Type": "Rear dock",
            "Key_Required": "No",
            "Trailer_Restriction": "Small truck only",
            "Status": "Active",
            "Status_Start_Date": "2026-05-19",
            "Status_Notes": "Ring first",
            "Assigned to": "John G",
        }
        normalized_overrides = {
            ("Assigned to" if key == "Assigned_to" else key): value
            for key, value in overrides.items()
        }
        row.update(normalized_overrides)
        return row


if __name__ == "__main__":
    unittest.main()
